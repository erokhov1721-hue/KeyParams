import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from .passport import format_number

DEFAULT_COLUMN_WIDTH = 8.43  # Excel's own default column width, in characters.
_ALIGNMENT_MAP = {
    "left": "left",
    "right": "right",
    "center": "center",
    "centerContinuous": "center",
}
# Default Office Open XML theme colors (used when workbook has no embedded theme).
# Order per the OOXML spec's <color theme="n"/> indices: lt1, dk1, lt2, dk2,
# accent1..accent6.
_DEFAULT_THEME_COLORS = [
    "FFFFFF",  # 0: lt1 (white / "Background 1")
    "000000",  # 1: dk1 (black / "Text 1")
    "E7E6E6",  # 2: lt2 (light gray / "Background 2")
    "44546A",  # 3: dk2 (dark blue-gray / "Text 2")
    "4472C4",  # 4: accent1 (blue)
    "ED7D31",  # 5: accent2 (orange)
    "A5A5A5",  # 6: accent3 (gray)
    "FFC000",  # 7: accent4 (yellow)
    "5B9BD5",  # 8: accent5 (light blue)
    "70AD47",  # 9: accent6 (green)
]

# Hard cap on how much of a sheet gets rendered. Real-world workbooks can
# report wildly inflated max_row/max_col (e.g. a border applied to an entire
# unused column) which would otherwise blow up render time and memory.
MAX_RENDERED_ROWS = 2000
MAX_RENDERED_COLS = 200


class EstimateReadError(Exception):
    pass


# Why an upload was refused, in words — looked up by code the same way
# ``cost_increase.PROBLEM_MESSAGES`` is, so an arbitrary ?estimate=... in the
# address bar puts nothing on the page.
PROBLEM_MESSAGES = {
    "format": "Смета должна быть в формате .xlsx",
    "unreadable": (
        "Не удалось прочитать смету — убедитесь, что это корректный файл .xlsx. "
        "Прежняя смета оставлена на месте."
    ),
}


def read_estimate(path) -> list:
    """Parse an .xlsx workbook into a list of rendered sheets.

    Each sheet is ``{"name": str, "rows": [[cell, ...], ...], "col_widths":
    [float, ...]}``. Hidden and very-hidden sheets are skipped. A cell
    covered by a merge (other than the merge's top-left cell) is omitted
    from its row — the top-left cell instead carries the merge's rowspan
    and colspan.
    """
    path = Path(path)
    try:
        # Read into memory once and open two independent BytesIO views on
        # it, rather than handing openpyxl the path directly: a workbook
        # that fails to load partway through (a zip missing its manifest,
        # say) leaves openpyxl's own archive reader without ever closing
        # it, and on Windows that keeps the file locked — the caller's own
        # cleanup of a rejected upload then fails with "file in use" on
        # top of the read error it was already reporting. A BytesIO has no
        # OS file handle to leak, so there is nothing left open either way.
        data = path.read_bytes()
        wb_styles = openpyxl.load_workbook(BytesIO(data), data_only=False)
        wb_values = openpyxl.load_workbook(BytesIO(data), data_only=True)
    except (zipfile.BadZipFile, KeyError, InvalidFileException, ValueError, OSError) as e:
        raise EstimateReadError(f"Cannot read {path}: {e}") from e

    value_bounds = _value_bounds(path)

    sheets = []
    for ws_styles in wb_styles.worksheets:
        if ws_styles.sheet_state != "visible":
            continue
        bounds = value_bounds.get(ws_styles.title, (0, 0))
        sheets.append(_render_sheet(ws_styles, wb_values[ws_styles.title], wb_styles, bounds))
    return sheets


def _value_bounds(path):
    """Scan every sheet with a read_only (streaming) workbook load and return
    ``{sheet_title: (max_row, max_col)}`` among cells whose value is not
    None.

    A sheet's declared dimension (``ws.max_row``/``ws.max_column`` on a
    normal-mode worksheet) can be wildly inflated by a single stray
    formatting artifact far outside the real data — e.g. a border applied to
    column XFD. Fetching cells at that scale via normal-mode
    ``ws.cell(r, c)`` is prohibitively slow, because each such call
    materializes a real ``Cell`` object. ``read_only=True`` streams the
    underlying XML and only visits cells that actually exist, so it finds the
    true content boundary cheaply regardless of how inflated the declared
    dimension is.

    IMPORTANT: this must stay read_only. Calling ``iter_rows()`` with no
    bounds on a *normal*-mode worksheet under the same inflated-dimension
    condition is catastrophic (it eagerly materializes the full declared
    rectangle) and can hang for minutes.

    Uses ``data_only=False`` (not True): a formula cell with no cached
    result (e.g. a workbook openpyxl itself just wrote, never opened in
    Excel) reads as ``None`` under ``data_only=True`` even though it is
    genuine content — that would wrongly shrink the boundary and drop real
    rows/columns. With ``data_only=False`` a formula cell's ``.value`` is
    its formula text, which is never ``None``, so it still counts.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        bounds = {}
        for ws in wb.worksheets:
            max_row = 0
            max_col = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        if cell.row > max_row:
                            max_row = cell.row
                        if cell.column > max_col:
                            max_col = cell.column
            bounds[ws.title] = (max_row, max_col)
        return bounds
    finally:
        wb.close()


def _render_sheet(ws_styles, ws_values, wb, value_bounds=(0, 0)):
    value_max_row, value_max_col = value_bounds
    span, covered = _merge_spans(ws_styles)
    merge_max_row = 0
    merge_max_col = 0
    for (top_row, top_col), (rowspan, colspan) in span.items():
        merge_max_row = max(merge_max_row, top_row + rowspan - 1)
        merge_max_col = max(merge_max_col, top_col + colspan - 1)

    content_max_row = max(value_max_row, merge_max_row)
    content_max_col = max(value_max_col, merge_max_col)

    max_row = min(content_max_row, MAX_RENDERED_ROWS)
    max_col = min(content_max_col, MAX_RENDERED_COLS)
    truncated = content_max_row > max_row or content_max_col > max_col

    rows = []
    for r in range(1, max_row + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            if (r, c) in covered:
                continue
            rowspan, colspan = span.get((r, c), (1, 1))
            row_cells.append(_render_cell(
                ws_styles.cell(row=r, column=c),
                ws_values.cell(row=r, column=c),
                rowspan, colspan,
                wb,
            ))
        rows.append(row_cells)

    col_widths = []
    for c in range(1, max_col + 1):
        dim = ws_styles.column_dimensions.get(get_column_letter(c))
        col_widths.append(dim.width if dim and dim.width else DEFAULT_COLUMN_WIDTH)

    sheet = {"name": ws_styles.title, "rows": rows, "col_widths": col_widths}
    if truncated:
        sheet["truncated"] = True
    return sheet


def _merge_spans(ws):
    """(span, covered): ``span`` maps a merge's top-left ``(row, col)`` to
    ``(rowspan, colspan)``; ``covered`` holds every other ``(row, col)``
    inside a merged range, which must be skipped when laying out a row."""
    span = {}
    covered = set()
    for merged_range in ws.merged_cells.ranges:
        top_left = (merged_range.min_row, merged_range.min_col)
        span[top_left] = (
            merged_range.max_row - merged_range.min_row + 1,
            merged_range.max_col - merged_range.min_col + 1,
        )
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                if (r, c) != top_left:
                    covered.add((r, c))
    return span, covered


def _render_cell(cell_style, cell_value, rowspan, colspan, wb):
    border = cell_style.border
    return {
        "value": _format_value(cell_value.value),
        "rowspan": rowspan,
        "colspan": colspan,
        "bold": bool(cell_style.font and cell_style.font.bold),
        "italic": bool(cell_style.font and cell_style.font.italic),
        "align": _alignment(cell_style.alignment),
        "bg": _fill_color(cell_style.fill, wb),
        "border_top": _has_border(border.top),
        "border_right": _has_border(border.right),
        "border_bottom": _has_border(border.bottom),
        "border_left": _has_border(border.left),
    }


def _format_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, (int, float)):
        return format_number(value)
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    return str(value)


def _alignment(alignment):
    if alignment is None or not alignment.horizontal:
        return None
    return _ALIGNMENT_MAP.get(alignment.horizontal)


def _fill_color(fill, wb):
    if fill is None or fill.fill_type != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None

    # Handle RGB colors directly
    if fg.type == "rgb":
        try:
            rgb = fg.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                if len(rgb) == 8:  # AARRGGBB -> drop the alpha channel
                    rgb = rgb[2:]
                return f"#{rgb}"
        except (AttributeError, ValueError):
            pass

    # Handle indexed colors from the standard palette
    if fg.type == "indexed":
        try:
            idx = fg.indexed
            if idx is not None and 0 <= idx < len(COLOR_INDEX):
                indexed_color = COLOR_INDEX[idx]
                if isinstance(indexed_color, str) and len(indexed_color) >= 6:
                    if len(indexed_color) == 8:
                        indexed_color = indexed_color[2:]
                    return f"#{indexed_color}"
        except (AttributeError, IndexError, TypeError):
            pass

    # Handle theme colors using the standard Office default theme colors.
    # Note: This uses an approximation (standard Office theme palette) rather than
    # attempting to parse the workbook's embedded theme XML. This is a known
    # limitation that provides reasonable color approximations for most estimates
    # while keeping the code simple and avoiding brittle XML parsing.
    if fg.type == "theme":
        try:
            theme_idx = fg.theme
            if theme_idx is not None and 0 <= theme_idx < len(_DEFAULT_THEME_COLORS):
                rgb = _DEFAULT_THEME_COLORS[theme_idx]
                if len(rgb) >= 6:
                    return f"#{rgb}"
        except (AttributeError, IndexError, TypeError):
            pass  # Theme resolution failed, return None

    return None


def _has_border(side):
    return bool(side and side.style)
