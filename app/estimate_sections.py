"""Reading an estimate's top-level sections into the report's cost lines.

A tender offer is a few thousand rows deep, but the comparison only wants
seventeen numbers: the total of each top-level section, laid against the line
it belongs to. The offer carries its own hierarchy in a "№ раздела" column —
1, 1.1, 1.2.3 — so the sections are the rows numbered with a bare integer,
and everything under them is already summed into them.

Which line a section belongs to is decided by its name rather than by its
number: the numbering is per-workbook and drifts (an offer whose sections run
1..15 numbers its own articles 1, 2, ... 12, 13, 99, 14, 16), while the names
are the industry's and stay put.
"""

import logging
import re
from collections import namedtuple
from pathlib import Path

import openpyxl

from . import extractors

logger = logging.getLogger(__name__)

# The report's cost lines, in the order they appear on the sheet, each with
# the words that identify its section in an estimate. First match wins, so
# order matters where names overlap: "Общестроительные работы (без отделки)"
# has to be claimed by the partitions line before the finishing line can see
# the word "отдел" in it.
CATEGORY_RULES = [
    ("rd", ("рабочей документации", "рабочая документация", "стадии \"р\"",
            "стадии р", "проектирование")),
    ("preparation", ("подготовительные работы", "содержание площадки")),
    ("excavation", ("котлован",)),
    ("waterproofing", ("гидроизоляц",)),
    ("concrete", ("конструктивные решения", "монолит", "несущих конструкций",
                  "конструкций здания")),
    ("partitions", ("общестроительные", "перегородк")),
    ("facade", ("фасад",)),
    ("roof", ("кровля", "кровли")),
    # Not a bare "отделка": the flats' own finishing is a package of its own
    # ("отделка квартир" = MR Base) and must not be swept in with the finishing
    # of the common areas.
    ("finishing", ("отделочные работы", "отделка моп", "отделка паркинга",
                   "отделка мест общего", "внутреняя отделка", "внутренняя отделка")),
    ("lifts", ("лифт",)),
    ("utilities", ("инженерн", "вис")),
    ("landscaping", ("благоустройств",)),
    ("technology", ("технологическ", "тх")),
    ("other", ("прочее", "зип", "другие", "дополнительные работы")),
    ("mr_base", ("mr-base", "mr base", "отделка квартир", "квартир")),
    ("mr_ready", ("mr-ready", "mr ready")),
    ("shell_core", ("shell", "нулевого цикла", "нулевой цикл")),
]

CATEGORY_KEYS = [key for key, _ in CATEGORY_RULES]

# What each line is called where a person reads it — the customer's own
# wording, from their comparison sheet. Kept here rather than beside either
# of the things that display it, so the Excel export and the page in the
# browser cannot end up naming the same line differently.
CATEGORY_LABELS = {
    "rd": 'Разработка стадии "Р"',
    "preparation": (
        "Подготовительные работы и содержание площадки (включая содержание "
        "прилегающей территории, аренда оборудования и механизмов и т.п.)"
    ),
    "excavation": "Устройство котлована",
    "waterproofing": "Гидроизоляция подземной части",
    "concrete": "Монолит + МК",
    "partitions": "Перегородки и стены",
    "facade": "Фасад",
    "roof": "Кровли",
    "finishing": "Отделка МОП, двери, ворота",
    "lifts": "Лифты",
    "utilities": "Инженерные системы",
    "landscaping": "Благоустройство",
    "technology": "Технологические решения",
    "other": "Другие (ЗИП и т.д.)",
    "mr_base": "MR Base",
    "mr_ready": "MR Ready",
    "shell_core": "SHELL & CORE",
}

# The last three lines are the finishing packages, which the sheet sets apart
# from the fourteen numbered kinds of work.
MR_CATEGORY_KEYS = ["mr_base", "mr_ready", "shell_core"]
WORK_CATEGORY_KEYS = [key for key in CATEGORY_KEYS if key not in MR_CATEGORY_KEYS]

# How far down to look for the header: an offer opens with the tender's name,
# the object, the address and the bidder's details before the table starts.
HEADER_SEARCH_ROWS = 40
HEADER_SEARCH_COLS = 40

SECTION_HEADER = "раздел"
# The stem, not the word: one offer heads the column "Статья СМР" and another
# "Справочник статей СМР", and matching the nominative alone left the second
# without an article column and so without any sections at all.
ARTICLE_HEADER = "стат"
ITEM_HEADER = "п/п"
NAME_HEADER = "наименование"
TOTAL_HEADER = "стоимость всего"
TOTAL_SUBHEADER = "всего"

# A section number, as opposed to a sub-section: "4" is one, "4.1" is part of
# it and is already counted inside it.
TOP_LEVEL_RE = re.compile(r"^\d+$")
# The article number an estimate prefixes its section name with ("12.
# Благоустройство"), dropped before matching so a renumbered offer still
# lands on the same line.
LEADING_NUMBER_RE = re.compile(r"^[\d.]+\s*")


class EstimateSectionsError(Exception):
    pass


_LETTER_RE = re.compile(r"[a-zа-яё]", re.IGNORECASE)


def _named(value):
    """The value as a name, or None if it is only a number or a code."""
    text = str(value or "").strip()
    return text if _LETTER_RE.search(text) else None


def _cell_text(ws, row, col):
    value = ws.cell(row=row, column=col).value
    return str(value).strip().lower() if value is not None else ""


def _heading_block(ws, row, col):
    """The columns a heading covers: up to just before the next heading.

    Read from the headings themselves rather than from the merged cells they
    are usually written in — one offer left "Стоимость всего" unmerged, and a
    reader that insisted on the merge found no totals column at all and gave
    up on the whole workbook.
    """
    for following in range(col + 1, HEADER_SEARCH_COLS + 1):
        if _cell_text(ws, row, following):
            return col, following - 1
    return col, HEADER_SEARCH_COLS


def _find_total_column(ws, header_row):
    """The column holding each section's total cost.

    An offer states costs twice over — per unit and in total — under headings
    with the same four sub-columns beneath each, so the heading alone doesn't
    identify a column: the one wanted is the "Всего" under "Стоимость всего".
    "Стоимость всего за объёмы заказчика" stands alongside with nothing
    underneath it, and is passed over for exactly that reason.
    """
    for col in range(1, HEADER_SEARCH_COLS + 1):
        if TOTAL_HEADER not in _cell_text(ws, header_row, col):
            continue
        first, last = _heading_block(ws, header_row, col)
        for sub in range(first, last + 1):
            if _cell_text(ws, header_row + 1, sub) == TOTAL_SUBHEADER:
                return sub
    return None


Header = namedtuple(
    "Header", "row item_col section_col article_col name_col total_col",
)

# One section of an estimate, already placed on a report line: ``key`` is the
# line, ``name`` the section as the estimate spells it. The name is carried
# along so the report can say where a figure came from — the placement is a
# judgement about wording, and a judgement nobody can check is worth little.
Section = namedtuple("Section", "key name amount")


def _find_header(ws):
    """Where the table starts and which columns matter, or None if this
    workbook isn't an offer laid out with numbered sections.

    The name column is found alongside the article column because a section is
    not always named in the one meant for it — in a real offer one section of
    fifteen had its "Статья СМР" cell left blank and only the works name
    filled in, and dropping it would have quietly lost 1.9 billion roubles
    from the report.
    """
    for row in range(1, HEADER_SEARCH_ROWS + 1):
        item_col = section_col = article_col = name_col = None
        for col in range(1, HEADER_SEARCH_COLS + 1):
            text = _cell_text(ws, row, col)
            if item_col is None and ITEM_HEADER in text:
                item_col = col
            elif section_col is None and SECTION_HEADER in text:
                section_col = col
            elif article_col is None and ARTICLE_HEADER in text:
                article_col = col
            elif name_col is None and NAME_HEADER in text:
                name_col = col
        if section_col is None or article_col is None:
            continue
        total_col = _find_total_column(ws, row)
        if total_col is not None:
            return Header(row, item_col, section_col, article_col, name_col, total_col)
    return None


def _amount(value):
    """The number in a total cell. A zero is a number like any other: written
    as ``value or ""`` this quietly dropped every section that came to nothing,
    and a section costing nothing is exactly what the report wants to show."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return extractors.parse_number(str(value))


def classify(name: str):
    """The report line a section with this name belongs to, or None."""
    text = LEADING_NUMBER_RE.sub("", str(name or "").strip().lower())
    if not text:
        return None
    for key, tokens in CATEGORY_RULES:
        for token in tokens:
            # Short tokens are matched as whole words: "ТХ" is a section of
            # its own, but as a substring it also sits inside "отходов".
            if len(token) <= 3:
                if re.search(rf"\b{re.escape(token)}\b", text):
                    return key
            elif token in text:
                return key
    return None


def read_section_totals(path) -> dict:
    """``{category key: total cost}`` for one estimate.

    Only the lines the estimate actually has appear in the result — a section
    it doesn't carry stays absent rather than becoming a zero, so the report
    can tell "nothing was spent here" from "this estimate doesn't say".
    """
    totals = {}
    for section in read_sections(path):
        totals[section.key] = totals.get(section.key, 0.0) + section.amount
    return totals


def read_sections(path) -> list:
    """Every section of one estimate that has a line in the report.

    Returns an empty list for a workbook that isn't laid out as a sectioned
    offer at all: the report then leaves its cost lines blank, exactly as it
    did before any estimate was attached.
    """
    path = Path(path)
    try:
        # Not read_only: the header spans merged cells, and a streaming
        # worksheet doesn't carry the merges needed to tell the two
        # cost blocks apart. Only three cells per row are read afterwards,
        # so the whole-workbook load is the cheaper half of this anyway.
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise EstimateSectionsError(f"Cannot read {path}: {e}") from e

    for ws in wb.worksheets:
        sections = _sections_from_sheet(ws)
        if sections:
            return sections
    return []


LEVEL_HEADER = "уровень"
LEVEL_TOTAL_HEADER = "всего"

Levels = namedtuple("Levels", "row first second third total_col")


def _find_levels_header(ws):
    """Where a "укрупнённая смета" states its levels, or None.

    A different shape of estimate from the offer above: instead of one column
    of section numbers it gives a column pair per level of nesting — "номер 1"
    and "уровень 1", "номер 2" and "уровень 2", and so on — with the money in
    a "Всего" column of its own.
    """
    for row in range(1, HEADER_SEARCH_ROWS + 1):
        levels, total_col = [], None
        for col in range(1, HEADER_SEARCH_COLS + 1):
            text = _cell_text(ws, row, col)
            if text.startswith(LEVEL_HEADER):
                levels.append(col)
            elif total_col is None and text.startswith(LEVEL_TOTAL_HEADER):
                total_col = col
        if len(levels) >= 2 and total_col is not None:
            return Levels(
                row=row, first=levels[0], second=levels[1],
                third=levels[2] if len(levels) > 2 else None,
                total_col=total_col,
            )
    return None


def _sections_from_levels(ws):
    """Section totals from an estimate written in levels.

    The money sits at the deepest level that has any: a section's own row is
    usually empty, and so is the row of a sub-section whose parts are listed
    beneath it. Each sub-section is therefore taken at its own figure where it
    has one, and as the sum of its parts where it hasn't — which is the one
    reading that neither loses money nor counts it twice.
    """
    header = _find_levels_header(ws)
    if header is None:
        return []

    sections, unmatched = [], []
    current_key = current_name = None
    current_own = None      # сумма, записанная на самой строке раздела
    total = 0.0
    pending = None          # (собственная сумма подраздела, сумма его частей)

    def close_subsection():
        nonlocal total, pending
        if pending is not None:
            own, parts = pending
            total += own if own is not None else parts
            pending = None

    def close_section():
        nonlocal current_key, current_own, total
        close_subsection()
        if current_key is not None:
            # A section that states its own total is taken at its word; one
            # that leaves the cell empty is the sum of what is listed under it.
            sections.append(Section(
                current_key, current_name,
                current_own if current_own is not None else total,
            ))
        current_key, current_own, total = None, None, 0.0

    for row in range(header.row + 1, ws.max_row + 1):
        amount = _amount(ws.cell(row=row, column=header.total_col).value)
        first = _named(ws.cell(row=row, column=header.first).value)
        second = _named(ws.cell(row=row, column=header.second).value)
        third = (
            _named(ws.cell(row=row, column=header.third).value)
            if header.third else None
        )

        if first:
            close_section()
            key = classify(first)
            if key is None:
                unmatched.append(first)
                current_name = None
            else:
                current_key, current_name, current_own = key, first, amount
            continue
        if current_key is None:
            continue
        if second:
            close_subsection()
            pending = (amount, 0.0)
        elif pending is not None and amount is not None:
            own, parts = pending
            pending = (own, parts + amount)
        elif amount is not None and third:
            total += amount

    close_section()
    _report_unmatched(unmatched)
    return sections


def _report_unmatched(unmatched):
    if unmatched:
        # Not an error: an estimate may carry sections this report has no line
        # for. Logged so that a section quietly going missing can be traced.
        logger.info("Разделы сметы без строки в отчёте: %s", "; ".join(unmatched))


def _sections_from_sheet(ws):
    sections = _sections_from_offer(ws)
    return sections if sections else _sections_from_levels(ws)


def _sections_from_offer(ws):
    header = _find_header(ws)
    if header is None:
        return []

    sections = []
    unmatched = []
    started = False
    for row in range(header.row + 2, ws.max_row + 1):
        number = ws.cell(row=row, column=header.section_col).value
        item = ws.cell(row=row, column=header.item_col).value if header.item_col else None
        article = ws.cell(row=row, column=header.article_col).value

        is_section = number is not None and TOP_LEVEL_RE.match(str(number).strip())
        # A row carrying neither a line number nor a section number is not
        # part of the priced work: it is something added underneath it. In the
        # offer this was written against, "Дополнительные работы" sits there
        # with 12.7 million against it, and skipping it would have left the
        # report 12.7 million short of the offer's own bottom line. The
        # "ИТОГО" rows below look the same but name themselves in the margin
        # rather than in a column this reads, so they fall out on their own.
        is_extra = started and number is None and item is None
        if not (is_section or is_extra):
            continue

        if not started:
            # An offer opens with a row for the lot itself: a section number,
            # no article, and the whole offer as its total. Sections start at
            # the first row that names an article, and anything above it is
            # that header — which must not be counted, or the report doubles.
            if not article:
                continue
            started = True

        # The column meant for the article name doesn't always hold one: one
        # offer numbers its articles there ("1", "1.1") and puts the wording
        # in the works column instead. A name has to have letters in it.
        name = _named(article) or _named(
            ws.cell(row=row, column=header.name_col).value if header.name_col else None
        )
        if not name:
            continue
        amount = _amount(ws.cell(row=row, column=header.total_col).value)
        if amount is None:
            continue
        key = classify(name)
        if key is None:
            unmatched.append(str(name).strip())
            continue
        sections.append(Section(key, str(name).strip(), amount))

    _report_unmatched(unmatched)
    return sections
