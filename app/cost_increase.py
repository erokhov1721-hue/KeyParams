"""Reading a project's cost-increase workbook into a percentage per kind of work.

The estimate says what the work was priced at; the cost-increase workbook says
what it costs now. Its "стало" column is that figure — always the latest, since
the workbook is kept cumulatively and the newest version already carries every
change agreed so far. Nothing is ever added up across versions: one file per
project, and a newer one replaces it outright.

So the increase is "стало" against the estimate, kind of work by kind of work.
"было" is not the baseline — the estimate is. "было" is only reached for where
"стало" says nothing at all, which is what an unfilled "стало" means: that row
hasn't been restated, and the last figure the workbook has for it is the one in
"было".

Where no estimate is attached to the project there is no baseline outside the
workbook, and the increase falls back to being "стало" against "было" — the only
comparison the file can make on its own. The page says so rather than passing it
off as a comparison with the estimate.

Rows are named the way the estimate names its sections, so which report line a
row belongs to is decided by ``estimate_sections.classify`` — the same judgement
the estimate reader makes, which is what lets a figure here and a figure there be
about the same kind of work. The one place the wording differs is the engineering
systems: the estimate carries them as a single section, this workbook splits them
in two ("ВИС - механические системы", "ВИС - Электрические и слаботочные
системы"), and both land on the one line and are added together.
"""

import logging
import re
from collections import namedtuple

import openpyxl

from . import estimate_sections, extractors
from .passport import format_number

logger = logging.getLogger(__name__)

# How far into the sheet to look for the header. The workbook opens straight
# into its table, but a title row or two above it costs nothing to allow for.
HEADER_SEARCH_ROWS = 20
HEADER_SEARCH_COLS = 40

WAS_HEADER = "было"
NOW_HEADER = "стало"

# The workbook's own bottom line. It is read past rather than through: the total
# shown under the table is the sum of the lines the table actually shows, so a
# row this reader dropped can be seen in the total rather than hidden by it.
TOTAL_ROW_RE = re.compile(r"^(итого|всего|сумма|total)\b")

_LETTER_RE = re.compile(r"[a-zа-яё]", re.IGNORECASE)


class CostIncreaseError(Exception):
    """The file isn't a cost-increase workbook this reader can make sense of."""


class FormulaWithoutCacheError(CostIncreaseError):
    """A "было"/"стало" cell holds a formula Excel never saved a computed
    result for. Read with ``data_only=True`` that cell is ``None`` — the
    same as one with nothing in it — and the two must not be treated alike:
    a line where both columns are genuinely blank is quietly skipped, which
    is right; a line where a formula just never got a cached value would be
    skipped the same way, silently dropping it from the increase report, or
    (if only one of the pair is affected) turning a missing figure into a
    false 100% swing once ``was or 0.0`` / ``now or 0.0`` fills it in.
    """


class _CheckedSheet:
    """A worksheet paired with its own formula-only twin (the same file
    loaded a second time with ``data_only=False``), so ``amount_at`` can
    tell "genuinely empty" apart from "a formula with no cached value" —
    everything else about the sheet passes straight through unchanged.
    """

    def __init__(self, ws, ws_formulas):
        self._ws = ws
        self._ws_formulas = ws_formulas

    def __getattr__(self, name):
        return getattr(self._ws, name)

    def amount_at(self, row, column):
        value = self._ws.cell(row=row, column=column).value
        if value is not None:
            return _amount(value)
        raw = self._ws_formulas.cell(row=row, column=column).value
        if isinstance(raw, str) and raw.startswith("="):
            coord = self._ws.cell(row=row, column=column).coordinate
            raise FormulaWithoutCacheError(
                f'Ячейка {coord} на листе «{self._ws.title}» — формула без '
                'сохранённого значения. Откройте файл в Excel, дайте ему '
                'пересчитаться (или просто сохраните заново) и загрузите его '
                'ещё раз.'
            )
        return None


# Why an upload was refused, in words. Looked up by code rather than passed as
# text, so an arbitrary ?increase=... in the address bar puts nothing on the page.
PROBLEM_MESSAGES = {
    "format": "Файл удорожания должен быть в формате .xlsx",
    "too_big": "Файл удорожания слишком большой — до 15 МБ",
    "unreadable": (
        "Не удалось прочитать файл удорожания. Нужна таблица, где у каждого "
        "вида работ есть столбцы «было» и «стало». Прежний файл оставлен на месте."
    ),
}


# One row of the workbook as it is written there.
Line = namedtuple("Line", "name was now")

# Where a line's current cost was taken from. "стало" is the answer almost
# always; the other two are the cases worth marking on the page, because there
# the figure is not what the workbook was asked for.
FROM_NOW = "now"        # столбец «стало» — как и задумано
FROM_WAS = "was"        # «стало» пусто, взято «было»
FROM_NOTHING = "none"   # файл про эти работы не говорит вовсе

# One line of the report.
#
# ``key`` and ``label`` name the kind of work. ``estimate`` is what the project's
# estimate prices it at, or None if there is no estimate or it carries no such
# section. ``was`` and ``now`` are the workbook's two columns as written, kept so
# the page can show where a figure came from and a test can pin it.
#
# ``baseline`` is what the increase is measured from — the estimate wherever
# there is one, "было" otherwise — and ``current`` what the work costs now, taken
# from ``source``. ``delta`` and ``percent`` are the increase; ``percent`` is None
# where the baseline is nothing, since the share by which nothing grew is not a
# number. ``sources`` are the workbook's own row names added together into this
# line, carried so the placement can be checked rather than taken on trust.
Row = namedtuple(
    "Row",
    "key label sources was now estimate baseline current delta percent source",
)

# ``rows`` in the report's own order, ``total`` their sum as a row of the same
# shape, ``unmatched`` the workbook's row names that belong to no report line,
# ``from_estimate`` whether the estimate supplied the baseline or the workbook's
# own "было" had to.
Report = namedtuple("Report", "rows total unmatched from_estimate")


def _text(ws, row, col):
    value = ws.cell(row=row, column=col).value
    return str(value).strip().lower() if value is not None else ""


def _named(value):
    """The value as a name, or None if it holds no letters — a row number, a
    ledger code, an empty cell."""
    text = str(value or "").strip()
    return text if _LETTER_RE.search(text) else None


def _amount(value):
    """The number in a money cell, or None if the cell holds no number.

    A zero is a number like any other: a kind of work that wasn't in the
    estimate at all starts at zero, and that zero is the whole reason its
    increase can't be stated as a percentage.
    """
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return extractors.parse_number(str(value))


Header = namedtuple("Header", "row name_col was_col now_col")


def _find_columns(ws, row):
    """The "было" and "стало" columns of this row, or None if it isn't the
    header. Matched as substrings, so "Было, руб." is still the column."""
    was_col = now_col = None
    for col in range(1, HEADER_SEARCH_COLS + 1):
        text = _text(ws, row, col)
        if was_col is None and WAS_HEADER in text:
            was_col = col
        elif now_col is None and NOW_HEADER in text:
            now_col = col
    if was_col is None or now_col is None or was_col == now_col:
        return None
    return was_col, now_col


def _name_column(ws, header_row, was_col, now_col):
    """The column the kinds of work are named in, or None if none is.

    Looked for to the left of the money first, which is where a table of this
    shape puts it — and where, in the real workbook, it sits with no heading of
    its own to be found by. The whole sheet is only searched if there is nothing
    on the left at all, so a wide comment column further right can't win a
    column that has the names.
    """
    def count(col):
        return sum(
            1 for row in range(header_row + 1, ws.max_row + 1)
            if _named(ws.cell(row=row, column=col).value)
        )

    for candidates in (
        range(was_col - 1, 0, -1),
        (c for c in range(1, HEADER_SEARCH_COLS + 1) if c not in (was_col, now_col)),
    ):
        counts = [(count(col), col) for col in candidates]
        best = max(counts, default=(0, None), key=lambda pair: pair[0])
        if best[0]:
            return best[1]
    return None


def _find_header(ws):
    for row in range(1, HEADER_SEARCH_ROWS + 1):
        columns = _find_columns(ws, row)
        if columns is None:
            continue
        was_col, now_col = columns
        name_col = _name_column(ws, row, was_col, now_col)
        if name_col is not None:
            return Header(row, name_col, was_col, now_col)
    return None


def read_lines(source) -> list:
    """Every priced row of a cost-increase workbook, as it is written there.

    ``source`` is a path or an open file — the upload is checked before it is
    saved, so a file that turns out to be unreadable can be refused without
    having overwritten the one that was already there.
    """
    try:
        # Not read_only: a streaming worksheet has no addressable cells, and
        # this reader looks columns up by number rather than walking rows in
        # order. The workbook is a page and a half long anyway.
        wb = openpyxl.load_workbook(source, data_only=True)
        # ``source`` is sometimes a path (reopening it is free) and sometimes
        # an already-open stream (a BytesIO the upload was buffered into) —
        # the first load consumed it, so it has to be rewound before the
        # second one can read it again.
        if hasattr(source, "seek"):
            source.seek(0)
        wb_formulas = openpyxl.load_workbook(source, data_only=False)
    except Exception as e:
        raise CostIncreaseError(f"файл не читается как .xlsx: {e}") from e

    for ws, ws_formulas in zip(wb.worksheets, wb_formulas.worksheets):
        header = _find_header(ws)
        if header is None:
            continue
        lines = _lines_from_sheet(_CheckedSheet(ws, ws_formulas), header)
        if lines:
            return lines

    raise CostIncreaseError(
        'в файле нет таблицы со столбцами «было» и «стало»'
    )


def _lines_from_sheet(ws, header) -> list:
    lines = []
    for row in range(header.row + 1, ws.max_row + 1):
        name = _named(ws.cell(row=row, column=header.name_col).value)
        if not name or TOTAL_ROW_RE.match(name.lower()):
            continue
        was = ws.amount_at(row, header.was_col)
        now = ws.amount_at(row, header.now_col)
        if was is None and now is None:
            continue
        lines.append(Line(name, was or 0.0, now or 0.0))
    return lines


def _percent(baseline, current):
    """The increase as a percentage, or None where there isn't one.

    Work the baseline priced at nothing grew from nothing, and the share by which
    nothing grew is not a number. Saying "+100%" there, or quietly writing a 0,
    would both read as a fact about the work rather than about the arithmetic.
    """
    if not baseline:
        return None if current else 0.0
    return (current / baseline - 1.0) * 100.0


def _current(was, now, baseline):
    """``(current cost, where it came from)`` for one line.

    "стало" is the answer wherever it has one — it is the whole point of the
    column, and of the file being kept cumulatively. An empty "стало" is not a
    zero cost, though: it means this row hasn't been restated, so the last figure
    the workbook holds for it is "было". And where neither column says anything,
    the workbook is silent about this work, which is not the same as the work
    having been struck out — the baseline stands unchanged rather than the line
    reading as the whole of it saved.
    """
    if now:
        return now, FROM_NOW
    if was:
        return was, FROM_WAS
    return baseline, FROM_NOTHING


def build_report(lines, estimate_totals=None) -> Report:
    """The workbook's rows measured against the estimate, line by line.

    Lines are returned in the report's own order rather than the workbook's, so
    the kinds of work read down the accordion in the order they read down the
    comparison.

    ``estimate_totals`` is ``{line: cost}`` from the project's own estimate, as
    ``estimate_sections.read_section_totals`` returns it, and it is what the
    increase is measured from. Every kind of work either of them names gets a
    line: a section the workbook leaves out still belongs in a table that claims
    to show the estimate's increase, and shows as unchanged with the workbook's
    silence marked on it.

    Without an estimate the only baseline left is the workbook's own "было", and
    the report says as much through ``from_estimate`` rather than presenting the
    two as the same thing.
    """
    estimate_totals = estimate_totals or {}
    from_estimate = bool(estimate_totals)

    gathered, unmatched = {}, []
    for line in lines:
        key = estimate_sections.classify(line.name)
        if key is None:
            unmatched.append(line.name)
            continue
        was, now, sources = gathered.get(key, (0.0, 0.0, []))
        gathered[key] = (was + line.was, now + line.now, sources + [line.name])

    rows = []
    for key in estimate_sections.CATEGORY_KEYS:
        if key not in gathered and key not in estimate_totals:
            continue
        was, now, sources = gathered.get(key, (0.0, 0.0, []))
        estimate = estimate_totals.get(key)
        # Work the estimate never priced starts from nothing, so every rouble of
        # it is increase. Work priced with no estimate to compare against starts
        # from "было", which is the only earlier figure there is.
        baseline = (estimate or 0.0) if from_estimate else was
        current, source = _current(was, now, baseline)
        # Ноль по смете и ноль в файле — строка, которая ничего не сообщает.
        # В таблице от неё только «0 · 0 · 0 %», и среди настоящих цифр это
        # мешает читать.
        if not baseline and not current:
            continue
        rows.append(Row(
            key=key,
            label=estimate_sections.CATEGORY_LABELS.get(key, key),
            sources=sources, was=was, now=now, estimate=estimate,
            baseline=baseline, current=current, delta=current - baseline,
            percent=_percent(baseline, current), source=source,
        ))

    if unmatched:
        # Not an error: the workbook may carry a kind of work this report has no
        # line for. Logged, and named on the page as well, so a row quietly
        # missing from the table can be traced to the row that was skipped.
        logger.info(
            "Разделы файла удорожания без строки в отчёте: %s", "; ".join(unmatched)
        )

    baseline_total = sum(row.baseline for row in rows)
    current_total = sum(row.current for row in rows)
    total = Row(
        key=None, label="Итого", sources=[],
        was=sum(row.was for row in rows), now=sum(row.now for row in rows),
        estimate=sum(row.estimate or 0.0 for row in rows) if from_estimate else None,
        baseline=baseline_total, current=current_total,
        delta=current_total - baseline_total,
        percent=_percent(baseline_total, current_total),
        source=FROM_NOW,
    )
    return Report(
        rows=rows, total=total, unmatched=unmatched, from_estimate=from_estimate,
    )


def read_report(source, estimate_totals=None) -> Report:
    return build_report(read_lines(source), estimate_totals)


def format_percent(value):
    """"+12,1 %" — the way the comparison writes a change, so the same number
    can't be spelled two ways in one app. None where there is no percentage.

    A line that didn't move is written "0 %" rather than "+0,0 %": most lines of
    most files haven't moved, and a column of signed zeroes reads as though
    something happened on every one of them.
    """
    if value is None:
        return None
    if not round(value, 1):
        return "0 %"
    return f"{value:+.1f} %".replace("-", "−").replace(".", ",")


def format_delta(value):
    """The change in roubles with its sign kept — "+13 340 083.64", "0".

    The sign is the point of the column: a line can as easily have come down as
    gone up, and an unsigned number in a column headed "удорожание" would read
    as an increase either way.
    """
    if value is None:
        return None
    if not round(value, 2):
        return "0"
    text = format_number(round(value, 2))
    return f"+{text}" if value > 0 else text.replace("-", "−")
