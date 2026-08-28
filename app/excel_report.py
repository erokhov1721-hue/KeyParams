"""Excel export of the project comparison, laid out like the customer's own
"Анализ стоимости выполнения генподрядных работ" workbook.

One sheet, one column pair per project: the money column ("ГП / Базовый
договор") and the per-square-metre column ("Стоимость на 1 м² ЖК"). Every
per-m² figure, every subtotal and both grand totals are written as live
formulas rather than computed here, so the file keeps working as a working
document: change a project's total area in Excel and its whole column
recalculates.
"""

import logging
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image

from . import estimate_sections, extractors, passport as passport_module, storage

logger = logging.getLogger(__name__)


class ExcelReportError(Exception):
    """Something the user can act on: nothing selected, a project that isn't
    there, a passport file that can't be read. Carries the message shown to
    them, so the route doesn't have to translate error types into words."""


# --- the adapter -----------------------------------------------------------

# Where each report field lives in passport.json. The first name in every
# tuple is the key the app writes today (see passport.PASSPORT_FIELDS and
# passport.CONTRACT_FIELDS); the rest are older or hand-edited spellings that
# have shown up in files, kept so an old passport still exports.
FIELD_ALIASES = {
    "project_name": ("project_name", "name"),
    "year_signed": ("year_signed", "signing_year", "contract_date", "date_signed"),
    "building_class": ("building_class", "class"),
    "general_contractor": ("general_contractor", "contractor"),
    "underground_area_sqm": ("underground_area_sqm", "underground_area"),
    "aboveground_area_sqm": ("aboveground_area_sqm", "aboveground_area"),
    "total_area_sqm": ("total_area_sqm", "total_area"),
    "contract_price_rub": ("contract_price_rub", "contract_price", "price_rub"),
    "vat": ("vat", "vat_rate", "nds"),
    "smr_term": ("smr_term", "smr_period"),
    "advance_payment": ("advance_payment", "advance"),
    "bank_guarantee": ("bank_guarantee", "bank_guarantee_advance"),
    "performance_bond_pct": ("performance_bond_pct", "performance_bond"),
}

# Keys a source-tagged value can hide its actual value behind. This app keeps
# the source in separate lists (ocr_fields / ai_fields) rather than wrapping
# the value, but a passport written by hand or by an older build can carry
# {"value": ..., "source": "ocr"} instead, and the report must read the value
# out of it rather than stringifying the whole dict.
_VALUE_KEYS = ("value", "значение", "text")

_DATE_FORMATS = ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y.%m.%d")


def _unwrap(value):
    if isinstance(value, dict):
        for key in _VALUE_KEYS:
            if key in value:
                return value[key]
        return None
    return value


def _lookup(passport, aliases):
    for key in aliases:
        if key in passport:
            value = _unwrap(passport[key])
            if value is not None and value != "":
                return value
    return None


def _text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _number(value):
    """A float, whatever the passport stored: a real number, or a string like
    "10 067 050 887,72" written the Russian way."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value)
    if text is None:
        return None
    return extractors.parse_number(text)


def _signing(value):
    """``(date_or_None, year_or_None)`` for whatever the passport calls a
    signing date: a bare year ("2025"), a full date in any of the usual
    orders, or a real date object. The date is returned separately because
    only a full date can be shown as "февраль 2025" — a bare year is written
    as a year."""
    if isinstance(value, datetime):
        return value.date(), value.year
    if isinstance(value, date):
        return value, value.year
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        year = int(value)
        return None, year if 1900 <= year <= 2100 else None

    text = _text(value)
    if text is None:
        return None, None
    for fmt in _DATE_FORMATS:
        try:
            parsed = datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        return parsed, parsed.year
    year = re.search(r"(19|20)\d{2}", text)
    return None, int(year.group()) if year else None


def normalize_passport(passport: dict) -> dict:
    """The report's view of one passport: canonical names, real numbers, a
    signing date split from its year, and the VAT rate.

    VAT follows the same rule as the rest of the app (see
    ``passport.vat_for_year``): it is derived from the signing year — 20%
    through 2025, 22% from 2026 — and only falls back to whatever the
    passport stored when no year is known. The rate is statutory, so a figure
    misread off a scan must not reach the report.
    """
    raw = {field: _lookup(passport, aliases) for field, aliases in FIELD_ALIASES.items()}
    signing_date, year = _signing(raw["year_signed"])

    return {
        "project_name": _text(raw["project_name"]),
        "signing_date": signing_date,
        "year_signed": year,
        "building_class": _text(raw["building_class"]),
        "general_contractor": _text(raw["general_contractor"]),
        "underground_area_sqm": _number(raw["underground_area_sqm"]),
        "aboveground_area_sqm": _number(raw["aboveground_area_sqm"]),
        "total_area_sqm": _number(raw["total_area_sqm"]),
        "contract_price_rub": _number(raw["contract_price_rub"]),
        "vat": passport_module.vat_for_year(year) or _text(raw["vat"]),
        "smr_term": _text(raw["smr_term"]),
        "advance_payment": _text(raw["advance_payment"]),
        "bank_guarantee": _text(raw["bank_guarantee"]),
        "performance_bond_pct": _text(raw["performance_bond_pct"]),
    }


def load_project(project_dir) -> dict:
    """Read one project folder into ``{"passport": ..., "cover": ...}``.

    Goes through the app's own storage and passport helpers rather than
    walking the folder itself, so the report sees a project exactly as the
    rest of the app does — same passport path, same cover-file rules.
    """
    project_dir = Path(project_dir)
    root, slug = project_dir.parent, project_dir.name

    path = storage.passport_path(root, slug)
    if not path.exists():
        raise ExcelReportError(
            f"Проект «{slug}» не найден или не достроен — у него нет паспорта. "
            "Обновите список проектов и выберите заново."
        )
    try:
        passport = passport_module.load_passport(path)
    except passport_module.PassportReadError as e:
        raise ExcelReportError(
            f"Не удалось прочитать паспорт проекта «{slug}» — файл повреждён. "
            "Создайте проект заново, загрузив исходные документы."
        ) from e

    costs, sources, _unmatched = estimate_costs(root, slug)
    return {
        "passport": passport,
        "cover": storage.cover_path(root, slug),
        "costs": costs,
        "cost_sources": sources,
    }


def estimate_costs(root, slug):
    """``({line: total}, {line: [section names]}, [unmatched section names])``
    from the project's estimate.

    An estimate that can't be parsed is not worth failing the whole export
    over: the passport half of the report is still worth having, and the cost
    lines fall back to being blank, which is what they were before an estimate
    was attached at all.
    """
    path = storage.estimate_path(root, slug)
    if not path.exists():
        return {}, {}, []
    try:
        sections, unmatched = estimate_sections.read_sections_with_warnings(path)
    except estimate_sections.EstimateSectionsError:
        logger.exception("Не удалось разобрать смету проекта «%s»", slug)
        return {}, {}, []

    totals, sources = {}, {}
    for section in sections:
        totals[section.key] = totals.get(section.key, Decimal("0")) + section.amount
        sources.setdefault(section.key, []).append(section.name)
    return totals, sources, unmatched


# --- the sheet -------------------------------------------------------------

SHEET_TITLE = "Проекты"
FONT_NAME = "Arial"

# Resolved from the customer workbook's own theme, then written as plain RGB:
# a theme colour would come out of a freshly created workbook as whatever
# theme openpyxl ships, which is not necessarily the one these were picked in.
COLOR_HEADER = "DAE3F3"     # project-name band (accent1, lighter)
COLOR_DARK = "203864"       # table header and the two total rows
COLOR_BAND = "D0CECE"       # ₽/м² strip, contract total, MR rows
COLOR_EMPTY = "FFF2CC"      # nothing was extracted — fill this in by hand
COLOR_GRID = "A6A6A6"
COLOR_WHITE = "FFFFFF"
COLOR_DEVIATION = "FF0000"

FMT_AREA = '#,##0.0" м2"'
FMT_MONEY = '#,##0\\ "₽"'
FMT_PER_SQM = '#,##0" ₽/m2"'
FMT_MONTH_YEAR = "[$-419]mmmm\\ yyyy;@"
FMT_YEAR = "0"
FMT_DEVIATION = "+0%;-0%"
FMT_INDEX = "@"

ROW_NAME = 3
ROW_PHOTO = 4
ROW_YEAR = 5
ROW_CLASS = 6
ROW_CONTRACTOR = 7
ROW_UNDERGROUND = 8
ROW_ABOVEGROUND = 9
ROW_TOTAL_AREA = 10
ROW_VAT = 11
ROW_PER_SQM = 13
ROW_TABLE_HEAD = 14
ROW_CONTRACT_TOTAL = 15
ROW_WORK_FIRST = 16
ROW_WORK_LAST = 29
ROW_SMR_TOTAL = 30
ROW_MR_FIRST = 31
ROW_MR_LAST = 33
ROW_GRAND_TOTAL = 34
ROW_TERMS_HEAD = 36
ROW_TERMS_FIRST = 37

CHARACTERISTIC_LABELS = [
    (ROW_YEAR, "Год подписания договора"),
    (ROW_CLASS, "Класс здания"),
    (ROW_CONTRACTOR, "Генподрядчик"),
    (ROW_UNDERGROUND, "Площадь подземной комплекса по СП / МР, м² (ТЭП)"),
    (ROW_ABOVEGROUND, "Площадь надземной комплекса по СП / МР, м² (ТЭП)"),
    (ROW_TOTAL_AREA, "Общая площадь комплекса по СП / МР, м² (ТЭП)"),
    (ROW_VAT, "Ставка НДС"),
]

# The report's cost lines and the categories an estimate is read into are the
# same list in the same order, which is what lets a section's total land on
# its own row. Both come from one place so they cannot drift apart.
WORK_KEYS = estimate_sections.WORK_CATEGORY_KEYS
MR_KEYS = estimate_sections.MR_CATEGORY_KEYS
WORK_LABELS = [estimate_sections.CATEGORY_LABELS[key] for key in WORK_KEYS]
MR_LABELS = [estimate_sections.CATEGORY_LABELS[key] for key in MR_KEYS]

TERMS_LABELS = [
    ("smr_term", "Срок СМР"),
    ("advance_payment", "Аванс"),
    ("bank_guarantee", "Банковская гарантия на возврат аванса"),
    ("performance_bond_pct", "Performance bond, %"),
]

COL_GUTTER = 1          # A — the narrow margin the customer's sheet starts with
COL_INDEX = 2           # B — № of the cost line / characteristic labels
COL_LABEL = 3           # C — cost line name
COL_FIRST_PROJECT = 5   # E — first project's money column
COLS_PER_PROJECT = 3    # money, per-m², separator

WIDTH_GUTTER = 3.14
WIDTH_INDEX = 13.71
WIDTH_LABEL = 69.29
WIDTH_MONEY = 27.71
# Wide enough for a per-m² figure in the customer's own format at Arial 12
# ("132 914 ₽/m2"); anything narrower and Excel shows the cell as ######.
WIDTH_PER_SQM = 20.0

ROW_HEIGHTS = {
    ROW_NAME: 67.5,
    ROW_PHOTO: 147.0,
    12: 15.6,
    ROW_PER_SQM: 89.25,
    ROW_TERMS_HEAD: 31.5,
}

# The photo cell, in pixels. Excel's own conversion: a column's width is in
# characters, a row's height in points.
PHOTO_BOX_PX = (
    round(WIDTH_MONEY * 7) + 5 + round(WIDTH_PER_SQM * 7) + 5,
    round(ROW_HEIGHTS[ROW_PHOTO] * 96 / 72),
)
PHOTO_PADDING_PX = 6

MEDIUM = Side(style="medium", color="FF000000")
THIN = Side(style="thin", color=COLOR_GRID)
THIN_BLACK = Side(style="thin", color="FF000000")
HAIR = Side(style="hair", color="FF000000")

_FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
_FILL_DARK = PatternFill("solid", fgColor=COLOR_DARK)
_FILL_BAND = PatternFill("solid", fgColor=COLOR_BAND)
_FILL_EMPTY = PatternFill("solid", fgColor=COLOR_EMPTY)


def _font(size=12, bold=True, color=None):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def _money_col(index):
    return COL_FIRST_PROJECT + index * COLS_PER_PROJECT


class _Formula(str):
    """Marks a string this module built itself as a formula — as opposed to
    a project name, a contract term, or any other text that came from a
    document or a person and happens to start with the same character. Only
    a ``_Formula`` is ever allowed to become a live formula in the sheet;
    every other string is pinned to plain text below, in ``_write``."""


def _write(ws, row, col, value=None, *, span=1, fmt=None, font=None, fill=None,
           align=None, top=None, bottom=None, left=None, right=None):
    """Write one cell, optionally merged across ``span`` columns.

    A merged range only takes its value and font from the top-left cell, but
    every cell in it has to carry the fill and the outer border or Excel draws
    the block with gaps in it.

    A plain string is pinned to text (``data_type = 's'``) after being set,
    regardless of what it starts with: openpyxl infers a leading "=" as a
    formula on its own, and a project name or a contract term is free text
    from a document or a person, not something this module wrote itself. A
    ``_Formula`` is the one exception — the report's own generated formulas
    (running totals, ₽/м², deviations) still need to compute.
    """
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if isinstance(value, str) and not isinstance(value, _Formula):
        cell.data_type = 's'
    if fmt is not None:
        cell.number_format = fmt
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)

    last = col + span - 1
    for c in range(col, last + 1):
        target = ws.cell(row=row, column=c)
        if fill is not None:
            target.fill = fill
        target.border = Border(
            top=top, bottom=bottom,
            left=left if c == col else None,
            right=right if c == last else None,
        )
    return cell


_CENTER = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _label_column(ws, projects):
    """Column B/C: the project-name corner, the characteristic labels, the
    numbered cost lines and the two total rows."""
    # The corner cell covers both the name row and the photo row, as it does
    # in the customer's sheet — there is no photo to put beside the labels.
    _write(ws, ROW_NAME, COL_INDEX, "Проекты", font=_font(16), fill=_FILL_HEADER,
           align=_CENTER, top=MEDIUM, left=MEDIUM)
    _write(ws, ROW_NAME, COL_LABEL, None, fill=_FILL_HEADER, top=MEDIUM, right=MEDIUM)
    _write(ws, ROW_PHOTO, COL_INDEX, None, fill=_FILL_HEADER, bottom=THIN, left=MEDIUM)
    _write(ws, ROW_PHOTO, COL_LABEL, None, fill=_FILL_HEADER, bottom=THIN, right=MEDIUM)
    ws.merge_cells(start_row=ROW_NAME, start_column=COL_INDEX,
                   end_row=ROW_PHOTO, end_column=COL_LABEL)

    for row, label in CHARACTERISTIC_LABELS:
        _write(ws, row, COL_INDEX, label, span=2, font=_font(),
               align=_LEFT_WRAP, top=THIN,
               bottom=MEDIUM if row == ROW_VAT else THIN,
               left=MEDIUM, right=MEDIUM)

    # The header spans down over the contract-total row, exactly as it does in
    # the customer's sheet: that row is a per-project figure, not a cost line.
    _write(ws, ROW_TABLE_HEAD, COL_INDEX, "№", font=_font(color=COLOR_WHITE),
           fill=_FILL_DARK, align=_CENTER, top=MEDIUM, left=MEDIUM)
    _write(ws, ROW_TABLE_HEAD, COL_LABEL, "Стоимость по видам работ/расход:",
           font=_font(color=COLOR_WHITE), fill=_FILL_DARK, align=_CENTER_WRAP,
           top=MEDIUM, left=MEDIUM, right=MEDIUM)
    _write(ws, ROW_CONTRACT_TOTAL, COL_INDEX, None, fill=_FILL_DARK,
           bottom=MEDIUM, left=MEDIUM)
    _write(ws, ROW_CONTRACT_TOTAL, COL_LABEL, None, fill=_FILL_DARK,
           bottom=MEDIUM, left=MEDIUM, right=MEDIUM)
    ws.merge_cells(start_row=ROW_TABLE_HEAD, start_column=COL_INDEX,
                   end_row=ROW_CONTRACT_TOTAL, end_column=COL_INDEX)
    ws.merge_cells(start_row=ROW_TABLE_HEAD, start_column=COL_LABEL,
                   end_row=ROW_CONTRACT_TOTAL, end_column=COL_LABEL)

    for offset, label in enumerate(WORK_LABELS):
        row = ROW_WORK_FIRST + offset
        _write(ws, row, COL_INDEX, str(offset + 1), fmt=FMT_INDEX, font=_font(),
               align=_CENTER, top=HAIR, bottom=HAIR, left=MEDIUM, right=THIN_BLACK)
        _write(ws, row, COL_LABEL, label, font=_font(), align=_LEFT_WRAP,
               top=HAIR, bottom=HAIR, left=THIN_BLACK, right=MEDIUM)

    vat = _shared_vat(projects)
    _write(ws, ROW_SMR_TOTAL, COL_INDEX, None, fill=_FILL_DARK,
           top=MEDIUM, bottom=MEDIUM, left=MEDIUM)
    _write(ws, ROW_SMR_TOTAL, COL_LABEL, f"Итого СМР, руб.{vat}",
           font=_font(color=COLOR_WHITE), fill=_FILL_DARK, align=_LEFT_WRAP,
           top=MEDIUM, bottom=MEDIUM, right=MEDIUM)

    for offset, label in enumerate(MR_LABELS):
        row = ROW_MR_FIRST + offset
        _write(ws, row, COL_INDEX, str(len(WORK_LABELS) + offset + 1), fmt=FMT_INDEX,
               font=_font(), fill=_FILL_BAND, align=_CENTER,
               top=MEDIUM if offset == 0 else HAIR, bottom=HAIR,
               left=MEDIUM, right=THIN_BLACK)
        _write(ws, row, COL_LABEL, label, font=_font(), fill=_FILL_BAND,
               align=_LEFT_WRAP, top=MEDIUM if offset == 0 else HAIR, bottom=HAIR,
               left=THIN_BLACK, right=MEDIUM)

    _write(ws, ROW_GRAND_TOTAL, COL_INDEX, None, fill=_FILL_DARK,
           top=MEDIUM, bottom=MEDIUM, left=MEDIUM)
    _write(ws, ROW_GRAND_TOTAL, COL_LABEL,
           f"Итого СМР, в тч Отделка и Нулевой цикл, руб.{vat}",
           font=_font(color=COLOR_WHITE), fill=_FILL_DARK, align=_LEFT_WRAP,
           top=MEDIUM, bottom=MEDIUM, right=MEDIUM)

    _write(ws, ROW_TERMS_HEAD, COL_INDEX, "Паспорт договора", span=2,
           font=_font(color=COLOR_WHITE), fill=_FILL_DARK, align=_LEFT,
           top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)
    for offset, (_, label) in enumerate(TERMS_LABELS):
        row = ROW_TERMS_FIRST + offset
        _write(ws, row, COL_INDEX, label, span=2, font=_font(), align=_LEFT_WRAP,
               top=THIN, bottom=MEDIUM if offset == len(TERMS_LABELS) - 1 else THIN,
               left=MEDIUM, right=MEDIUM)


def _shared_vat(projects):
    """" с НДС 20%" when every project agrees on the rate, "" when they don't —
    one label serves all the columns, so it can only name a rate they share."""
    rates = {p["vat"] for p in projects if p["vat"]}
    return f" с НДС {rates.pop()}" if len(rates) == 1 else ""


def _value_style(value):
    """Empty values are highlighted rather than zero-filled: a missing figure
    is something to go and find, not a nought."""
    return _FILL_EMPTY if value is None else None


COMMENT_AUTHOR = "Паспорт объекта"
COMMENT_PREFIX = "Из сметы:"


def _note_source(cell, names):
    """Say which sections of the estimate a figure was made of.

    Placing a section on a line is a judgement about wording, and the one
    place it can be checked is next to the number itself. Two sections landing
    on one line — an offer's "Прочее" and its unnumbered "Дополнительные
    работы" both belong to "Другие" — are both named, so a total that looks
    too big can be taken apart without opening the estimate.
    """
    if not names:
        return
    cell.comment = Comment(f"{COMMENT_PREFIX} {'; '.join(names)}", COMMENT_AUTHOR)


def _missing_cost_style(value, costs):
    """A cost line is only worth highlighting once there is an estimate to
    compare it against. With no estimate attached every line is empty by
    design and is meant to be typed in, so marking all fourteen would be a
    wall of yellow saying nothing."""
    return _FILL_EMPTY if value is None and costs else None


def _project_column(ws, index, project, cover, costs=None, sources=None):
    costs = costs or {}
    sources = sources or {}
    col = _money_col(index)
    money = get_column_letter(col)
    per_sqm = get_column_letter(col + 1)
    first_money = get_column_letter(_money_col(0))

    _write(ws, ROW_NAME, col, project["project_name"] or "Без названия", span=2,
           font=_font(18), fill=_FILL_HEADER, align=_CENTER_WRAP,
           top=MEDIUM, bottom=THIN, left=MEDIUM, right=MEDIUM)
    _write(ws, ROW_PHOTO, col, None, span=2, align=_CENTER,
           top=THIN, bottom=THIN, left=MEDIUM, right=MEDIUM)
    if cover is not None:
        _insert_cover(ws, col, cover)

    if project["signing_date"] is not None:
        year_value, year_fmt = project["signing_date"], FMT_MONTH_YEAR
    elif project["year_signed"] is not None:
        year_value, year_fmt = project["year_signed"], FMT_YEAR
    else:
        year_value, year_fmt = None, None

    areas = (project["underground_area_sqm"], project["aboveground_area_sqm"])
    total_area = project["total_area_sqm"]
    if total_area is None and any(a is not None for a in areas):
        # The customer's own sheet derives it this way, and keeping it a
        # formula means correcting either part fixes the total too.
        total_value = _Formula(f"=+{money}{ROW_UNDERGROUND}+{money}{ROW_ABOVEGROUND}")
    else:
        total_value = total_area

    cells = [
        (ROW_YEAR, year_value, year_fmt),
        (ROW_CLASS, project["building_class"], None),
        (ROW_CONTRACTOR, project["general_contractor"], None),
        (ROW_UNDERGROUND, areas[0], FMT_AREA),
        (ROW_ABOVEGROUND, areas[1], FMT_AREA),
        (ROW_TOTAL_AREA, total_value, FMT_AREA),
        (ROW_VAT, project["vat"], None),
    ]
    for row, value, fmt in cells:
        _write(ws, row, col, value, span=2, fmt=fmt, font=_font(),
               fill=_value_style(value), align=_CENTER_WRAP, top=THIN,
               bottom=MEDIUM if row == ROW_VAT else THIN,
               left=MEDIUM, right=MEDIUM)

    # The headline ₽/м², and next to it how far this project sits from the
    # first one — the column everything else is being compared against.
    _write(ws, ROW_PER_SQM, col, _Formula(f"=+{per_sqm}{ROW_CONTRACT_TOTAL}"), fmt=FMT_PER_SQM,
           font=_font(), fill=_FILL_BAND, align=_CENTER,
           top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=THIN_BLACK)
    deviation = None if index == 0 else _Formula(
        f"=IFERROR({money}{ROW_PER_SQM}/${first_money}${ROW_PER_SQM}-1,\"\")"
    )
    _write(ws, ROW_PER_SQM, col + 1, deviation, fmt=FMT_DEVIATION,
           font=_font(color=COLOR_DEVIATION), align=_CENTER,
           top=MEDIUM, bottom=MEDIUM, left=THIN_BLACK, right=MEDIUM)

    _write(ws, ROW_TABLE_HEAD, col, "ГП  / Базовый договор",
           font=_font(color=COLOR_WHITE), fill=_FILL_DARK, align=_CENTER_WRAP,
           top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=THIN_BLACK)
    _write(ws, ROW_TABLE_HEAD, col + 1, "Стоимость\nна 1 м² ЖК",
           font=_font(color=COLOR_WHITE), fill=_FILL_DARK, align=_CENTER_WRAP,
           top=MEDIUM, bottom=MEDIUM, left=THIN_BLACK, right=MEDIUM)

    _write(ws, ROW_CONTRACT_TOTAL, col, _Formula(f"=+{money}{ROW_GRAND_TOTAL}"), fmt=FMT_MONEY,
           font=_font(), fill=_FILL_BAND, align=_CENTER,
           top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=THIN_BLACK)
    _per_sqm_cell(ws, ROW_CONTRACT_TOTAL, col, money, fill=_FILL_BAND,
                  top=MEDIUM, bottom=MEDIUM)

    for offset in range(len(WORK_LABELS)):
        row = ROW_WORK_FIRST + offset
        value = costs.get(WORK_KEYS[offset])
        cell = _write(ws, row, col, value, fmt=FMT_MONEY, font=_font(), align=_CENTER,
                      fill=_missing_cost_style(value, costs),
                      top=HAIR, bottom=HAIR, left=MEDIUM, right=THIN_BLACK)
        _note_source(cell, sources.get(WORK_KEYS[offset]))
        _per_sqm_cell(ws, row, col, money, top=HAIR, bottom=HAIR)

    _write(ws, ROW_SMR_TOTAL, col,
           _Formula(f"=SUBTOTAL(9,{money}{ROW_WORK_FIRST}:{money}{ROW_WORK_LAST})"),
           fmt=FMT_MONEY, font=_font(color=COLOR_WHITE), fill=_FILL_DARK,
           align=_CENTER, top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=THIN_BLACK)
    _per_sqm_cell(ws, ROW_SMR_TOTAL, col, money, fill=_FILL_DARK,
                  color=COLOR_WHITE, top=MEDIUM, bottom=MEDIUM)

    for offset in range(len(MR_LABELS)):
        row = ROW_MR_FIRST + offset
        cell = _write(ws, row, col, costs.get(MR_KEYS[offset]), fmt=FMT_MONEY,
                      font=_font(), fill=_FILL_BAND,
                      align=_CENTER, top=MEDIUM if offset == 0 else HAIR, bottom=HAIR,
                      left=MEDIUM, right=THIN_BLACK)
        _note_source(cell, sources.get(MR_KEYS[offset]))
        _per_sqm_cell(ws, row, col, money, fill=_FILL_BAND,
                      top=MEDIUM if offset == 0 else HAIR, bottom=HAIR)

    _write(ws, ROW_GRAND_TOTAL, col, _grand_total_formula(money, project),
           fmt=FMT_MONEY, font=_font(color=COLOR_WHITE), fill=_FILL_DARK,
           align=_CENTER, top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=THIN_BLACK)
    _per_sqm_cell(ws, ROW_GRAND_TOTAL, col, money, fill=_FILL_DARK,
                  color=COLOR_WHITE, top=MEDIUM, bottom=MEDIUM)

    _write(ws, ROW_TERMS_HEAD, col, None, span=2, fill=_FILL_DARK,
           top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)
    for offset, (field, _) in enumerate(TERMS_LABELS):
        row = ROW_TERMS_FIRST + offset
        value = project[field]
        _write(ws, row, col, value, span=2, font=_font(), fill=_value_style(value),
               align=_CENTER_WRAP, top=THIN,
               bottom=MEDIUM if offset == len(TERMS_LABELS) - 1 else THIN,
               left=MEDIUM, right=MEDIUM)


def _grand_total_formula(money, project):
    """The bottom total, and the one cell the contract price reaches.

    While the cost lines above are still empty — they are filled in from the
    estimate by hand — the total falls back to the price agreed in the
    contract, so a freshly exported report already shows real money. The
    moment anything is entered above, the sum takes over on its own.
    """
    parts = "+".join(
        f"{money}{row}" for row in
        [ROW_SMR_TOTAL] + list(range(ROW_MR_FIRST, ROW_MR_LAST + 1))
    )
    price = project["contract_price_rub"]
    if price is None:
        return _Formula(f"=+{parts}")
    return _Formula(f"=IF({parts}=0,{price:.2f},{parts})")


def _per_sqm_cell(ws, row, col, money, *, fill=None, color=None, top=None, bottom=None):
    """The "Стоимость на 1 м² ЖК" cell next to a money cell — always the same
    division, always a formula, wrapped so a project with no total area shows
    an empty cell instead of #DIV/0! down the whole column."""
    _write(
        ws, row, col + 1,
        _Formula(f"=IFERROR(+{money}{row}/${money}${ROW_TOTAL_AREA},\"\")"),
        fmt=FMT_PER_SQM, font=_font(color=color), fill=fill, align=_CENTER,
        top=top, bottom=bottom, left=THIN_BLACK, right=MEDIUM,
    )


def _insert_cover(ws, col, cover_path):
    """Put the object's photo in the picture row, scaled to fit the cell and
    centred in it. A photo that can't be read is simply left out — an export
    must not fail over a broken image file."""
    try:
        with Image.open(cover_path) as source:
            image = source.convert("RGB")
            box_w = PHOTO_BOX_PX[0] - 2 * PHOTO_PADDING_PX
            box_h = PHOTO_BOX_PX[1] - 2 * PHOTO_PADDING_PX
            scale = min(box_w / image.width, box_h / image.height, 1.0)
            width = max(1, round(image.width * scale))
            height = max(1, round(image.height * scale))
            if scale < 1.0:
                image = image.resize((width, height), Image.LANCZOS)
            buffer = BytesIO()
            image.save(buffer, format="PNG")
    except (OSError, ValueError):
        return

    buffer.seek(0)
    picture = XLImage(buffer)
    picture.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=col - 1, colOff=pixels_to_EMU(max(0, (PHOTO_BOX_PX[0] - width) // 2)),
            row=ROW_PHOTO - 1, rowOff=pixels_to_EMU(max(0, (PHOTO_BOX_PX[1] - height) // 2)),
        ),
        ext=XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height)),
    )
    ws.add_image(picture)


def _layout(ws, count):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 60
    ws.freeze_panes = ws.cell(row=1, column=COL_FIRST_PROJECT - 1)

    ws.column_dimensions[get_column_letter(COL_GUTTER)].width = WIDTH_GUTTER
    ws.column_dimensions[get_column_letter(COL_INDEX)].width = WIDTH_INDEX
    ws.column_dimensions[get_column_letter(COL_LABEL)].width = WIDTH_LABEL
    for index in range(count):
        col = _money_col(index)
        ws.column_dimensions[get_column_letter(col - 1)].width = WIDTH_GUTTER
        ws.column_dimensions[get_column_letter(col)].width = WIDTH_MONEY
        ws.column_dimensions[get_column_letter(col + 1)].width = WIDTH_PER_SQM

    for row in range(ROW_YEAR, ROW_VAT + 1):
        ws.row_dimensions[row].height = 24.95
    for row in range(ROW_TABLE_HEAD, ROW_TERMS_FIRST + len(TERMS_LABELS)):
        ws.row_dimensions[row].height = 31.5
    # The one label long enough to need a second line at this column width.
    ws.row_dimensions[ROW_WORK_FIRST + 1].height = 47.25
    for row, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height


def _as_projects(projects):
    normalized = []
    covers = []
    costs = []
    sources = []
    for item in projects:
        if isinstance(item, dict) and "passport" in item:
            normalized.append(normalize_passport(item["passport"]))
            covers.append(item.get("cover"))
            costs.append(item.get("costs") or {})
            sources.append(item.get("cost_sources") or {})
        else:
            normalized.append(normalize_passport(item))
            covers.append(None)
            costs.append({})
            sources.append({})
    return normalized, covers, costs, sources


def build_comparison_report(projects, out_path):
    """Write the comparison workbook for ``projects`` to ``out_path``.

    ``projects`` is what ``load_project`` returns — ``{"passport": ...,
    "cover": ...}`` — or bare passport dicts when there are no photos to
    place. ``out_path`` is a path or any file object openpyxl can save to.
    """
    normalized, covers, costs, sources = _as_projects(projects)
    if not normalized:
        raise ExcelReportError(
            "Не выбрано ни одного проекта. Отметьте проекты галочками в списке "
            "и нажмите «Выгрузить в Excel»."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE

    _layout(ws, len(normalized))
    _label_column(ws, normalized)
    for index, column in enumerate(zip(normalized, covers, costs, sources)):
        _project_column(ws, index, *column)

    wb.save(out_path)
    return out_path
