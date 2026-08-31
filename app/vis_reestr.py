"""Reads the "Реестр претензий ВИС" workbook and aggregates it for the
"Прогнозируемое удорожание" dashboard.

Ported from a standalone Node/Express tool (github.com/erokhov1721-hue/Reestr)
that read the same workbook and served the aggregates to a Chart.js frontend.
This app has no client-side charting library and no CDN dependency anywhere
(see app/chart_render.py), so the aggregation logic is kept — same column
layout, same grouping rules — but rendering is server-side, into the same
proportional bar rows the rest of the app already uses (see
passport.build_comparison_charts).
"""

from datetime import date, datetime, timedelta
from decimal import Decimal

import openpyxl

from . import extractors, workbook_cache
from . import passport as passport_module

SHEET_NAME = "Общая сводная"

# 0-based column indices, matching the real workbook's layout — the first
# column is blank, so data starts at index 1.
COL_TYPE = 1
COL_OBJECT = 2
COL_DATE = 5
COL_INITIAL = 8
COL_AGREED = 11
COL_STATUS = 15

HEADER_SEARCH_ROWS = 20
DEFAULT_HEADER_ROW = 4

PRICE_INCREASE_MARKER = "УДОРОЖАНИЕ"
COMPLETED_MARKER = "ЗАВЕРШЕН"

_EXCEL_EPOCH = datetime(1899, 12, 30)
_DATE_STRING_FORMATS = ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S")


class VisReestrError(Exception):
    """The workbook exists but can't be read as a claims registry — wrong
    sheet, or not a workbook at all. Its message is shown on the page
    as-is, so it names the actual problem rather than just "failed"."""


# Looked up by a fixed code from the query string, the same way
# app.estimate.PROBLEM_MESSAGES and passport.DGP_PROBLEM_MESSAGES are — an
# arbitrary ?reestr_problem=... in the address bar shows nothing on the page.
PROBLEM_MESSAGES = {
    "format": "Загрузите файл реестра в формате .xlsx",
    "too_big": "Файл слишком большой — не более 15 МБ.",
    "unreadable": (
        "Не удалось прочитать файл — убедитесь, что это корректный .xlsx "
        "с листом «Общая сводная». Прежний реестр оставлен на месте."
    ),
}


def _cell(row, index):
    return row[index] if index < len(row) else None


def _find_header_row(rows):
    """The first row whose TYPE column says "тип" — the header. Falls back
    to a fixed row, same as the original tool, rather than failing outright:
    a header row that moved by a row or two still leaves the columns below
    it in the right place."""
    for i, row in enumerate(rows[:HEADER_SEARCH_ROWS]):
        cell = _cell(row, COL_TYPE)
        if cell and "тип" in str(cell).strip().lower():
            return i
    return DEFAULT_HEADER_ROW


def _to_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (_EXCEL_EPOCH + timedelta(days=value)).date()
        except (OverflowError, OSError):
            return None
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for fmt in _DATE_STRING_FORMATS:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _to_money(value):
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        return extractors.parse_money(value) or Decimal("0")
    return Decimal("0")


def _upper(value):
    return str(value).strip().upper() if value else ""


def _records_from_workbook(wb):
    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(f'"{name}"' for name in wb.sheetnames)
        raise VisReestrError(
            f'Лист "{SHEET_NAME}" не найден в файле. Доступные листы: {available}'
        )
    sheet = wb[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    header_idx = _find_header_row(rows)

    records = []
    for row in rows[header_idx + 1:]:
        object_name = _cell(row, COL_OBJECT)
        object_name = str(object_name).strip() if object_name else ""
        if not object_name:
            continue
        records.append({
            "type": _upper(_cell(row, COL_TYPE)),
            "object": object_name,
            "date": _to_date(_cell(row, COL_DATE)),
            "initial_sum": _to_money(_cell(row, COL_INITIAL)),
            "agreed_sum": _to_money(_cell(row, COL_AGREED)),
            "status": _upper(_cell(row, COL_STATUS)),
        })
    return records


def parse_records(source):
    """Records from a path or a file-like object — used to validate an
    upload before it's saved (mirrors ``cost_increase.read_lines``)."""
    try:
        wb = openpyxl.load_workbook(source, data_only=True)
    except (openpyxl.utils.exceptions.InvalidFileException, KeyError, OSError) as e:
        raise VisReestrError(f"Не удалось открыть файл Excel: {e}") from e
    return _records_from_workbook(wb)


def parse_records_cached(path):
    """Records from a path already saved to disk, reusing the parsed
    workbook across repeat page views (date-filter changes) the way
    ``app.estimate`` already does for project estimates."""
    wb = workbook_cache.get_or_load(path, data_only=True)
    return _records_from_workbook(wb)


def _filter_by_date(records, date_from, date_to):
    if date_from is None and date_to is None:
        return list(records)
    result = []
    for r in records:
        if r["date"] is None:
            continue
        if date_from and r["date"] < date_from:
            continue
        if date_to and r["date"] > date_to:
            continue
        result.append(r)
    return result


def _group_count(items, key_fn):
    counts = {}
    for item in items:
        key = key_fn(item)
        if not key:
            continue
        counts[key] = counts.get(key, 0) + 1
    return counts


def _group_sum(items, key_fn, value_fn):
    sums = {}
    for item in items:
        key = key_fn(item)
        if not key:
            continue
        sums[key] = sums.get(key, Decimal("0")) + value_fn(item)
    return sums


def _sorted_counts(counts):
    return [
        {"name": name, "count": count}
        for name, count in sorted(counts.items(), key=lambda kv: kv[1], reverse=True)
    ]


def _sorted_sums(sums):
    return [
        {"name": name, "sum": total}
        for name, total in sorted(sums.items(), key=lambda kv: kv[1], reverse=True)
    ]


def build_analytics(records, date_from=None, date_to=None):
    """Aggregates ``records`` (as returned by ``parse_records``) the same
    way the original tool's ``/api/analytics`` endpoint did — grouped counts
    and sums, ready to hand a template. ``date_from``/``date_to`` are
    ``date`` objects, inclusive, or None for no bound on that side."""
    filtered = _filter_by_date(records, date_from, date_to)

    requests_by_object = _sorted_counts(_group_count(filtered, lambda r: r["object"]))

    total_initial_sum = sum((r["initial_sum"] for r in filtered), Decimal("0"))
    total_agreed_sum = sum((r["agreed_sum"] for r in filtered), Decimal("0"))

    price_rows = [r for r in filtered if PRICE_INCREASE_MARKER in r["type"]]
    price_increase = _sorted_sums(
        _group_sum(price_rows, lambda r: r["object"], lambda r: r["agreed_sum"])
    )

    by_type = _sorted_counts(_group_count(filtered, lambda r: r["type"] or "Не указано"))
    by_status = _sorted_counts(_group_count(filtered, lambda r: r["status"] or "Не указан"))

    completed_count = sum(1 for r in filtered if COMPLETED_MARKER in r["status"])

    return {
        "meta": {
            "total_records": len(records),
            "filtered_records": len(filtered),
        },
        "kpi": {
            "total_requests": len(filtered),
            "completed_count": completed_count,
            "total_initial_sum": total_initial_sum,
            "total_agreed_sum": total_agreed_sum,
        },
        "requests_by_object": requests_by_object,
        "price_increase": price_increase,
        "by_type": by_type,
        "by_status": by_status,
    }


_MONEY_SCALE = [(Decimal("1000000000"), "млрд"), (Decimal("1000000"), "млн"), (Decimal("1000"), "тыс")]


def format_money_short(value):
    """A rouble figure abbreviated to its largest round unit — "24.25 млрд
    ₽" rather than every digit. Same idea as ``passport._format_money_short``,
    kept local rather than imported so this module has no dependency on
    passport's own (underscore-prefixed, not a public API) formatting."""
    if value is None:
        return "—"
    magnitude = abs(value)
    sign = "-" if value < 0 else ""
    for scale, suffix in _MONEY_SCALE:
        if magnitude >= scale:
            return f"{sign}{magnitude / scale:.2f} {suffix} ₽"
    grouped = f"{magnitude:,.0f}".replace(",", " ")
    return f"{sign}{grouped} ₽"


def _finalize_bars(rows, value_key, *, money):
    """Adds ``width_pct`` and ``display`` (and, for money rows,
    ``short_display``) to each row in place — the same shape
    ``passport._finalize_chart`` builds for the comparison page's bar
    charts, reused here via the same template macro pattern."""
    if not rows:
        return rows
    max_value = max(row[value_key] for row in rows) or 1
    for row in rows:
        row["width_pct"] = round(float(row[value_key]) / float(max_value) * 100, 1)
        if money:
            row["display"] = f"{passport_module.format_number(row[value_key])} ₽"
            row["short_display"] = format_money_short(row[value_key])
        else:
            row["display"] = str(row[value_key])
    return rows


def finalize_requests_by_object(rows):
    return _finalize_bars(rows, "count", money=False)


def finalize_price_increase(rows):
    rows = _finalize_bars(rows, "sum", money=True)
    total = sum((row["sum"] for row in rows), Decimal("0"))
    for i, row in enumerate(rows, start=1):
        row["rank"] = i
        row["share_pct"] = round(float(row["sum"] / total * 100), 1) if total else None
    return rows


def finalize_breakdown(rows):
    """``by_type``/``by_status`` — a count-based breakdown, same bars as
    ``finalize_requests_by_object`` plus each row's share of the total, so
    the list reads as a distribution rather than just a ranking."""
    rows = _finalize_bars(rows, "count", money=False)
    total = sum(row["count"] for row in rows)
    for row in rows:
        row["share_pct"] = round(row["count"] / total * 100, 1) if total else None
    return rows
