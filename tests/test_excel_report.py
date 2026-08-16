import io
from datetime import date

from openpyxl import load_workbook

from app import create_app, excel_report, passport as passport_module, storage
from app.excel_report import (
    ROW_CONTRACT_TOTAL, ROW_GRAND_TOTAL, ROW_NAME, ROW_PER_SQM, ROW_TOTAL_AREA,
    ROW_VAT, ROW_YEAR,
)


def _make_project_with_passport(root, name, **fields):
    slug = storage.create_project(root, name)
    data = {
        "project_name": name, "address": None, "year_signed": None,
        "building_class": None, "general_contractor": None,
        "contract_price_rub": None, "underground_area_sqm": None,
        "aboveground_area_sqm": None, "total_area_sqm": None, "ocr_fields": [],
    }
    data.update(fields)
    passport_module.save_passport(data, storage.passport_path(root, slug))
    return slug


def _passport(name, **fields):
    data = {"project_name": name}
    data.update(fields)
    return data


def _build(projects, tmp_path, filename="report.xlsx"):
    out = tmp_path / filename
    excel_report.build_comparison_report(projects, out)
    return load_workbook(out, data_only=False).active


def test_report_builds_for_a_single_project(tmp_path):
    ws = _build([_passport(
        "MIRA", year_signed="2025", building_class="Бизнес",
        general_contractor="ООО «АНТТЕК»", underground_area_sqm=13341.3,
        aboveground_area_sqm=62399.7, total_area_sqm=75740.9,
        contract_price_rub=10067050887.72,
    )], tmp_path)

    assert ws["B3"].value == "Проекты"
    assert ws["E3"].value == "MIRA"
    assert ws[f"B{ROW_YEAR}"].value == "Год подписания договора"
    assert ws["E6"].value == "Бизнес"
    assert ws["E7"].value == "ООО «АНТТЕК»"
    assert ws[f"E{ROW_TOTAL_AREA}"].value == 75740.9
    assert ws["C14"].value == "Стоимость по видам работ/расход:"
    assert ws["E14"].value.startswith("ГП")
    assert ws["C30"].value.startswith("Итого СМР")
    assert ws["C34"].value.startswith("Итого СМР, в тч Отделка")


def test_report_puts_each_project_in_its_own_column_pair(tmp_path):
    names = ["П1", "П2", "П3", "П4", "П5"]
    ws = _build([_passport(name, total_area_sqm=1000.0) for name in names], tmp_path)

    # E, H, K, N, Q — three columns per project, two of them used.
    assert [ws.cell(row=ROW_NAME, column=5 + 3 * i).value for i in range(5)] == names
    assert ws.cell(row=ROW_NAME, column=5 + 3 * 4).value == "П5"


def test_report_survives_a_project_with_empty_fields(tmp_path):
    ws = _build([_passport("Пустой")], tmp_path)

    assert ws["E3"].value == "Пустой"
    assert ws["E7"].value is None                      # генподрядчик не найден
    assert ws[f"E{ROW_TOTAL_AREA}"].value is None      # общая площадь не найдена
    assert ws[f"E{ROW_GRAND_TOTAL}"].value.startswith("=")


def test_empty_values_are_highlighted_rather_than_zero_filled(tmp_path):
    ws = _build([_passport("Пустой", total_area_sqm=1000.0)], tmp_path)

    contractor = ws["E7"]
    assert contractor.value is None
    assert contractor.fill.fgColor.rgb.endswith(excel_report.COLOR_EMPTY)
    # A value that was found keeps the plain background.
    total_area = ws[f"E{ROW_TOTAL_AREA}"]
    assert total_area.value == 1000.0
    assert not str(total_area.fill.fgColor.rgb).endswith(excel_report.COLOR_EMPTY)


def test_total_area_is_summed_by_formula_when_only_the_parts_are_known(tmp_path):
    ws = _build([_passport(
        "П", underground_area_sqm=100.0, aboveground_area_sqm=200.0,
    )], tmp_path)

    assert ws[f"E{ROW_TOTAL_AREA}"].value == "=+E8+E9"


def test_formulas_are_saved_as_formulas(tmp_path):
    ws = _build([_passport(
        "П", total_area_sqm=1000.0, contract_price_rub=2000000.0,
    )], tmp_path)

    # "Стоимость на 1 м² ЖК" must recalculate in Excel, not carry a number
    # this program worked out once.
    assert ws[f"F{ROW_CONTRACT_TOTAL}"].value.startswith("=")
    assert f"$E${ROW_TOTAL_AREA}" in ws[f"F{ROW_CONTRACT_TOTAL}"].value
    assert ws["E30"].value.startswith("=SUBTOTAL(9,")
    assert ws[f"E{ROW_CONTRACT_TOTAL}"].value.startswith("=")


def test_contract_price_shows_in_the_totals_while_the_cost_lines_are_empty(tmp_path):
    ws = _build([_passport(
        "П", total_area_sqm=1000.0, contract_price_rub=2000000.0,
    )], tmp_path)

    grand_total = ws[f"E{ROW_GRAND_TOTAL}"].value
    assert grand_total.startswith("=IF(")
    assert "2000000.00" in grand_total


def test_deviation_from_the_first_project_is_a_formula_on_the_others(tmp_path):
    ws = _build([
        _passport("П1", total_area_sqm=1000.0, contract_price_rub=1000.0),
        _passport("П2", total_area_sqm=1000.0, contract_price_rub=2000.0),
    ], tmp_path)

    assert ws[f"F{ROW_PER_SQM}"].value is None          # первому не с чем сравнивать
    assert ws[f"I{ROW_PER_SQM}"].value.startswith("=")
    assert f"$E${ROW_PER_SQM}" in ws[f"I{ROW_PER_SQM}"].value


def test_number_formats_follow_the_customer_template(tmp_path):
    ws = _build([_passport("П", underground_area_sqm=1.0, contract_price_rub=1.0)], tmp_path)

    assert ws["E8"].number_format == '#,##0.0" м2"'
    assert ws[f"E{ROW_CONTRACT_TOTAL}"].number_format == '#,##0\\ "₽"'
    assert ws[f"F{ROW_CONTRACT_TOTAL}"].number_format == '#,##0" ₽/m2"'
    assert ws["E3"].font.name == "Arial"


def test_contract_terms_reach_the_sheet(tmp_path):
    ws = _build([_passport(
        "П", smr_term="30 месяцев", advance_payment="15%",
        bank_guarantee="Включено", performance_bond_pct="5%",
    )], tmp_path)

    assert ws["B36"].value == "Паспорт договора"
    assert ws["B37"].value == "Срок СМР"
    assert ws["E37"].value == "30 месяцев"
    assert ws["E40"].value == "5%"


# --- normalize_passport ----------------------------------------------------


def test_normalize_reads_a_russian_number_written_as_text():
    data = excel_report.normalize_passport(
        _passport("П", contract_price_rub="10 067 050 887,72")
    )

    assert data["contract_price_rub"] == 10067050887.72


def test_normalize_reads_areas_written_as_text():
    data = excel_report.normalize_passport(
        _passport("П", total_area_sqm="67 413", underground_area_sqm="13 341,3")
    )

    assert data["total_area_sqm"] == 67413.0
    assert data["underground_area_sqm"] == 13341.3


def test_normalize_accepts_a_bare_year():
    data = excel_report.normalize_passport(_passport("П", year_signed="2025"))

    assert data["year_signed"] == 2025
    assert data["signing_date"] is None


def test_normalize_accepts_dates_in_different_formats():
    for value, expected in [
        ("20.02.2025", date(2025, 2, 20)),
        ("2025-02-20", date(2025, 2, 20)),
        ("20/02/2025", date(2025, 2, 20)),
        (date(2025, 2, 20), date(2025, 2, 20)),
    ]:
        data = excel_report.normalize_passport(_passport("П", year_signed=value))
        assert data["signing_date"] == expected, value
        assert data["year_signed"] == 2025, value


def test_normalize_finds_a_year_inside_a_longer_phrase():
    data = excel_report.normalize_passport(_passport("П", year_signed="от 2024 г."))

    assert data["year_signed"] == 2024


def test_normalize_derives_vat_from_the_signing_year():
    assert excel_report.normalize_passport(_passport("П", year_signed="2024"))["vat"] == "20%"
    assert excel_report.normalize_passport(_passport("П", year_signed="2025"))["vat"] == "20%"
    assert excel_report.normalize_passport(_passport("П", year_signed="2026"))["vat"] == "22%"
    assert excel_report.normalize_passport(_passport("П", year_signed="2030"))["vat"] == "22%"


def test_normalize_keeps_a_stored_vat_when_no_year_is_known():
    data = excel_report.normalize_passport(_passport("П", vat="без НДС"))

    assert data["vat"] == "без НДС"


def test_normalize_unwraps_a_value_stored_with_its_source():
    data = excel_report.normalize_passport(_passport(
        "П", general_contractor={"value": "ООО «АНТТЕК»", "source": "ai"},
        total_area_sqm={"value": "1 000", "source": "ocr"},
    ))

    assert data["general_contractor"] == "ООО «АНТТЕК»"
    assert data["total_area_sqm"] == 1000.0


def test_normalize_leaves_missing_fields_empty():
    data = excel_report.normalize_passport(_passport("П"))

    assert data["general_contractor"] is None
    assert data["total_area_sqm"] is None
    assert data["contract_price_rub"] is None


def test_building_an_empty_report_is_refused_with_an_explanation(tmp_path):
    try:
        excel_report.build_comparison_report([], tmp_path / "report.xlsx")
    except excel_report.ExcelReportError as e:
        assert "Не выбрано ни одного проекта" in str(e)
    else:
        raise AssertionError("пустой отчёт должен быть отклонён")


# --- load_project ----------------------------------------------------------


def test_load_project_reads_the_passport_and_finds_the_cover(tmp_path):
    slug = _make_project_with_passport(tmp_path, "ПроектА", total_area_sqm=1000.0)
    cover = storage.project_dir(tmp_path, slug) / "cover.png"
    cover.write_bytes(b"not really a png")

    project = excel_report.load_project(storage.project_dir(tmp_path, slug))

    assert project["passport"]["project_name"] == "ПроектА"
    assert project["cover"] == cover


def test_load_project_explains_a_missing_project(tmp_path):
    try:
        excel_report.load_project(tmp_path / "нет-такого")
    except excel_report.ExcelReportError as e:
        assert "не найден" in str(e)
    else:
        raise AssertionError("несуществующий проект должен быть отклонён")


def test_load_project_explains_a_broken_passport(tmp_path):
    slug = _make_project_with_passport(tmp_path, "ПроектА")
    storage.passport_path(tmp_path, slug).write_text("{не json", encoding="utf-8")

    try:
        excel_report.load_project(storage.project_dir(tmp_path, slug))
    except excel_report.ExcelReportError as e:
        assert "повреждён" in str(e)
    else:
        raise AssertionError("повреждённый паспорт должен быть отклонён")


def test_an_unreadable_cover_does_not_break_the_export(tmp_path):
    slug = _make_project_with_passport(tmp_path, "ПроектА", total_area_sqm=1000.0)
    (storage.project_dir(tmp_path, slug) / "cover.png").write_bytes(b"not really a png")

    out = tmp_path / "report.xlsx"
    excel_report.build_comparison_report(
        [excel_report.load_project(storage.project_dir(tmp_path, slug))], out,
    )

    assert load_workbook(out).active["E3"].value == "ПроектА"


# --- the route -------------------------------------------------------------


def test_excel_route_returns_a_workbook(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(
        tmp_path, "ПроектА", contract_price_rub=100.0, year_signed="2024", total_area_sqm=10.0,
    )
    slug2 = _make_project_with_passport(
        tmp_path, "ПроектБ", contract_price_rub=200.0, year_signed="2026", total_area_sqm=20.0,
    )

    resp = client.get(f"/report/excel?slug={slug1}&slug={slug2}")

    assert resp.status_code == 200
    assert resp.content_type.startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["Content-Disposition"]
    assert resp.data[:2] == b"PK"

    ws = load_workbook(io.BytesIO(resp.data), data_only=False).active
    assert ws["E3"].value == "ПроектА"
    assert ws["H3"].value == "ПроектБ"
    assert ws[f"E{ROW_VAT}"].value == "20%"
    assert ws[f"H{ROW_VAT}"].value == "22%"


def test_excel_route_without_a_selection_explains_itself(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.get("/report/excel")

    assert resp.status_code == 400
    assert "Не выбрано ни одного проекта" in resp.data.decode("utf-8")


def test_excel_route_reports_projects_that_no_longer_exist(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.get("/report/excel?slug=нет-такого")

    assert resp.status_code == 400
    assert "не найдены" in resp.data.decode("utf-8")


def test_excel_route_reports_a_broken_passport(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")
    storage.passport_path(tmp_path, slug).write_text("{не json", encoding="utf-8")

    resp = client.get(f"/report/excel?slug={slug}")

    assert resp.status_code == 400
    assert "повреждён" in resp.data.decode("utf-8")


def test_index_page_offers_the_excel_export(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get("/").data.decode("utf-8")

    assert "Выгрузить в Excel" in body
    assert "/report/excel" in body


def test_compare_page_offers_the_excel_export(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get(f"/compare?slug={slug}").data.decode("utf-8")

    assert "Выгрузить в Excel" in body
    assert "Сохранить в PDF" in body


# --- строки затрат из сметы ---


def test_report_row_keys_line_up_with_their_labels():
    # The totals are matched to rows by position, so a key list that drifts
    # from its labels would put a facade's cost on the roof line.
    assert len(excel_report.WORK_KEYS) == len(excel_report.WORK_LABELS)
    assert len(excel_report.MR_KEYS) == len(excel_report.MR_LABELS)
    assert excel_report.WORK_KEYS[0] == "rd"
    assert excel_report.MR_KEYS == ["mr_base", "mr_ready", "shell_core"]


def test_estimate_totals_fill_the_cost_lines(tmp_path):
    costs = {"rd": 192801180.0, "facade": 3034129955.4, "lifts": 220000000.0}
    ws = _build([{
        "passport": _passport("П", total_area_sqm=67413.0),
        "cover": None,
        "costs": costs,
    }], tmp_path)

    assert ws["E16"].value == 192801180.0            # Разработка стадии "Р"
    assert ws["E22"].value == 3034129955.4           # Фасад
    assert ws["E25"].value == 220000000.0            # Лифты
    # The totals stay formulas: they must follow the figures, not freeze them.
    assert ws["E30"].value.startswith("=SUBTOTAL(9,")
    assert ws["F22"].value.startswith("=")


def test_mr_lines_are_filled_from_the_estimate_too(tmp_path):
    ws = _build([{
        "passport": _passport("П", total_area_sqm=1000.0),
        "cover": None,
        "costs": {"mr_base": 1900000000.0},
    }], tmp_path)

    assert ws["E31"].value == 1900000000.0
    assert ws["E32"].value is None


def test_a_cost_line_the_estimate_lacks_is_highlighted(tmp_path):
    ws = _build([{
        "passport": _passport("П", total_area_sqm=1000.0),
        "cover": None,
        "costs": {"facade": 100.0},
    }], tmp_path)

    assert ws["E22"].value == 100.0
    missing = ws["E25"]                               # Лифты — в смете нет
    assert missing.value is None
    assert missing.fill.fgColor.rgb.endswith(excel_report.COLOR_EMPTY)


def test_without_an_estimate_the_cost_lines_stay_plain(tmp_path):
    # Nothing to compare against: every line is empty by design and meant to
    # be typed in, so marking all fourteen would say nothing.
    ws = _build([_passport("П", total_area_sqm=1000.0)], tmp_path)

    for row in (16, 22, 25, 29):
        cell = ws.cell(row=row, column=5)
        assert cell.value is None
        assert not str(cell.fill.fgColor.rgb).endswith(excel_report.COLOR_EMPTY)


def test_a_zero_cost_line_is_written_as_zero(tmp_path):
    ws = _build([{
        "passport": _passport("П", total_area_sqm=1000.0),
        "cover": None,
        "costs": {"landscaping": 0.0},
    }], tmp_path)

    landscaping = ws["E27"]
    assert landscaping.value == 0
    assert not str(landscaping.fill.fgColor.rgb).endswith(excel_report.COLOR_EMPTY)


def test_load_project_reads_the_estimate_when_there_is_one(tmp_path):
    from openpyxl import Workbook

    slug = _make_project_with_passport(tmp_path, "Со сметой")
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="№ п/п")
    ws.cell(row=1, column=2, value="№ раздела")
    ws.cell(row=1, column=3, value="Статья СМР")
    ws.cell(row=1, column=4, value="Наименование работ")
    ws.cell(row=1, column=5, value="Стоимость всего, RUB")
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
    ws.cell(row=2, column=6, value="Всего")
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value=1)
    ws.cell(row=3, column=3, value="6. Фасадные работы")
    ws.cell(row=3, column=6, value=555.0)
    wb.save(storage.estimate_path(tmp_path, slug))

    project = excel_report.load_project(storage.project_dir(tmp_path, slug))

    assert project["costs"] == {"facade": 555.0}


def test_load_project_without_an_estimate_has_no_costs(tmp_path):
    slug = _make_project_with_passport(tmp_path, "Без сметы")

    project = excel_report.load_project(storage.project_dir(tmp_path, slug))

    assert project["costs"] == {}


def test_an_unreadable_estimate_does_not_stop_the_export(tmp_path):
    slug = _make_project_with_passport(tmp_path, "Битая смета")
    storage.estimate_path(tmp_path, slug).write_bytes(b"not a workbook")

    project = excel_report.load_project(storage.project_dir(tmp_path, slug))

    assert project["costs"] == {}
    assert project["passport"]["project_name"] == "Битая смета"


def test_each_filled_cost_line_says_which_sections_it_came_from(tmp_path):
    # Placing a section on a line is a judgement about wording; the one place
    # it can be checked is next to the number.
    ws = _build([{
        "passport": _passport("П", total_area_sqm=1000.0),
        "cover": None,
        "costs": {"other": 12.5},
        "cost_sources": {"other": ["99. Прочее", "Дополнительные работы"]},
    }], tmp_path)

    comment = ws["E29"].comment
    assert comment is not None
    assert "99. Прочее" in comment.text
    assert "Дополнительные работы" in comment.text


def test_an_empty_cost_line_carries_no_note(tmp_path):
    ws = _build([{
        "passport": _passport("П", total_area_sqm=1000.0),
        "cover": None,
        "costs": {"facade": 1.0},
        "cost_sources": {"facade": ["6. Фасадные работы"]},
    }], tmp_path)

    assert ws["E22"].comment is not None
    assert ws["E25"].comment is None


def test_load_project_reports_the_sections_behind_each_line(tmp_path):
    from openpyxl import Workbook

    slug = _make_project_with_passport(tmp_path, "Со сметой")
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="№ п/п")
    ws.cell(row=1, column=2, value="№ раздела")
    ws.cell(row=1, column=3, value="Статья СМР")
    ws.cell(row=1, column=4, value="Наименование работ")
    ws.cell(row=1, column=5, value="Стоимость всего, RUB")
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
    ws.cell(row=2, column=6, value="Всего")
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value=1)
    ws.cell(row=3, column=3, value="99. Прочее")
    ws.cell(row=3, column=6, value=10.0)
    ws.cell(row=4, column=4, value="Дополнительные работы")
    ws.cell(row=4, column=6, value=2.5)
    wb.save(storage.estimate_path(tmp_path, slug))

    project = excel_report.load_project(storage.project_dir(tmp_path, slug))

    assert project["costs"] == {"other": 12.5}
    assert project["cost_sources"] == {"other": ["99. Прочее", "Дополнительные работы"]}
