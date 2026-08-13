import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

from app import estimate


def _save(tmp_path, wb, filename="smeta.xlsx"):
    path = tmp_path / filename
    wb.save(path)
    return path


def test_read_estimate_simple_values(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Раздел", "Кол-во", "Цена"])
    ws.append(["Земляные работы", 10, 1500.5])
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)

    assert len(sheets) == 1
    rows = sheets[0]["rows"]
    assert [c["value"] for c in rows[0]] == ["Раздел", "Кол-во", "Цена"]
    assert [c["value"] for c in rows[1]] == ["Земляные работы", "10", "1 500.50"]


def test_read_estimate_merged_cells_collapse_into_one(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Заголовок"
    ws.merge_cells("A1:C1")
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)

    row = sheets[0]["rows"][0]
    assert len(row) == 1
    assert row[0]["value"] == "Заголовок"
    assert row[0]["colspan"] == 3
    assert row[0]["rowspan"] == 1


def test_read_estimate_multiple_sheets_in_order(tmp_path):
    wb = Workbook()
    wb.active.title = "Смета"
    wb.create_sheet("Материалы")
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)

    assert [s["name"] for s in sheets] == ["Смета", "Материалы"]


def test_read_estimate_skips_hidden_sheet(tmp_path):
    wb = Workbook()
    wb.active.title = "Видимый"
    hidden = wb.create_sheet("Скрытый")
    hidden.sheet_state = "hidden"
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)

    assert [s["name"] for s in sheets] == ["Видимый"]


def test_read_estimate_formula_cell_never_shows_raw_formula_text(tmp_path):
    """openpyxl never evaluates formulas — a workbook it just created has no
    cached value for a formula cell, so it must render blank rather than
    leaking the raw "=..." text. (A real Excel-saved file *does* carry a
    cached value, which openpyxl's own data_only mode is trusted to read —
    that part isn't ours to re-test.)"""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = "=A1*2"
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)

    assert sheets[0]["rows"][1][0]["value"] == ""


def test_read_estimate_captures_basic_formatting(tmp_path):
    wb = Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = "Итого"
    cell.font = Font(bold=True, italic=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    rendered = sheets[0]["rows"][0][0]

    assert rendered["value"] == "Итого"
    assert rendered["bold"] is True
    assert rendered["italic"] is True
    assert rendered["bg"] == "#FF0000"
    assert rendered["align"] == "center"
    assert rendered["border_top"] is True
    assert rendered["border_bottom"] is True
    assert rendered["border_left"] is False


def test_read_estimate_reports_column_widths(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B1"] = "y"
    ws.column_dimensions["B"].width = 40
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    widths = sheets[0]["col_widths"]

    assert len(widths) == 2
    assert widths[1] == 40


def test_read_estimate_raises_on_corrupted_file(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not a real xlsx file")

    with pytest.raises(estimate.EstimateReadError):
        estimate.read_estimate(path)


def test_read_estimate_indexed_color_fill(tmp_path):
    """Test that indexed color fills (from standard palette) are resolved to #RRGGBB."""
    wb = Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = "Indexed"
    # Indexed color 10 is a standard red in Excel's palette
    cell.fill = PatternFill(fill_type="solid", fgColor=Color(index=10))
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    rendered = sheets[0]["rows"][0][0]

    assert rendered["value"] == "Indexed"
    # Should resolve to an actual color, not None
    assert rendered["bg"] is not None
    assert isinstance(rendered["bg"], str)
    assert rendered["bg"].startswith("#")


def test_read_estimate_theme_color_fill(tmp_path):
    """Test that theme color fills are resolved to #RRGGBB using the default
    Office theme palette. (Note: uses standard palette approximation, not
    workbook's actual embedded theme.)"""
    wb = Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = "Theme"
    # Theme color 0 from the standard Office default palette
    cell.fill = PatternFill(fill_type="solid", fgColor=Color(theme=0, tint=0))
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    rendered = sheets[0]["rows"][0][0]

    assert rendered["value"] == "Theme"
    # Should resolve to the default palette color for theme index 0
    assert rendered["bg"] is not None
    assert isinstance(rendered["bg"], str)
    # Per the OOXML theme order (lt1, dk1, lt2, dk2, accent1..6), theme index
    # 0 is lt1 ("White, Background 1") -> #FFFFFF.
    assert rendered["bg"] == "#FFFFFF"


def test_read_estimate_theme_color_1_is_black(tmp_path):
    """Theme index 1 is dk1 ("Black, Text 1") -> #000000."""
    wb = Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = "Theme"
    cell.fill = PatternFill(fill_type="solid", fgColor=Color(theme=1, tint=0))
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    rendered = sheets[0]["rows"][0][0]

    assert rendered["bg"] == "#000000"


def test_read_estimate_theme_color_accent_unaffected(tmp_path):
    """Accent indices (4..9) are unaffected by the lt1/dk1/lt2/dk2 reorder."""
    wb = Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = "Theme"
    # Theme index 4 is accent1 -> #4472C4, unchanged by the fix.
    cell.fill = PatternFill(fill_type="solid", fgColor=Color(theme=4, tint=0))
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    rendered = sheets[0]["rows"][0][0]

    assert rendered["bg"] == "#4472C4"


def test_render_sheet_truncates_oversized_dimensions(tmp_path, monkeypatch):
    """A sheet with real *values* extending past the render cap gets
    clamped, and the returned sheet dict signals truncation."""
    monkeypatch.setattr(estimate, "MAX_RENDERED_ROWS", 5)
    monkeypatch.setattr(estimate, "MAX_RENDERED_COLS", 3)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    # A real value (not just formatting) far outside the cap means genuine
    # content is being cut off.
    ws.cell(row=20, column=10).value = "far"
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    sheet = sheets[0]

    assert sheet["truncated"] is True
    assert len(sheet["rows"]) == 5
    assert all(len(row) <= 3 for row in sheet["rows"])
    assert len(sheet["col_widths"]) == 3


def test_render_sheet_small_sheet_not_truncated(tmp_path):
    """An ordinary small sheet is unaffected by the cap: no truncation flag,
    and it renders its full content as before."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Раздел", "Кол-во", "Цена"])
    ws.append(["Земляные работы", 10, 1500.5])
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    sheet = sheets[0]

    assert "truncated" not in sheet or sheet["truncated"] is False
    assert len(sheet["rows"]) == 2
    assert [c["value"] for c in sheet["rows"][0]] == ["Раздел", "Кол-во", "Цена"]
    assert [c["value"] for c in sheet["rows"][1]] == ["Земляные работы", "10", "1 500.50"]


def test_render_sheet_inflated_dimension_but_modest_content_not_truncated(tmp_path):
    """The exact real-world scenario this fix targets: a stray far-out cell
    with only formatting (no value) inflates ws.max_row/ws.max_column, but
    the real content is a small table. The sheet should render its full real
    content and NOT report truncation, since nothing real was cut off."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Раздел", "Кол-во", "Цена"])
    ws.append(["Земляные работы", 10, 1500.5])
    # Stray formatting-only cell far outside the real table. Before this fix,
    # ws.max_column would report ~300 and the render loop would materialize
    # rows x 200 (the old cap) cells needlessly.
    ws.cell(row=5, column=300).fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    sheet = sheets[0]

    assert "truncated" not in sheet or sheet["truncated"] is False
    assert len(sheet["rows"]) == 2
    assert len(sheet["col_widths"]) == 3
    assert [c["value"] for c in sheet["rows"][0]] == ["Раздел", "Кол-во", "Цена"]


def test_render_sheet_merge_extends_past_last_valued_cell(tmp_path):
    """A merged range whose top-left cell has a value but which extends into
    columns/rows with no values at all must still be fully accounted for in
    the render boundary."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Заголовок"
    # Merge extends to column F (6) and row 3, well past any valued cell.
    ws.merge_cells("A1:F3")
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    sheet = sheets[0]

    assert "truncated" not in sheet or sheet["truncated"] is False
    assert len(sheet["rows"]) == 3
    assert len(sheet["col_widths"]) == 6
    assert sheet["rows"][0][0]["value"] == "Заголовок"
    assert sheet["rows"][0][0]["rowspan"] == 3
    assert sheet["rows"][0][0]["colspan"] == 6


def test_render_sheet_merge_extent_still_capped(tmp_path, monkeypatch):
    """A merge extending past the render cap is still clamped, and reported
    as truncated, even though the merge itself has no genuine cell values
    beyond its top-left."""
    monkeypatch.setattr(estimate, "MAX_RENDERED_ROWS", 5)
    monkeypatch.setattr(estimate, "MAX_RENDERED_COLS", 3)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Заголовок"
    ws.merge_cells("A1:J1")  # 10 columns, past the 3-column cap.
    path = _save(tmp_path, wb)

    sheets = estimate.read_estimate(path)
    sheet = sheets[0]

    assert sheet["truncated"] is True
    assert len(sheet["col_widths"]) == 3
