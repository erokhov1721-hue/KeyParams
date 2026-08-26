import io

from openpyxl import Workbook

from app import create_app
from tests.helpers import build_docx_bytes, document_xml


def _dgp_bytes():
    return build_docx_bytes(document_xml(paragraphs=[
        "Общество с ограниченной ответственностью «Ромашка» (ООО «Ромашка»), "
        "именуемое в дальнейшем «Генподрядчик», с третьей стороны,"
    ]))


def _tz_bytes():
    return build_docx_bytes(document_xml(tables=[[["1", "Площадь подземной части", "м2", "1 000"]]]))


def _smeta_bytes():
    wb = Workbook()
    wb.active.append(["Раздел", "Сумма"])
    wb.active.append(["Фундамент", 1000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_index_page_loads_when_empty(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.get("/")
    assert resp.status_code == 200


def test_new_project_form_loads(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.get("/projects/new")
    assert resp.status_code == 200


def test_the_top_bar_does_not_duplicate_the_sidebar_link(tmp_path):
    # Заводить проект — одна ссылка, в левом меню. Кнопка посреди верхней
    # панели повторяла её и больше ничего не делала.
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Тест")
    storage.passport_path(tmp_path, slug).write_text("{}", encoding="utf-8")

    for path in ("/", "/compare/select", f"/projects/{slug}"):
        body = client.get(path).get_data(as_text=True)
        assert "+ Создать проект" not in body, path
        assert "Добавить проект" in body, path


def test_every_page_offers_the_theme_toggle(tmp_path):
    # The toggle lives in the shared layout, so a page that renders its own
    # header block must not end up without it.
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Тест")
    storage.passport_path(tmp_path, slug).write_text("{}", encoding="utf-8")

    for path in ("/", "/projects/new", "/compare/select", f"/projects/{slug}"):
        body = client.get(path).get_data(as_text=True)
        assert 'id="theme-toggle"' in body, path


def test_native_dropdowns_follow_the_theme(tmp_path):
    # The popup of a <select> is drawn by the browser, not by this
    # stylesheet. With no colour scheme declared it opens on white while its
    # options inherit the field's near-white text, which left the list of
    # building classes unreadable in the dark theme.
    app = create_app(tmp_path)
    client = app.test_client()

    css = client.get("/static/style.css").get_data(as_text=True)

    assert "color-scheme: dark" in css
    assert "color-scheme: light" in css
    assert "select option" in css


def test_the_header_height_is_one_value_used_once(tmp_path):
    # Высота панели не должна зависеть от того, сколько на странице кнопок:
    # min-height или auto поднимали её от содержимого, и полоса гуляла от
    # страницы к странице. Кнопки в панели больше не выравниваются абсолютом —
    # иначе они снова перестанут задавать её высоту и наедут на края.
    app = create_app(tmp_path)
    client = app.test_client()

    css = client.get("/static/style.css").get_data(as_text=True)
    topbar = css.split(".topbar {", 1)[1].split("}", 1)[0]
    actions = css.split(".topbar-actions {", 1)[1].split("}", 1)[0]

    assert "--header-height: 64px" in css
    assert "height: var(--header-height)" in topbar
    assert "min-height" not in topbar
    assert "position: absolute" not in actions
    assert "gap: 12px" in actions


def test_theme_choice_is_applied_before_the_page_paints(tmp_path):
    # The saved theme has to be read in <head>: doing it lower down makes a
    # light-theme user watch the dark theme flash on every single load.
    app = create_app(tmp_path)
    client = app.test_client()

    body = client.get("/").get_data(as_text=True)
    head = body.split("</head>")[0]

    assert "localStorage.getItem('theme')" in head


def test_create_project_rejects_missing_name(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 400


def test_create_project_rejects_non_docx(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Тест",
        "dgp_file": (io.BytesIO(b"not docx"), "dgp.txt"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 400


def test_create_project_rejects_corrupted_docx(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Тест",
        "dgp_file": (io.BytesIO(b"this has a .docx name but is not a real zip"), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 400


def test_create_project_then_view_passport(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Тестовый проект",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 302

    page = client.get(resp.headers["Location"])
    assert page.status_code == 200
    assert "ООО «Ромашка»".encode("utf-8") in page.data


def test_index_lists_created_project(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    client.post("/projects", data={
        "project_name": "Видимый проект",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    resp = client.get("/")
    # Именно название, а не имя папки: в списке стоит то, как проект назвали.
    assert "Видимый проект".encode("utf-8") in resp.data


def test_update_project_saves_manual_field(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    create_resp = client.post("/projects", data={
        "project_name": "Правка",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    project_url = create_resp.headers["Location"]
    slug = "Правка"

    update_resp = client.post(project_url, data={
        "year_signed": "2025",
        "building_class": "Бизнес",
        "general_contractor": "ООО «Ромашка»",
        "underground_area_sqm": "1000",
        "aboveground_area_sqm": "2000",
        "total_area_sqm": "3000",
    })
    assert update_resp.status_code == 302

    from app import passport, storage
    saved = passport.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["building_class"] == "Бизнес"
    assert saved["aboveground_area_sqm"] == 2000.0


def _root_with_passport_above(tmp_path):
    """Projects root with a real passport.json planted one level above it.

    Without the slug whitelist guard, the slug ".." would resolve to that
    planted file, so the route would answer 200 instead of 404.
    """
    root = tmp_path / "projects"
    root.mkdir()
    (tmp_path / "passport.json").write_text(
        '{"project_name": "Похищенный"}', encoding="utf-8"
    )
    return root


def test_path_traversal_blocked_on_get(tmp_path):
    root = _root_with_passport_above(tmp_path)
    assert (root / ".." / "passport.json").exists(), "test setup must be a real repro"

    client = create_app(root).test_client()
    resp = client.get("/projects/..")
    assert resp.status_code == 404


def test_path_traversal_blocked_on_post(tmp_path):
    root = _root_with_passport_above(tmp_path)
    assert (root / ".." / "passport.json").exists(), "test setup must be a real repro"

    client = create_app(root).test_client()
    resp = client.post("/projects/..", data={"building_class": "Бизнес"})
    assert resp.status_code == 404
    # The planted passport must be untouched.
    assert "Похищенный" in (tmp_path / "passport.json").read_text(encoding="utf-8")


def test_create_project_with_name_that_slugifies_to_empty(tmp_path):
    """A name made only of stripped characters is a 400, not a 500."""
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": '???',
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 400

    from app import storage
    assert storage.list_project_slugs(tmp_path) == []


def test_create_project_with_hash_in_name_redirect_is_followable(tmp_path):
    """A '#' in the project name must not truncate the redirect target."""
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Дом #5 100%",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 302

    page = client.get(resp.headers["Location"])
    assert page.status_code == 200


def test_docx_error_message_hides_server_paths(tmp_path):
    """The user-facing error must not leak absolute filesystem paths."""
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Тест",
        "dgp_file": (io.BytesIO(b"this has a .docx name but is not a real zip"), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 400
    body = resp.data.decode("utf-8")
    assert "dgp.docx" not in body
    assert str(tmp_path) not in body


def test_no_orphan_on_corrupted_docx(tmp_path):
    """Verify that a failed project creation doesn't leave an orphaned directory."""
    app = create_app(tmp_path)
    client = app.test_client()
    project_name = "Проверка"

    client.post("/projects", data={
        "project_name": project_name,
        "dgp_file": (io.BytesIO(b"this has a .docx name but is not a real zip"), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")

    from app import storage
    slugs = storage.list_project_slugs(tmp_path)
    assert len(slugs) == 0, "Expected no projects after failed upload"


def test_delete_project_removes_it_and_redirects_to_index(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Тест")
    storage.passport_path(tmp_path, slug).write_text("{}", encoding="utf-8")

    resp = client.post(f"/projects/{slug}/delete", follow_redirects=True)

    assert resp.status_code == 200
    assert slug not in storage.list_project_slugs(tmp_path)
    assert not (tmp_path / slug).exists()


def test_delete_project_unknown_slug_redirects_instead_of_404(tmp_path):
    # Deleting what isn't there already got the user what they wanted, so it
    # sends them back to the dashboard. A double-click on the delete button
    # used to land the second POST on a bare "Not Found" page with no way back.
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects/does-not-exist/delete")
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/")


def test_dashboard_sweeps_up_a_delete_that_could_not_finish(tmp_path):
    # Whatever the user deletes has to end up actually deleted, even when a
    # file was in use at the time and the removal had to be left for later.
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Недоудалённый")
    storage.estimate_path(tmp_path, slug).write_bytes(b"PK\x03\x04")
    (tmp_path / slug / storage.DELETED_MARKER).write_bytes(b"")

    resp = client.get("/")

    assert resp.status_code == 200
    assert not (tmp_path / slug).exists()


def test_rename_project_updates_name_and_redirects_to_index(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Старое имя")
    passport_module.save_passport(
        {"project_name": "Старое имя"}, storage.passport_path(tmp_path, slug),
    )

    resp = client.post(f"/projects/{slug}/rename", data={"project_name": "Новое имя"}, follow_redirects=True)

    assert resp.status_code == 200
    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["project_name"] == "Новое имя"


def test_rename_project_rejects_empty_name(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Старое имя")
    passport_module.save_passport(
        {"project_name": "Старое имя"}, storage.passport_path(tmp_path, slug),
    )

    resp = client.post(f"/projects/{slug}/rename", data={"project_name": "   "})

    assert resp.status_code == 400
    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["project_name"] == "Старое имя"


def test_rename_project_unknown_slug_404s(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects/does-not-exist/rename", data={"project_name": "Новое имя"})
    assert resp.status_code == 404


def test_index_page_shows_project_name_instead_of_slug(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Проспект Мира")
    passport_module.save_passport(
        {"project_name": "Проспект Мира — очень длинное имя"},
        storage.passport_path(tmp_path, slug),
    )

    resp = client.get("/")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "Проспект Мира — очень длинное имя" in body


def _make_project_with_passport(root, name, **fields):
    from app import storage, passport as passport_module

    slug = storage.create_project(root, name)
    data = {
        "project_name": name, "year_signed": None, "building_class": None,
        "general_contractor": None, "underground_area_sqm": None,
        "aboveground_area_sqm": None, "total_area_sqm": None, "ocr_fields": [],
    }
    data.update(fields)
    passport_module.save_passport(data, storage.passport_path(root, slug))
    return slug


def test_compare_projects_shows_selected_projects_side_by_side(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(tmp_path, "ПроектА", total_area_sqm=1000.0)
    slug2 = _make_project_with_passport(tmp_path, "ПроектБ", total_area_sqm=2000.0)

    resp = client.get(f"/compare?slug={slug1}&slug={slug2}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "ПроектА" in body
    assert "ПроектБ" in body
    assert "1 000" in body
    assert "2 000" in body


def test_compare_projects_table_shows_price_per_sqm_row(tmp_path):
    from app import passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(
        tmp_path, "ПроектА", contract_price_rub=10067050887.72, total_area_sqm=67413.0,
    )

    resp = client.get(f"/compare?slug={slug}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "Цена за м²" in body
    expected = passport_module.format_number(10067050887.72 / 67413.0)
    assert expected in body


def test_compare_projects_table_shows_dash_for_price_per_sqm_when_area_missing(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА", contract_price_rub=100.0)

    resp = client.get(f"/compare?slug={slug}")

    assert resp.status_code == 200
    assert "Цена за м²" in resp.data.decode("utf-8")


def test_compare_projects_ignores_unknown_slug(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    resp = client.get(f"/compare?slug={slug}&slug=does-not-exist")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "ПроектА" in body
    assert "does-not-exist" not in body


def test_compare_projects_shows_price_charts(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(
        tmp_path, "ПроектА", contract_price_rub=100.0, year_signed="2024",
        building_class="Бизнес", total_area_sqm=1.0,
    )
    slug2 = _make_project_with_passport(
        tmp_path, "ПроектБ", contract_price_rub=200.0, year_signed="2023",
        building_class="Комфорт", total_area_sqm=1.0,
    )

    resp = client.get(f"/compare?slug={slug1}&slug={slug2}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "Цена работ по году подписания" in body
    assert "Цена работ по классу жилья" in body
    assert "Цена работ по проектам" in body
    assert "Цена за м² по проектам" in body
    assert "ПроектА (2024)" in body
    assert "ПроектБ (Комфорт)" in body


def test_compare_projects_shows_empty_chart_message_without_data(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    resp = client.get(f"/compare?slug={slug}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "Недостаточно данных" in body


def test_compare_projects_price_charts_share_one_card_with_tabs(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(
        tmp_path, "ПроектА", contract_price_rub=100.0, year_signed="2024",
        building_class="Бизнес", total_area_sqm=1.0,
    )
    slug2 = _make_project_with_passport(
        tmp_path, "ПроектБ", contract_price_rub=200.0, year_signed="2023",
        building_class="Комфорт", total_area_sqm=1.0,
    )

    body = client.get(f"/compare?slug={slug1}&slug={slug2}").get_data(as_text=True)

    assert 'data-chart-tab="price"' in body
    assert 'data-chart-tab="price_by_year"' in body
    assert 'data-chart-tab="price_by_class"' in body
    # Полное описание переехало в подсказку вкладки, а не пропало совсем.
    assert 'title="Цена работ по году подписания договора"' in body


def test_compare_projects_shows_a_color_legend_for_each_project(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(tmp_path, "ПроектА")
    slug2 = _make_project_with_passport(tmp_path, "ПроектБ")

    body = client.get(f"/compare?slug={slug1}&slug={slug2}").get_data(as_text=True)

    assert "project-legend" in body
    assert "#059669" in body
    assert "#4f46e5" in body


def test_compare_projects_price_per_sqm_renders_as_a_bar_chart_like_the_rest(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(
        tmp_path, "ПроектА", contract_price_rub=100.0, total_area_sqm=1.0,
    )
    slug2 = _make_project_with_passport(
        tmp_path, "ПроектБ", contract_price_rub=150.0, total_area_sqm=1.0,
    )

    body = client.get(f"/compare?slug={slug1}&slug={slug2}").get_data(as_text=True)

    assert "kpi-tile" not in body
    assert "150 ₽" in body
    assert "100 ₽" in body


def test_compare_page_shows_the_terms_table(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _make_project_with_passport(
        tmp_path, "ПроектА", smr_term="33 (тридцать три месяца)",
        performance_bond_pct="3%",
    )
    b = _make_project_with_passport(tmp_path, "ПроектБ", vat="20%")

    body = client.get(f"/compare?slug={a}&slug={b}").data.decode("utf-8")

    assert "Условия" in body
    assert "Срок СМР" in body
    assert "33 (тридцать три месяца)" in body
    assert "Performance bond, %" in body


def test_compare_page_hides_the_terms_table_without_a_protocol(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _make_project_with_passport(tmp_path, "ПроектА")
    b = _make_project_with_passport(tmp_path, "ПроектБ")

    body = client.get(f"/compare?slug={a}&slug={b}").data.decode("utf-8")

    assert "Срок СМР" not in body


def test_compare_pdf_carries_the_terms_table(tmp_path):
    # PDF повторяет страницу целиком, поэтому «Условия» должны дойти и до него.
    from app import routes

    seen = {}
    original = routes.pdf_export.build_compare_pdf

    def spy(*args, **kwargs):
        seen.update(kwargs)
        return original(*args, **kwargs)

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА", smr_term="33 мес")
    routes.pdf_export.build_compare_pdf = spy
    try:
        resp = client.get(f"/compare/pdf?slug={slug}")
    finally:
        routes.pdf_export.build_compare_pdf = original

    assert resp.status_code == 200
    assert seen["terms"]["rows"][0]["cells"] == ["33 мес"]


def test_compare_projects_pdf_returns_pdf_file(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(
        tmp_path, "ПроектА", contract_price_rub=100.0, year_signed="2024", total_area_sqm=1.0,
    )
    slug2 = _make_project_with_passport(
        tmp_path, "ПроектБ", contract_price_rub=200.0, year_signed="2023", total_area_sqm=1.0,
    )

    resp = client.get(f"/compare/pdf?slug={slug1}&slug={slug2}")

    assert resp.status_code == 200
    assert resp.content_type == "application/pdf"
    assert "attachment" in resp.headers["Content-Disposition"]
    assert resp.data[:4] == b"%PDF"


def test_compare_projects_pdf_redirects_to_the_selection_when_none_chosen(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.get("/compare/pdf", follow_redirects=True)

    assert resp.status_code == 200
    assert resp.request.path == "/compare/select"


def test_compare_projects_redirects_to_the_selection_when_none_chosen(tmp_path):
    # Нечего сравнивать — значит, проекты ещё не выбраны, и вести надо туда,
    # где их выбирают, а не на общий список.
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.get("/compare", follow_redirects=True)

    assert resp.status_code == 200
    assert resp.request.path == "/compare/select"


# --- выбор проектов для сравнения живёт на своей странице ---


def test_compare_select_page_lists_projects_with_checkboxes(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    resp = client.get("/compare/select")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "ПроектА" in body
    assert f'name="slug" value="{slug}"' in body


def test_compare_select_page_offers_comparison_and_excel(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get("/compare/select").data.decode("utf-8")

    assert "Сравнить проекты" in body
    assert 'action="/compare"' in body
    assert "/report/excel" in body


def test_compare_select_page_narrows_the_list_by_filter(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    _make_project_with_passport(tmp_path, "Старый", year_signed="2024")
    _make_project_with_passport(tmp_path, "Новый", year_signed="2026")

    body = client.get("/compare/select?year=2026").data.decode("utf-8")

    assert "Новый" in body
    assert 'value="Старый"' not in body


def test_index_page_leaves_comparison_to_its_own_page(tmp_path):
    # Две задачи — заведение проектов и их сравнение — разведены по разным
    # страницам. Галочки и кнопка сравнения на общем списке означали бы, что
    # разделения не случилось.
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get("/").data.decode("utf-8")

    assert "ПроектА" in body
    assert "Сравнить проекты" not in body
    assert f'name="slug" value="{slug}"' not in body


def test_sidebar_leads_to_both_adding_and_comparing(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get("/").data.decode("utf-8")

    assert "Добавить проект" in body
    assert "Сравнить объекты" in body
    assert 'href="/projects/new"' in body
    assert 'href="/compare/select"' in body


def test_compare_select_page_offers_renaming_and_deleting(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get("/compare/select").data.decode("utf-8")

    assert "btn-edit" in body
    assert "btn-delete" in body
    assert f"/rename" in body
    assert f"/delete" in body


def test_index_keeps_its_row_actions(tmp_path):
    # Кнопки в строках общего списка — те же самые, и фото правится только там.
    app = create_app(tmp_path)
    client = app.test_client()
    _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get("/").data.decode("utf-8")

    assert "btn-edit" in body
    assert "btn-delete" in body
    assert "btn-cover-upload" in body


def test_compare_select_page_has_no_cover_button(tmp_path):
    # На выборе проектов нужны только карандаш и урна. Проверяется сама
    # кнопка, а не имя класса: обработчик фото лежит в общем скрипте обеих
    # страниц и упоминает его безобидно.
    app = create_app(tmp_path)
    client = app.test_client()
    _make_project_with_passport(tmp_path, "ПроектА")

    body = client.get("/compare/select").data.decode("utf-8")

    assert "btn-cover-upload" not in body


def test_deleting_from_the_selection_returns_to_the_selection(tmp_path):
    # Удалил проект, выбирая, что с чем сравнить, — и остался там же, а не
    # уехал на общий список.
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    resp = client.post(f"/projects/{slug}/delete", data={"back": "select"})

    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/compare/select")
    assert slug not in storage.list_project_slugs(tmp_path)


def test_renaming_from_the_selection_returns_to_the_selection(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    resp = client.post(
        f"/projects/{slug}/rename", data={"project_name": "Новое", "back": "select"},
    )

    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/compare/select")


def test_an_unknown_return_page_leads_to_the_project_list(tmp_path):
    # В форму можно подставить что угодно, поэтому «куда вернуться» — это
    # название страницы из короткого списка, а не адрес: увести человека на
    # чужой сайт так нельзя.
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")

    resp = client.post(
        f"/projects/{slug}/rename",
        data={"project_name": "Новое", "back": "https://example.com/"},
    )

    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/")


def test_project_page_shows_a_long_value_in_full_on_hover(tmp_path):
    # Условия из протокола приходят строкой длиной с абзац, а поле узкое:
    # видно только начало. Окошко с полным текстом и его обработчик страница
    # обязана нести — иначе подсказка тихо исчезнет при правке шаблона.
    long_value = (
        "Срок выполнения работ, мес.: 33 (тридцать три месяца) до момента "
        "получения разрешения на ввод объекта в эксплуатацию"
    )
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА", smr_term=long_value)
    # Паспорт договора показывается только там, где протокол загружен.
    storage.contract_terms_path(tmp_path, slug).write_bytes(b"%PDF-fake")

    body = client.get(f"/projects/{slug}").data.decode("utf-8")

    assert 'class="value-hint"' in body
    assert "mouseenter" in body
    assert long_value in body


def test_project_page_flags_ocr_filled_field(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "ОКР проект")
    passport_module.save_passport({
        "project_name": "ОКР проект",
        "year_signed": None,
        "building_class": None,
        "general_contractor": None,
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": 67413.0,
        "ocr_fields": ["total_area_sqm"],
    }, storage.passport_path(tmp_path, slug))

    resp = client.get(f"/projects/{slug}")
    assert resp.status_code == 200
    assert "С картинки".encode("utf-8") in resp.data


def test_project_page_shows_numeric_field_with_thousands_spaces(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Проект с суммой")
    passport_module.save_passport({
        "project_name": "Проект с суммой", "year_signed": None, "building_class": None,
        "general_contractor": None, "contract_price_rub": 10067050887.72,
        "underground_area_sqm": None, "aboveground_area_sqm": None,
        "total_area_sqm": None, "ocr_fields": [],
    }, storage.passport_path(tmp_path, slug))

    resp = client.get(f"/projects/{slug}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert 'value="10 067 050 887.72"' in body


def test_project_page_building_class_is_a_dropdown_with_fixed_options(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Проект без класса")
    passport_module.save_passport({
        "project_name": "Проект без класса", "year_signed": None, "building_class": None,
        "general_contractor": None, "contract_price_rub": None,
        "underground_area_sqm": None, "aboveground_area_sqm": None,
        "total_area_sqm": None, "ocr_fields": [],
    }, storage.passport_path(tmp_path, slug))

    resp = client.get(f"/projects/{slug}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert '<select name="building_class">' in body
    for option in passport_module.BUILDING_CLASS_OPTIONS:
        assert f">{option}<" in body


def test_project_page_building_class_dropdown_preselects_current_value(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Проект с классом")
    passport_module.save_passport({
        "project_name": "Проект с классом", "year_signed": None, "building_class": "Бизнес",
        "general_contractor": None, "contract_price_rub": None,
        "underground_area_sqm": None, "aboveground_area_sqm": None,
        "total_area_sqm": None, "ocr_fields": [],
    }, storage.passport_path(tmp_path, slug))

    resp = client.get(f"/projects/{slug}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert '<option value="Бизнес" selected>Бизнес</option>' in body


def test_project_page_shows_computed_price_per_sqm(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Проект с ценой")
    passport_module.save_passport({
        "project_name": "Проект с ценой",
        "year_signed": None,
        "building_class": None,
        "general_contractor": None,
        "contract_price_rub": 10067050887.72,
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": 67413.0,
        "ocr_fields": [],
    }, storage.passport_path(tmp_path, slug))

    resp = client.get(f"/projects/{slug}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "Цена за м²" in body
    expected = passport_module.format_number(10067050887.72 / 67413.0)
    assert expected in body


def test_project_page_shows_dash_for_price_per_sqm_when_area_missing(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "Проект без площади")
    passport_module.save_passport({
        "project_name": "Проект без площади",
        "year_signed": None,
        "building_class": None,
        "general_contractor": None,
        "contract_price_rub": 10067050887.72,
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": None,
        "ocr_fields": [],
    }, storage.passport_path(tmp_path, slug))

    resp = client.get(f"/projects/{slug}")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert "Цена за м²" in body


def test_update_project_clears_ocr_flag_when_value_changed(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "ОКР правка")
    path = storage.passport_path(tmp_path, slug)
    passport_module.save_passport({
        "project_name": "ОКР правка",
        "year_signed": None,
        "building_class": None,
        "general_contractor": None,
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": 67413.0,
        "ocr_fields": ["total_area_sqm"],
    }, path)

    client.post(f"/projects/{slug}", data={"total_area_sqm": "70000"})

    saved = passport_module.load_passport(path)
    assert saved["total_area_sqm"] == 70000.0
    assert saved["ocr_fields"] == []


def test_update_project_keeps_ocr_flag_when_value_unchanged(tmp_path):
    from app import storage, passport as passport_module

    app = create_app(tmp_path)
    client = app.test_client()
    slug = storage.create_project(tmp_path, "ОКР без правки")
    path = storage.passport_path(tmp_path, slug)
    passport_module.save_passport({
        "project_name": "ОКР без правки",
        "year_signed": None,
        "building_class": None,
        "general_contractor": None,
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": 67413.0,
        "ocr_fields": ["total_area_sqm"],
    }, path)

    client.post(f"/projects/{slug}", data={"total_area_sqm": "67413"})

    saved = passport_module.load_passport(path)
    assert saved["ocr_fields"] == ["total_area_sqm"]


def test_project_page_shows_ai_badge_for_ai_filled_field(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    from app import storage, passport as passport_module

    slug = storage.create_project(tmp_path, "Проект с AI-полем")
    passport_module.save_passport({
        "project_name": "Проект с AI-полем",
        "year_signed": None,
        "building_class": None,
        "general_contractor": "ООО «Из AI»",
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": None,
        "ocr_fields": [],
        "ai_fields": ["general_contractor"],
    }, storage.passport_path(tmp_path, slug))

    resp = client.get(f"/projects/{slug}")

    assert resp.status_code == 200
    assert "Найдено через AI".encode("utf-8") in resp.data


def test_update_project_clears_ai_flag_when_value_changes(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    from app import storage, passport as passport_module

    slug = storage.create_project(tmp_path, "Проект с AI-полем")
    path = storage.passport_path(tmp_path, slug)
    passport_module.save_passport({
        "project_name": "Проект с AI-полем",
        "year_signed": None,
        "building_class": None,
        "general_contractor": None,
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": 67413.0,
        "ocr_fields": [],
        "ai_fields": ["total_area_sqm"],
    }, path)

    client.post(f"/projects/{slug}", data={"total_area_sqm": "70000"})

    saved = passport_module.load_passport(path)
    assert saved["total_area_sqm"] == 70000.0
    assert saved["ai_fields"] == []


def test_update_project_keeps_ai_flag_when_value_unchanged(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    from app import storage, passport as passport_module

    slug = storage.create_project(tmp_path, "Проект с AI-полем")
    path = storage.passport_path(tmp_path, slug)
    passport_module.save_passport({
        "project_name": "Проект с AI-полем",
        "year_signed": None,
        "building_class": None,
        "general_contractor": None,
        "underground_area_sqm": None,
        "aboveground_area_sqm": None,
        "total_area_sqm": 67413.0,
        "ocr_fields": [],
        "ai_fields": ["total_area_sqm"],
    }, path)

    client.post(f"/projects/{slug}", data={"total_area_sqm": "67413"})

    saved = passport_module.load_passport(path)
    assert saved["ai_fields"] == ["total_area_sqm"]


def test_create_project_with_estimate_saves_and_serves_it(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Со сметой",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "smeta_file": (io.BytesIO(_smeta_bytes()), "smeta.xlsx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 302
    slug = "Со_сметой"

    from app import storage
    assert storage.estimate_path(tmp_path, slug).exists()

    page = client.get(f"/projects/{slug}")
    assert page.status_code == 200
    assert "Фундамент".encode("utf-8") in page.data


def test_project_page_says_so_when_there_is_no_estimate(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Без сметы",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 302
    slug = "Без_сметы"

    page = client.get(f"/projects/{slug}")

    assert page.status_code == 200
    assert "Смета не загружена".encode("utf-8") in page.data


def test_create_project_rejects_non_xlsx_estimate(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Плохая смета",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "smeta_file": (io.BytesIO(b"not excel"), "smeta.txt"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 400

    from app import storage
    assert storage.list_project_slugs(tmp_path) == []


def test_new_project_form_has_optional_estimate_field(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.get("/projects/new")

    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert 'name="smeta_file"' in body
    assert 'accept=".xlsx"' in body


def test_project_page_shows_the_estimate_itself_when_file_present(tmp_path):
    # Смета живёт в гармошке на самой странице проекта — одним входом, а не
    # ссылкой на отдельную страницу с той же таблицей.
    app = create_app(tmp_path)
    client = app.test_client()
    client.post("/projects", data={
        "project_name": "Есть смета",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "smeta_file": (io.BytesIO(_smeta_bytes()), "smeta.xlsx"),
    }, content_type="multipart/form-data")

    page = client.get("/projects/Есть_смета")

    assert page.status_code == 200
    body = page.data.decode("utf-8")
    assert 'class="estimate-table"' in body
    assert "Фундамент" in body


def test_project_page_has_no_estimate_table_without_a_file(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    client.post("/projects", data={
        "project_name": "Нет сметы",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")

    page = client.get("/projects/Нет_сметы")

    assert page.status_code == 200
    assert 'class="estimate-table"' not in page.data.decode("utf-8")


def _offer_with_concrete_quantity(volume_m3):
    """Смета с разделом «Возведение несущих конструкций здания» и колонкой
    «Предлагаемое количество»: как в настоящей оферте, объём монолита сидит в
    строке-детали под разделом, а не на его собственной строке."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=9, column=1, value="№ п/п")
    ws.cell(row=9, column=2, value="№ раздела")
    ws.cell(row=9, column=3, value="Статья СМР")
    ws.cell(row=9, column=4, value="Наименование работ")
    ws.cell(row=9, column=7, value="Ед. изм")
    ws.cell(row=10, column=10, value="Предлагаемое количество")
    ws.cell(row=9, column=12, value="Стоимость всего")
    ws.cell(row=10, column=12, value="Всего")

    ws.cell(row=11, column=2, value=4)
    ws.cell(row=11, column=3, value="4. Конструктивные решения")
    ws.cell(row=11, column=4, value="Возведение несущих конструкций здания")
    ws.cell(row=11, column=7, value="м3")
    ws.cell(row=11, column=12, value=1)

    ws.cell(row=12, column=4, value="Фундаментная плита")
    ws.cell(row=12, column=7, value="м3")
    ws.cell(row=12, column=10, value=volume_m3)
    ws.cell(row=12, column=12, value=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _offer_with_facade_quantity(area_m2):
    """Смета с фасадным разделом и колонкой «Предлагаемое количество»,
    построенная так же, как ``_offer_with_concrete_quantity``, но для
    площади фасада."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=9, column=1, value="№ п/п")
    ws.cell(row=9, column=2, value="№ раздела")
    ws.cell(row=9, column=3, value="Статья СМР")
    ws.cell(row=9, column=4, value="Наименование работ")
    ws.cell(row=9, column=7, value="Ед. изм")
    ws.cell(row=10, column=10, value="Предлагаемое количество")
    ws.cell(row=9, column=12, value="Стоимость всего")
    ws.cell(row=10, column=12, value="Всего")

    ws.cell(row=11, column=2, value=6)
    ws.cell(row=11, column=3, value="6. Устройство фасадов")
    ws.cell(row=11, column=4, value="Устройство фасадов")
    ws.cell(row=11, column=7, value="м2")
    ws.cell(row=11, column=12, value=1)

    ws.cell(row=12, column=4, value="Панель навесного фасада")
    ws.cell(row=12, column=7, value="м2")
    ws.cell(row=12, column=10, value=area_m2)
    ws.cell(row=12, column=12, value=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _offer_with_concrete_and_facade_quantity(volume_m3, area_m2):
    """One estimate carrying both a concrete and a facade section, so a
    project can show every coefficient at once."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=9, column=1, value="№ п/п")
    ws.cell(row=9, column=2, value="№ раздела")
    ws.cell(row=9, column=3, value="Статья СМР")
    ws.cell(row=9, column=4, value="Наименование работ")
    ws.cell(row=9, column=7, value="Ед. изм")
    ws.cell(row=10, column=10, value="Предлагаемое количество")
    ws.cell(row=9, column=12, value="Стоимость всего")
    ws.cell(row=10, column=12, value="Всего")

    ws.cell(row=11, column=2, value=4)
    ws.cell(row=11, column=3, value="4. Конструктивные решения")
    ws.cell(row=11, column=4, value="Возведение несущих конструкций здания")
    ws.cell(row=11, column=7, value="м3")
    ws.cell(row=11, column=12, value=1)
    ws.cell(row=12, column=4, value="Фундаментная плита")
    ws.cell(row=12, column=7, value="м3")
    ws.cell(row=12, column=10, value=volume_m3)
    ws.cell(row=12, column=12, value=1)

    ws.cell(row=13, column=2, value=6)
    ws.cell(row=13, column=3, value="6. Устройство фасадов")
    ws.cell(row=13, column=4, value="Устройство фасадов")
    ws.cell(row=13, column=7, value="м2")
    ws.cell(row=13, column=12, value=1)
    ws.cell(row=14, column=4, value="Панель навесного фасада")
    ws.cell(row=14, column=7, value="м2")
    ws.cell(row=14, column=10, value=area_m2)
    ws.cell(row=14, column=12, value=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_project_page_shows_the_concrete_coefficient(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "СБетоном", total_area_sqm=1000.0)
    storage.estimate_path(tmp_path, slug).write_bytes(_offer_with_concrete_quantity(500.0))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Расчётные коэффициенты бетонных и фасадных конструкций" in body
    assert "Коэффициент монолита за общую площадь по СП" in body
    assert passport_module.format_number(500.0) in body   # объём монолита
    assert passport_module.format_number(0.5) in body     # 500 м³ / 1000 м²


def test_project_page_explains_missing_concrete_section_instead_of_a_blank(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "БезБетона", total_area_sqm=1000.0)
    storage.estimate_path(tmp_path, slug).write_bytes(_smeta_bytes())

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Расчётные коэффициенты бетонных и фасадных конструкций" in body
    assert "нет раздела" in body


def test_project_page_shows_the_facade_coefficient(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "СФасадом", total_area_sqm=1000.0)
    storage.estimate_path(tmp_path, slug).write_bytes(_offer_with_facade_quantity(2500.0))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Расчётные коэффициенты бетонных и фасадных конструкций" in body
    assert "Коэффициент фасада за общую площадь по СП, м²(фас)/м²" in body
    assert passport_module.format_number(2500.0) in body   # площадь фасада
    assert passport_module.format_number(2.5) in body      # 2500 м² / 1000 м²


def test_project_page_explains_missing_facade_section_instead_of_a_blank(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "БезФасада", total_area_sqm=1000.0)
    storage.estimate_path(tmp_path, slug).write_bytes(_smeta_bytes())

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Расчётные коэффициенты бетонных и фасадных конструкций" in body
    assert "Площадь фасада по смете" in body
    assert "впишите площадь вручную" in body


def test_manual_coefficients_form_saves_all_three_values(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "САрматурой")

    resp = client.post(
        f"/projects/{slug}/manual-coefficients",
        data={
            "rebar_coefficient_avg": "0,12", "facade_area_manual": "1 500,5",
            "concrete_volume_manual": "600,25",
        },
    )

    assert resp.status_code == 302
    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["rebar_coefficient_avg"] == 0.12
    assert saved["facade_area_manual"] == 1500.5
    assert saved["concrete_volume_manual"] == 600.25

    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "Коэффициент арматуры (средний)" in body
    assert passport_module.format_number(0.12) in body
    assert passport_module.format_number(1500.5) in body
    assert passport_module.format_number(600.25) in body


def test_manual_coefficients_clear_on_an_empty_submit(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(
        tmp_path, "БезАрматуры", rebar_coefficient_avg=0.2, facade_area_manual=1000.0,
        concrete_volume_manual=500.0,
    )

    client.post(
        f"/projects/{slug}/manual-coefficients",
        data={"rebar_coefficient_avg": "", "facade_area_manual": "", "concrete_volume_manual": ""},
    )

    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["rebar_coefficient_avg"] is None
    assert saved["facade_area_manual"] is None
    assert saved["concrete_volume_manual"] is None


def test_manual_concrete_volume_overrides_the_one_read_from_the_estimate(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "СРучнымОбъёмом", total_area_sqm=1000.0)
    storage.estimate_path(tmp_path, slug).write_bytes(_offer_with_concrete_quantity(500.0))

    client.post(
        f"/projects/{slug}/manual-coefficients",
        data={"rebar_coefficient_avg": "", "facade_area_manual": "", "concrete_volume_manual": "700"},
    )

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert passport_module.format_number(700.0) in body   # ручное значение
    assert passport_module.format_number(0.7) in body     # 700 м³ / 1000 м²


def test_concrete_volume_falls_back_to_the_estimate_once_the_manual_value_is_cleared(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(
        tmp_path, "СОчищеннымОбъёмом", total_area_sqm=1000.0, concrete_volume_manual=700.0,
    )
    storage.estimate_path(tmp_path, slug).write_bytes(_offer_with_concrete_quantity(500.0))

    client.post(
        f"/projects/{slug}/manual-coefficients",
        data={"rebar_coefficient_avg": "", "facade_area_manual": "", "concrete_volume_manual": ""},
    )

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert passport_module.format_number(500.0) in body   # снова из сметы


def test_project_page_explains_missing_concrete_estimate_instead_of_a_blank(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "БезСметы", total_area_sqm=1000.0)

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Расчётные коэффициенты бетонных и фасадных конструкций" in body
    assert "Объём монолита по смете" in body
    assert "впишите объём вручную" in body


def test_manual_facade_area_overrides_the_one_read_from_the_estimate(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "СРучнымФасадом", total_area_sqm=1000.0)
    storage.estimate_path(tmp_path, slug).write_bytes(_offer_with_facade_quantity(2500.0))

    client.post(
        f"/projects/{slug}/manual-coefficients",
        data={"rebar_coefficient_avg": "", "facade_area_manual": "3000"},
    )

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert passport_module.format_number(3000.0) in body   # ручное значение
    assert passport_module.format_number(3.0) in body      # 3000 м² / 1000 м²


def test_facade_area_falls_back_to_the_estimate_once_the_manual_value_is_cleared(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(
        tmp_path, "СОчищеннымФасадом", total_area_sqm=1000.0, facade_area_manual=3000.0,
    )
    storage.estimate_path(tmp_path, slug).write_bytes(_offer_with_facade_quantity(2500.0))

    client.post(
        f"/projects/{slug}/manual-coefficients",
        data={"rebar_coefficient_avg": "", "facade_area_manual": ""},
    )

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert passport_module.format_number(2500.0) in body   # снова из сметы


def test_compare_page_shows_the_concrete_facade_and_rebar_coefficient_charts(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(
        tmp_path, "СКоэффициентами", total_area_sqm=1000.0, rebar_coefficient_avg=120.5,
    )
    storage.estimate_path(tmp_path, slug).write_bytes(
        _offer_with_concrete_and_facade_quantity(500.0, 2500.0)
    )

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    assert "Коэффициент монолита за общую площадь по СП, м³/м²" in body
    assert "Коэффициент фасада за общую площадь по СП, м²(фас)/м²" in body
    assert "Коэффициент арматуры (средний), кг/м³" in body
    assert passport_module.format_number(0.5) in body      # 500 м³ / 1000 м²
    assert passport_module.format_number(2.5) in body      # 2500 м² / 1000 м²
    assert passport_module.format_number(120.5) in body
    # Больше не дублируется строкой в «Общих сведениях» — только график.
    assert body.count("Коэффициент монолита за общую площадь по СП, м³/м²") == 1


def test_compare_projects_facade_coefficient_zero_shows_placeholder_not_bar(tmp_path):
    # 0.00 — настоящее значение (площадь фасада вписана нулём), а не "нет
    # данных": закрашенная нулевая полоска рядом с обычными выглядела бы как
    # баг, поэтому вместо неё — пунктирная плашка.
    app = create_app(tmp_path)
    client = app.test_client()
    slug1 = _make_project_with_passport(
        tmp_path, "ПроектА", total_area_sqm=1000.0, facade_area_manual=0.0,
    )
    slug2 = _make_project_with_passport(
        tmp_path, "ПроектБ", total_area_sqm=1000.0, facade_area_manual=2000.0,
    )

    body = client.get(f"/compare?slug={slug1}&slug={slug2}").get_data(as_text=True)

    assert "bar-track is-zero" in body


def test_compare_page_coefficient_charts_say_so_when_data_is_missing(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "БезКоэффициентов")

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    assert "Коэффициент монолита за общую площадь по СП, м³/м²" in body
    assert "Коэффициент фасада за общую площадь по СП, м²(фас)/м²" in body
    assert "Коэффициент арматуры (средний), кг/м³" in body
    assert "Недостаточно данных для этого графика." in body


def test_project_page_renders_multiple_estimate_sheets_as_tabs(tmp_path):
    wb = Workbook()
    wb.active.title = "Смета"
    wb.active["A1"] = "Итого"
    wb.create_sheet("Материалы")["A1"] = "Цемент"
    buf = io.BytesIO()
    wb.save(buf)

    app = create_app(tmp_path)
    client = app.test_client()
    client.post("/projects", data={
        "project_name": "Многолистовая",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "smeta_file": (io.BytesIO(buf.getvalue()), "smeta.xlsx"),
    }, content_type="multipart/form-data")

    page = client.get("/projects/Многолистовая")

    assert page.status_code == 200
    body = page.data.decode("utf-8")
    assert "Итого" in body
    assert "Цемент" in body
    assert "Материалы" in body


def test_create_project_rejects_corrupted_estimate(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    resp = client.post("/projects", data={
        "project_name": "Битая смета",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "smeta_file": (io.BytesIO(b"this is not a real xlsx file"), "smeta.xlsx"),
    }, content_type="multipart/form-data")
    assert resp.status_code == 400

    from app import storage
    assert storage.list_project_slugs(tmp_path) == []


# --- contract-terms upload: the page explains why nothing was filled ---

def _upload_contract_terms(client, slug, follow=True):
    return client.post(
        f"/projects/{slug}/contract-terms",
        data={"contract_terms_file": (io.BytesIO(b"%PDF-fake"), "protocol.pdf")},
        content_type="multipart/form-data",
        follow_redirects=follow,
    )


def _stub_scan_returning(monkeypatch, fields, problem):
    """Make the upload path behave as a scan whose recognition gave this result."""
    from app import passport as passport_module

    monkeypatch.setattr(passport_module.pdf_reader, "read_pdf_text", lambda path: "")
    monkeypatch.setattr(
        passport_module.pdf_reader, "render_pages_to_images",
        lambda path, **kwargs: [b"png"],
    )
    monkeypatch.setattr(
        passport_module.ai_extractor, "extract_contract_terms_from_images",
        lambda images: (fields, problem),
    )


def test_contract_terms_upload_explains_missing_api_key(tmp_path, monkeypatch):
    from app import ai_extractor

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")
    _stub_scan_returning(monkeypatch, {}, ai_extractor.PROBLEM_NO_KEY)

    resp = _upload_contract_terms(client, slug)

    assert resp.status_code == 200
    assert "ANTHROPIC_API_KEY" in resp.data.decode("utf-8")


def test_contract_terms_upload_explains_empty_balance(tmp_path, monkeypatch):
    from app import ai_extractor

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")
    _stub_scan_returning(monkeypatch, {}, ai_extractor.PROBLEM_NO_CREDIT)

    resp = _upload_contract_terms(client, slug)

    # "Plans & Billing" renders escaped, so assert on plain-text wording.
    assert "нет средств" in resp.data.decode("utf-8")


def test_contract_terms_upload_says_nothing_recognized_when_read_but_empty(tmp_path, monkeypatch):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")
    _stub_scan_returning(monkeypatch, {}, None)

    resp = _upload_contract_terms(client, slug)

    assert "распознать не удалось" in resp.data.decode("utf-8")


def test_contract_terms_upload_shows_no_warning_when_fields_were_filled(tmp_path, monkeypatch):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")
    _stub_scan_returning(monkeypatch, {"performance_bond_pct": "3%"}, None)

    resp = _upload_contract_terms(client, slug)

    body = resp.data.decode("utf-8")
    assert "3%" in body
    assert "распознать не удалось" not in body
    assert "ANTHROPIC_API_KEY" not in body


def test_contract_terms_upload_saves_vat_from_recognition(tmp_path, monkeypatch):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА")
    _stub_scan_returning(monkeypatch, {"vat": "20%"}, None)

    _upload_contract_terms(client, slug)

    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["vat"] == "20%"
    assert "vat" in saved["contract_auto_fields"]


# --- contract-terms PDF as part of project creation ---

def _dgp_bytes_signed_in(year):
    return build_docx_bytes(document_xml(paragraphs=[
        "Общество с ограниченной ответственностью «Ромашка» (ООО «Ромашка»), "
        "именуемое в дальнейшем «Генподрядчик», с третьей стороны,",
        f"{year} год",
    ]))


def _create_data(dgp=None, contract_terms=None):
    data = {
        "project_name": "ПроектА",
        "dgp_file": (io.BytesIO(dgp or _dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }
    if contract_terms is not None:
        data["contract_terms_file"] = contract_terms
    return data


def test_create_project_accepts_optional_contract_terms_pdf(tmp_path, monkeypatch):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    _stub_scan_returning(monkeypatch, {"performance_bond_pct": "3%"}, None)

    resp = client.post("/projects", data=_create_data(
        contract_terms=(io.BytesIO(b"%PDF-fake"), "protocol.pdf"),
    ), content_type="multipart/form-data", follow_redirects=True)

    assert resp.status_code == 200
    slug = storage.list_project_slugs(tmp_path)[0]
    assert storage.contract_terms_path(tmp_path, slug).exists()
    assert "3%" in resp.data.decode("utf-8")


def test_create_project_works_without_contract_terms(tmp_path):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.post(
        "/projects", data=_create_data(), content_type="multipart/form-data",
    )

    assert resp.status_code == 302
    slug = storage.list_project_slugs(tmp_path)[0]
    assert not storage.contract_terms_path(tmp_path, slug).exists()
    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["vat"] is None


def test_create_project_rejects_contract_terms_that_is_not_pdf(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.post("/projects", data=_create_data(
        contract_terms=(io.BytesIO(b"not a pdf"), "protocol.docx"),
    ), content_type="multipart/form-data")

    assert resp.status_code == 400
    assert "PDF" in resp.data.decode("utf-8")


def test_create_project_rejects_oversized_contract_terms(tmp_path):
    from app import routes

    app = create_app(tmp_path)
    client = app.test_client()
    too_big = b"x" * (routes.MAX_CONTRACT_TERMS_SIZE + 1)

    resp = client.post("/projects", data=_create_data(
        contract_terms=(io.BytesIO(too_big), "protocol.pdf"),
    ), content_type="multipart/form-data")

    assert resp.status_code == 400


def test_create_project_sets_vat_from_signing_year_2026(tmp_path, monkeypatch):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    _stub_scan_returning(monkeypatch, {"performance_bond_pct": "3%"}, None)

    client.post("/projects", data=_create_data(
        dgp=_dgp_bytes_signed_in(2026),
        contract_terms=(io.BytesIO(b"%PDF-fake"), "protocol.pdf"),
    ), content_type="multipart/form-data")

    slug = storage.list_project_slugs(tmp_path)[0]
    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["year_signed"] == "2026"
    assert saved["vat"] == "22%"


def test_create_project_sets_vat_from_signing_year_2025(tmp_path, monkeypatch):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    _stub_scan_returning(monkeypatch, {"performance_bond_pct": "3%"}, None)

    client.post("/projects", data=_create_data(
        dgp=_dgp_bytes_signed_in(2025),
        contract_terms=(io.BytesIO(b"%PDF-fake"), "protocol.pdf"),
    ), content_type="multipart/form-data")

    slug = storage.list_project_slugs(tmp_path)[0]
    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["vat"] == "20%"


def test_create_project_explains_contract_terms_problem(tmp_path, monkeypatch):
    from app import ai_extractor

    app = create_app(tmp_path)
    client = app.test_client()
    _stub_scan_returning(monkeypatch, {}, ai_extractor.PROBLEM_NO_KEY)

    resp = client.post("/projects", data=_create_data(
        contract_terms=(io.BytesIO(b"%PDF-fake"), "protocol.pdf"),
    ), content_type="multipart/form-data", follow_redirects=True)

    assert "ANTHROPIC_API_KEY" in resp.data.decode("utf-8")


def test_contract_terms_upload_sets_vat_from_passport_year(tmp_path, monkeypatch):
    from app import passport as passport_module, storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "ПроектА", year_signed="2026")
    _stub_scan_returning(monkeypatch, {"performance_bond_pct": "3%"}, None)

    _upload_contract_terms(client, slug)

    saved = passport_module.load_passport(storage.passport_path(tmp_path, slug))
    assert saved["vat"] == "22%"


def test_new_project_form_offers_a_contract_terms_field(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    body = client.get("/projects/new").data.decode("utf-8")

    assert 'name="contract_terms_file"' in body


# --- фото объекта на форме создания проекта ---

def _cover_bytes(size=64):
    return b"\x89PNG\r\n\x1a\n" + b"x" * size


def test_new_project_form_offers_a_cover_field(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    body = client.get("/projects/new").data.decode("utf-8")

    assert 'name="cover_file"' in body
    assert "Фото объекта" in body


def test_create_project_with_a_cover_saves_and_serves_it(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.post("/projects", data={
        "project_name": "С фото",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "cover_file": (io.BytesIO(_cover_bytes()), "photo.PNG"),
    }, content_type="multipart/form-data")

    assert resp.status_code == 302
    slug = "С_фото"
    assert storage.cover_path(tmp_path, slug) is not None
    assert client.get(f"/projects/{slug}/cover").status_code == 200


def test_create_project_without_a_cover_leaves_the_project_without_one(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()

    client.post("/projects", data={
        "project_name": "Без фото",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")

    assert storage.cover_path(tmp_path, "Без_фото") is None
    assert client.get("/projects/Без_фото/cover").status_code == 404


def test_create_project_rejects_a_cover_in_the_wrong_format(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.post("/projects", data={
        "project_name": "Плохое фото",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "cover_file": (io.BytesIO(b"not a picture"), "photo.txt"),
    }, content_type="multipart/form-data")

    assert resp.status_code == 400
    assert "JPG, PNG или WEBP" in resp.data.decode("utf-8")
    # Отказ до создания: нерабочей папки после него не остаётся.
    assert not (tmp_path / "Плохое_фото").exists()


def test_create_project_rejects_an_oversized_cover(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    too_big = b"\x89PNG\r\n\x1a\n" + b"x" * (5 * 1024 * 1024)

    resp = client.post("/projects", data={
        "project_name": "Тяжёлое фото",
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
        "cover_file": (io.BytesIO(too_big), "photo.png"),
    }, content_type="multipart/form-data")

    assert resp.status_code == 400
    assert "до 5 МБ" in resp.data.decode("utf-8")


def test_cover_upload_from_the_project_list_still_works(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Проект")

    resp = client.post(
        f"/projects/{slug}/cover",
        data={"cover_file": (io.BytesIO(_cover_bytes()), "photo.jpg")},
        content_type="multipart/form-data",
    )

    assert resp.status_code == 302
    assert storage.cover_path(tmp_path, slug).name == "cover.jpg"


def test_cover_upload_from_the_project_list_refuses_a_wrong_format(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Проект")

    resp = client.post(
        f"/projects/{slug}/cover",
        data={"cover_file": (io.BytesIO(b"nope"), "photo.txt")},
        content_type="multipart/form-data",
    )

    assert resp.status_code == 400


# --- сравнение по разделам сметы ---

def _offer_bytes(sections):
    """Смета, похожая на настоящую оферту: разделы с номерами и итогами."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="№ п/п")
    ws.cell(row=1, column=2, value="№ раздела")
    ws.cell(row=1, column=3, value="Статья СМР")
    ws.cell(row=1, column=4, value="Наименование работ")
    ws.cell(row=1, column=5, value="Стоимость всего, RUB, с учетом НДС 20%")
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
    ws.cell(row=2, column=6, value="Всего")
    for offset, (article, total) in enumerate(sections):
        row = 3 + offset
        ws.cell(row=row, column=1, value=offset + 1)
        ws.cell(row=row, column=2, value=offset + 1)
        ws.cell(row=row, column=3, value=article)
        ws.cell(row=row, column=6, value=total)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _project_with_offer(root, name, sections, **fields):
    from app import passport as passport_module, storage

    slug = storage.create_project(root, name)
    data = {"project_name": name, "year_signed": "2025", "total_area_sqm": 1000.0}
    data.update(fields)
    passport_module.save_passport(data, storage.passport_path(root, slug))
    storage.estimate_path(root, slug).write_bytes(_offer_bytes(sections))
    return slug


def test_compare_page_shows_the_section_table(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(
        tmp_path, "СоСметой", [("6. Фасадные работы", 3_000_000.0)],
    )

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    assert "Стоимость по разделам" in body
    assert "Фасад" in body
    assert "3 000" in body           # 3 000 000 ₽ / 1 000 м²
    assert "Итого СМР" in body


def test_compare_page_has_no_section_table_without_estimates(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "БезСметы", total_area_sqm=1000.0)

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    assert "Стоимость по разделам" not in body


def test_section_table_corrections_are_off_by_default(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(
        tmp_path, "Проект", [("6. Фасадные работы", 1_200_000.0)], year_signed="2020",
    )

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    assert "1 200" in body


def test_section_table_applies_inflation_from_the_address(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(
        tmp_path, "Проект", [("6. Фасадные работы", 1_000_000.0)], year_signed="2024",
    )

    body = client.get(
        f"/compare?slug={slug}&inflation_on=1&inflation=10&year=2026"
    ).get_data(as_text=True)

    assert "1 210" in body          # 1 000 × 1,1²


def test_section_table_applies_vat_from_the_address(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(
        tmp_path, "Проект", [("6. Фасадные работы", 1_200_000.0)], year_signed="2025",
    )

    body = client.get(f"/compare?slug={slug}&vat_on=1&vat=22").get_data(as_text=True)

    assert "1 220" in body          # 1 200 × 122/120


def test_section_table_marks_a_project_whose_year_is_unknown(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(
        tmp_path, "Проект", [("6. Фасадные работы", 1_000_000.0)], year_signed=None,
    )

    body = client.get(
        f"/compare?slug={slug}&inflation_on=1&inflation=10"
    ).get_data(as_text=True)

    assert "без поправки на инфляцию" in body


def test_compare_page_shows_the_pair_cards(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_offer(
        tmp_path, "Левый", [("6. Фасадные работы", 1_000_000_000.0)],
        contract_price_rub=1_000_000_000.0,
    )
    b = _project_with_offer(
        tmp_path, "Правый", [("6. Фасадные работы", 2_000_000_000.0)],
        contract_price_rub=2_000_000_000.0,
    )

    body = client.get(f"/compare?slug={a}&slug={b}").get_data(as_text=True)

    assert "Сравнение двух объектов" in body
    assert "Цена работ по договору" in body
    assert "1 000 000 000 ₽" in body
    assert "2 000 000 000 ₽" in body
    assert "+100,0 %" in body
    assert "Дельта по разделам" in body


def test_the_pair_can_be_chosen_from_the_page(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slugs = [
        _project_with_offer(tmp_path, f"П{i}", [("6. Фасадные работы", i * 1_000_000_000.0)],
                            contract_price_rub=i * 1_000_000_000.0)
        for i in (1, 2, 3)
    ]
    query = "&".join(f"slug={slug}" for slug in slugs)

    body = client.get(f"/compare?{query}&left={slugs[0]}&right={slugs[2]}").get_data(as_text=True)

    assert "3 000 000 000 ₽" in body
    assert "+200,0 %" in body


def test_one_project_alone_gets_no_pair_cards(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(tmp_path, "Один", [("6. Фасадные работы", 1.0)])

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    assert "Сравнение двух объектов" not in body


# --- удорожание объекта ----------------------------------------------------

def _increase_bytes(rows, *, header_row=1):
    """Файл удорожания той же формы, что настоящий: названия работ — в столбце
    без заголовка, деньги — под «было» и «стало»."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=header_row, column=3, value="было")
    ws.cell(row=header_row, column=4, value="стало")
    for offset, (name, was, now) in enumerate(rows):
        row = header_row + 1 + offset
        ws.cell(row=row, column=1, value=offset + 1)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=was)
        ws.cell(row=row, column=4, value=now)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_increase(client, slug, data, filename="udorozhanie.xlsx"):
    return client.post(
        f"/projects/{slug}/cost-increase",
        data={"cost_increase_file": (io.BytesIO(data), filename)},
        content_type="multipart/form-data",
    )


# --- справка по объекту в PDF -----------------------------------------------

def test_the_project_page_offers_to_save_the_pdf_summary(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Сохранить справку в PDF" in body


def test_the_project_pdf_route_returns_a_pdf(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    resp = client.get(f"/projects/{slug}/pdf")

    assert resp.status_code == 200
    assert resp.data.startswith(b"%PDF")


def test_the_project_pdf_for_an_unknown_project_is_not_found(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = client.get("/projects/нет-такого/pdf")

    assert resp.status_code == 404


def test_the_project_pdf_route_works_with_estimate_and_increase(tmp_path):
    slug = _project_with_offer(tmp_path, "Тест", [("8. Кровля", 1_000_000.0)])
    app = create_app(tmp_path)
    client = app.test_client()
    _upload_increase(client, slug, _increase_bytes([("Кровля", 1_100_000.0, 1_300_000.0)]))

    resp = client.get(f"/projects/{slug}/pdf")

    assert resp.status_code == 200
    assert resp.data.startswith(b"%PDF")


def test_the_project_page_offers_to_upload_a_cost_increase_file(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Удорожание объекта" in body
    assert "Загрузить файл удорожания" in body


def test_an_uploaded_cost_increase_file_shows_a_percentage_per_kind_of_work(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    resp = _upload_increase(client, slug, _increase_bytes([
        ("Устройство гидроизоляции подземной части здания", 110099464.84, 123439548.48),
        ("Кровля", 145926597.99, 145926597.99),
    ]))

    assert resp.status_code == 302
    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "Гидроизоляция подземной части" in body
    assert "+12,1 %" in body
    assert "+13 340 083.64" in body
    # Кровля не сдвинулась — и это видно как «0 %», а не как «+0,0 %».
    assert "0 %" in body
    assert "Заменить файл удорожания" in body


def test_the_two_vis_rows_are_shown_as_one_line_of_engineering_systems(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    _upload_increase(client, slug, _increase_bytes([
        ("ВИС - механические системы", 1120700487.64, 1120700487.64),
        ("ВИС - Электрические и слаботочные системы", 819317798.55, 819317798.55),
    ]))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "1 940 018 286.19" in body
    # Одна строка, а не две: суммы отдельных «ВИС» на странице нет.
    assert "1 120 700 487.64" not in body


def test_work_that_was_not_in_the_estimate_is_shown_as_new_work(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    _upload_increase(client, slug, _increase_bytes([("Благоустройство, дороги", 0, 207987892.97)]))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "новые работы" in body
    assert "+207 987 892.97" in body


def test_uploading_a_newer_cost_increase_file_replaces_the_previous_one(tmp_path):
    # Файл накопительный: в новой версии уже учтено всё, что было в прежней,
    # поэтому она её заменяет, а не дополняет.
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    _upload_increase(client, slug, _increase_bytes([("Кровля", 100.0, 110.0)]))
    _upload_increase(client, slug, _increase_bytes([("Кровля", 100.0, 130.0)]))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "+30,0 %" in body
    assert "+10,0 %" not in body


def test_an_unreadable_cost_increase_file_is_refused_and_the_old_one_kept(tmp_path):
    # Замена, которая не удалась, не должна оставлять человека вообще без
    # удорожания: прежний файл на месте, а сказано ровно то, что произошло.
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")
    _upload_increase(client, slug, _increase_bytes([("Кровля", 100.0, 110.0)]))

    resp = _upload_increase(client, slug, b"not a workbook at all")

    assert resp.status_code == 302
    assert "increase=unreadable" in resp.headers["Location"]
    body = client.get(f"/projects/{slug}?increase=unreadable").get_data(as_text=True)
    assert "Прежний файл оставлен на месте" in body
    assert "+10,0 %" in body


def test_a_cost_increase_file_without_the_money_columns_is_refused(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    wb = Workbook()
    wb.active.append(["Раздел", "Сумма"])
    wb.active.append(["Кровля", 100])
    buf = io.BytesIO()
    wb.save(buf)

    resp = _upload_increase(client, slug, buf.getvalue())

    assert "increase=unreadable" in resp.headers["Location"]


def test_a_cost_increase_file_in_the_wrong_format_is_refused(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    resp = _upload_increase(client, slug, b"whatever", filename="udorozhanie.pdf")

    assert "increase=format" in resp.headers["Location"]
    body = client.get(f"/projects/{slug}?increase=format").get_data(as_text=True)
    assert "должен быть в формате .xlsx" in body


def test_an_oversized_cost_increase_file_is_refused(tmp_path):
    from app import routes

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    resp = _upload_increase(client, slug, b"x" * (routes.MAX_COST_INCREASE_SIZE + 1))

    assert "increase=too_big" in resp.headers["Location"]


def test_uploading_a_cost_increase_file_to_an_unknown_project_is_not_found(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = _upload_increase(client, "нет-такого", _increase_bytes([("Кровля", 1.0, 2.0)]))

    assert resp.status_code == 404


def test_rows_of_the_cost_increase_file_with_no_line_in_the_report_are_named(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    _upload_increase(client, slug, _increase_bytes([
        ("Кровля", 100.0, 110.0),
        ("Аренда вертолётной площадки", 50.0, 60.0),
    ]))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "Аренда вертолётной площадки" in body


def test_a_cost_increase_file_broken_after_it_was_saved_does_not_break_the_page(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")
    storage.cost_increase_path(tmp_path, slug).write_bytes(b"no longer a workbook")

    resp = client.get(f"/projects/{slug}")

    assert resp.status_code == 200
    assert "прочитать его не удалось" in resp.get_data(as_text=True)


# --- удорожание считается от сметы ------------------------------------------

def test_without_an_estimate_the_page_says_the_baseline_is_the_file_itself(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")
    _upload_increase(client, slug, _increase_bytes([("Кровля", 100.0, 110.0)]))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Смета не загружена, поэтому удорожание считается от столбца «было»" in body
    assert "Было, руб." in body


def test_with_an_estimate_the_increase_is_stalo_against_the_estimate(tmp_path):
    # Ровно то правило, ради которого всё это и делается: база — смета,
    # «стало» — сколько работы стоят теперь.
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(tmp_path, "Тест", [("8. Кровля", 1_000_000.0)])

    _upload_increase(client, slug, _increase_bytes([("Кровля", 1_100_000.0, 1_300_000.0)]))
    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Удорожание считается от сметы" in body
    assert "Смета, руб." in body
    # 1 300 000 против 1 000 000 по смете, а не против 1 100 000 из «было».
    assert "+300 000" in body
    assert "+30,0 %" in body
    assert "+200 000" not in body


def test_an_empty_stalo_is_taken_from_was_and_said_so(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(tmp_path, "Тест", [("8. Кровля", 1_000_000.0)])

    _upload_increase(client, slug, _increase_bytes([("Кровля", 1_100_000.0, 0)]))
    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "взято «было»" in body
    assert "+100 000" in body


def test_a_section_of_the_estimate_absent_from_the_file_is_shown_unchanged(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(
        tmp_path, "Тест",
        [("8. Кровля", 100.0), ("6. Фасадные работы", 3_000_000.0)],
    )

    _upload_increase(client, slug, _increase_bytes([("Кровля", 100.0, 110.0)]))
    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Фасад" in body
    assert "в файле нет данных" in body
    # И стоимость раздела осталась на месте, а не прочиталась как экономия.
    assert "3 000 000" in body
    assert "−3 000 000" not in body


def test_work_the_estimate_never_priced_reads_as_new_work(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(tmp_path, "Тест", [("8. Кровля", 100.0)])

    _upload_increase(client, slug, _increase_bytes([
        ("Кровля", 100.0, 100.0),
        ("Благоустройство, дороги", 0, 5_000_000.0),
    ]))
    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "новые работы" in body
    assert "+5 000 000" in body


def test_the_upload_is_never_refused_over_the_estimate(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(tmp_path, "Тест", [("8. Кровля", 1_000_000.0)])

    resp = _upload_increase(client, slug, _increase_bytes([("Кровля", 1_100_000.0, 1_200_000.0)]))

    assert resp.status_code == 302
    assert "increase=" not in resp.headers["Location"]
    assert storage.cost_increase_path(tmp_path, slug).exists()


def test_an_estimate_of_nothing_does_not_put_the_word_none_on_the_page(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_offer(tmp_path, "Тест", [("8. Кровля", 0.0)])

    _upload_increase(client, slug, _increase_bytes([("Кровля", 100.0, 110.0)]))
    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "+110" in body
    assert "None" not in body


# --- удорожание на странице сравнения ---------------------------------------

def _project_with_increase(root, client, name, sections, rows, **fields):
    """Проект со сметой и загруженным файлом удорожания."""
    slug = _project_with_offer(root, name, sections, **fields)
    _upload_increase(client, slug, _increase_bytes(rows))
    return slug


def test_the_comparison_shows_the_increase_block(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_increase(
        tmp_path, client, "Левый", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_300_000.0)],
    )
    b = _project_with_increase(
        tmp_path, client, "Правый", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_100_000.0)],
    )

    body = client.get(f"/compare?slug={a}&slug={b}").get_data(as_text=True)

    assert "Удорожание проектов" in body
    # Плитки: средний процент и сумма по всем проектам.
    assert "Средний % удорожания" in body
    assert "+20,0 %" in body
    assert "Удорожание по всем проектам" in body
    assert "+400 000 ₽" in body
    # Диаграмма по проектам и таблица видов работ.
    assert "Общее увеличение стоимости по проектам" in body
    assert "Виды работ, которые делают смету дороже" in body
    assert "+30,0 %" in body
    assert "+10,0 %" in body


def test_the_works_table_says_how_often_a_kind_of_work_gets_dearer(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_increase(
        tmp_path, client, "Левый",
        [("8. Кровля", 100.0), ("6. Фасадные работы", 100.0)],
        [("Кровля", 100.0, 110.0), ("Фасадные работы", 100.0, 100.0)],
    )
    b = _project_with_increase(
        tmp_path, client, "Правый",
        [("8. Кровля", 100.0), ("6. Фасадные работы", 100.0)],
        [("Кровля", 100.0, 120.0), ("Фасадные работы", 100.0, 100.0)],
    )

    body = client.get(f"/compare?slug={a}&slug={b}").get_data(as_text=True)

    assert "2 из 2" in body
    assert "0 из 2" in body


def test_without_any_cost_increase_file_the_comparison_has_no_such_block(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_offer(tmp_path, "Левый", [("8. Кровля", 100.0)])
    b = _project_with_offer(tmp_path, "Правый", [("8. Кровля", 100.0)])

    body = client.get(f"/compare?slug={a}&slug={b}").get_data(as_text=True)

    assert "Удорожание проектов" not in body


def test_a_single_project_gets_the_figures_without_a_one_bar_chart(tmp_path):
    # Одно число диаграммой не рисуют: плитки остаются, полоска по проектам —
    # нет, потому что сравнивать её не с чем.
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_increase(
        tmp_path, client, "Один", [("8. Кровля", 100.0)], [("Кровля", 100.0, 110.0)],
    )

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    assert "Средний % удорожания" in body
    assert "Общее увеличение стоимости по проектам" not in body


def test_a_broken_cost_increase_file_does_not_break_the_comparison(tmp_path):
    from app import storage

    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_increase(
        tmp_path, client, "Целый", [("8. Кровля", 100.0)], [("Кровля", 100.0, 110.0)],
    )
    b = _project_with_offer(tmp_path, "Битый", [("8. Кровля", 100.0)])
    storage.cost_increase_path(tmp_path, b).write_bytes(b"not a workbook")

    resp = client.get(f"/compare?slug={a}&slug={b}")

    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "Удорожание проектов" in body
    assert "Учтены 1 из 2" in body


def test_the_pdf_carries_the_increase_block_too(tmp_path):
    # Файл показывает страницу целиком — иначе цифра на экране и цифра в файле
    # расходятся, а виновата в этом выгрузка.
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_increase(
        tmp_path, client, "Левый", [("8. Кровля", 100.0)], [("Кровля", 100.0, 130.0)],
    )
    b = _project_with_increase(
        tmp_path, client, "Правый", [("8. Кровля", 100.0)], [("Кровля", 100.0, 110.0)],
    )

    resp = client.get(f"/compare/pdf?slug={a}&slug={b}")

    assert resp.status_code == 200
    assert resp.data.startswith(b"%PDF")


def test_the_works_table_is_sortable_like_the_sections_one(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _project_with_increase(
        tmp_path, client, "Один", [("8. Кровля", 100.0)], [("Кровля", 100.0, 110.0)],
    )

    body = client.get(f"/compare?slug={slug}").get_data(as_text=True)

    # Обработчик сортировки один на все таблицы страницы, а не только на первую.
    assert "document.querySelectorAll('.sections-table')" in body
    assert "works-table" in body


def test_everything_in_a_table_card_starts_on_the_same_line(tmp_path):
    # У карточки с таблицей боковых отступов нет — их держат ячейки. Значит
    # всё остальное, что в ней лежит, должно отодвинуться само и ровно на ту
    # же величину. Иначе плитки, подзаголовки и диаграмма блока удорожания
    # прижимаются к краю карточки, а колонки таблицы отступают, и один блок
    # читается как два, сдвинутых друг относительно друга.
    app = create_app(tmp_path)
    client = app.test_client()

    css = client.get("/static/style.css").get_data(as_text=True)
    block = css.split(".sections-card > .sections-head,", 1)[1].split("}", 1)[0]

    assert "margin-inline: var(--table-edge)" in block
    for selector in (
        ".sections-card > .adjust-form",
        ".sections-card > .stat-tiles",
        ".sections-card > .pair-subtitle",
        ".sections-card > .delta-list",
    ):
        assert selector in block, selector
    # Сама таблица в этот список не входит: край она держит своими ячейками, и
    # внешний отступ сложился бы с ними вдвое.
    assert "sections-table-wrap" not in block


def test_the_works_table_keeps_the_shared_cell_padding(tmp_path):
    # Таблица удорожания стоит в сравнении рядом с таблицей стоимости по
    # разделам, и отступы в ячейках у них должны быть одни. Поэтому своих
    # отступов у неё нет вовсе — она берёт общие для .sections-table.
    import re

    app = create_app(tmp_path)
    client = app.test_client()

    css = client.get("/static/style.css").get_data(as_text=True)
    shared = css.split(".sections-table th,", 1)[1].split("{", 1)[1].split("}", 1)[0]
    works_rules = re.findall(r"\.works-table[^{]*\{([^}]*)\}", css)

    assert "padding: var(--table-pad-y) var(--table-pad-x)" in shared
    assert works_rules, "у таблицы удорожания должны быть свои правила"
    assert not any("padding" in rule for rule in works_rules)


def test_the_comparison_shows_the_increase_per_square_metre(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_increase(
        tmp_path, client, "Левый", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_300_000.0)], total_area_sqm=1000.0,
    )
    b = _project_with_increase(
        tmp_path, client, "Правый", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_100_000.0)], total_area_sqm=1000.0,
    )

    body = client.get(f"/compare?slug={a}&slug={b}").get_data(as_text=True)

    # Плитка, отдельная диаграмма и столбец в таблице видов работ.
    assert "Удорожание на м²" in body
    assert "Удорожание на м² по проектам" in body
    assert "удорожание на м²" in body
    # 400 000 ₽ на 2000 м² — это 200 ₽/м².
    assert "+200 ₽/м²" in body
    assert "+300 ₽/м²" in body
    assert "+100 ₽/м²" in body


def test_a_project_without_an_area_leaves_the_per_metre_figures_out(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_increase(
        tmp_path, client, "С площадью", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_300_000.0)], total_area_sqm=1000.0,
    )
    b = _project_with_increase(
        tmp_path, client, "Без площади", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_100_000.0)], total_area_sqm=None,
    )

    body = client.get(f"/compare?slug={a}&slug={b}").get_data(as_text=True)

    assert "Удорожание проектов" in body
    assert "Удорожание на м²" not in body
    # Только внутри самого блока: ниже идёт сравнение двух объектов, и там
    # «₽/м²» стоит по праву — это цена работ на метр, а не удорожание.
    block = body.split("Удорожание проектов", 1)[1].split("Сравнение двух объектов", 1)[0]
    assert "₽/м²" not in block


def test_the_pdf_carries_the_per_metre_figures_too(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    a = _project_with_increase(
        tmp_path, client, "Левый", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_300_000.0)], total_area_sqm=1000.0,
    )
    b = _project_with_increase(
        tmp_path, client, "Правый", [("8. Кровля", 1_000_000.0)],
        [("Кровля", 1_000_000.0, 1_100_000.0)], total_area_sqm=1000.0,
    )

    resp = client.get(f"/compare/pdf?slug={a}&slug={b}")

    assert resp.status_code == 200
    assert resp.data.startswith(b"%PDF")


# --- заменить ДГП ------------------------------------------------------------

def _dgp_bytes_alt():
    return build_docx_bytes(document_xml(paragraphs=[
        "Общество с ограниченной ответственностью «Вектор» (ООО «Вектор»), "
        "именуемое в дальнейшем «Генподрядчик», с третьей стороны,"
    ]))


def _create_full_project(client, name):
    resp = client.post("/projects", data={
        "project_name": name,
        "dgp_file": (io.BytesIO(_dgp_bytes()), "dgp.docx"),
        "tz_file": (io.BytesIO(_tz_bytes()), "tz.docx"),
    }, content_type="multipart/form-data")
    return resp.headers["Location"].rsplit("/", 1)[-1]


def _upload_dgp(client, slug, data, filename="dgp.docx"):
    return client.post(
        f"/projects/{slug}/dgp",
        data={"dgp_file": (io.BytesIO(data), filename)},
        content_type="multipart/form-data",
    )


def test_the_passport_offers_to_replace_the_dgp(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _create_full_project(client, "Тест")

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Заменить ДГП" in body


def test_replacing_the_dgp_rebuilds_the_passport_fields(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _create_full_project(client, "Тест")

    resp = _upload_dgp(client, slug, _dgp_bytes_alt())

    assert resp.status_code == 302
    assert "dgp=" not in resp.headers["Location"]
    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "ООО «Вектор»" in body
    assert "ООО «Ромашка»" not in body


def test_a_dgp_in_the_wrong_format_is_refused(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _create_full_project(client, "Тест")

    resp = _upload_dgp(client, slug, b"whatever", filename="dgp.txt")

    assert "dgp=format" in resp.headers["Location"]
    body = client.get(f"/projects/{slug}?dgp=format").get_data(as_text=True)
    assert "формате .docx" in body


def test_an_unreadable_dgp_is_refused_and_the_old_one_kept(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _create_full_project(client, "Тест")

    resp = _upload_dgp(client, slug, b"this has a .docx name but is not a real zip")

    assert "dgp=unreadable" in resp.headers["Location"]
    body = client.get(f"/projects/{slug}?dgp=unreadable").get_data(as_text=True)
    assert "Прежний ДГП оставлен на месте" in body
    # Паспорт не тронут — старый генподрядчик по-прежнему на странице.
    assert "ООО «Ромашка»" in body


def test_replacing_the_dgp_for_an_unknown_project_is_not_found(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = _upload_dgp(client, "нет-такого", _dgp_bytes_alt())

    assert resp.status_code == 404


# --- заменить смету -----------------------------------------------------------

def _upload_estimate(client, slug, data, filename="smeta.xlsx"):
    return client.post(
        f"/projects/{slug}/estimate",
        data={"estimate_file": (io.BytesIO(data), filename)},
        content_type="multipart/form-data",
    )


def test_the_estimate_accordion_offers_to_upload_when_there_is_none(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    body = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "Загрузить смету" in body


def test_uploading_an_estimate_offers_to_replace_it_afterwards(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    resp = _upload_estimate(client, slug, _offer_bytes([("8. Кровля", 1_000_000.0)]))

    assert resp.status_code == 302
    assert "estimate=" not in resp.headers["Location"]
    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "Заменить смету" in body


def test_uploading_a_newer_estimate_replaces_the_previous_one(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")
    _upload_estimate(client, slug, _offer_bytes([("8. Кровля", 1_000_000.0)]))

    _upload_estimate(client, slug, _offer_bytes([("6. Фасадные работы", 3_000_000.0)]))

    body = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "Фасадные работы" in body
    assert "Кровля" not in body


def test_an_estimate_in_the_wrong_format_is_refused(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")

    resp = _upload_estimate(client, slug, b"whatever", filename="smeta.pdf")

    assert "estimate=format" in resp.headers["Location"]
    body = client.get(f"/projects/{slug}?estimate=format").get_data(as_text=True)
    assert "формате .xlsx" in body


def test_an_unreadable_estimate_is_refused_and_the_old_one_kept(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")
    _upload_estimate(client, slug, _offer_bytes([("8. Кровля", 1_000_000.0)]))

    resp = _upload_estimate(client, slug, b"not a workbook at all")

    assert "estimate=unreadable" in resp.headers["Location"]
    body = client.get(f"/projects/{slug}?estimate=unreadable").get_data(as_text=True)
    assert "Прежняя смета оставлена на месте" in body
    assert "Кровля" in body


def test_replacing_the_estimate_for_an_unknown_project_is_not_found(tmp_path):
    app = create_app(tmp_path)
    client = app.test_client()

    resp = _upload_estimate(client, "нет-такого", _offer_bytes([("8. Кровля", 1.0)]))

    assert resp.status_code == 404


def test_replacing_the_estimate_recalculates_the_cost_increase(tmp_path):
    # Удорожание считается от сметы — значит, после замены сметы цифры на
    # странице обязаны отвечать новой смете, а не той, что была при загрузке
    # файла удорожания.
    app = create_app(tmp_path)
    client = app.test_client()
    slug = _make_project_with_passport(tmp_path, "Тест")
    _upload_estimate(client, slug, _offer_bytes([("8. Кровля", 1_000_000.0)]))
    _upload_increase(client, slug, _increase_bytes([("Кровля", 1_100_000.0, 1_300_000.0)]))
    before = client.get(f"/projects/{slug}").get_data(as_text=True)
    assert "+300 000" in before

    _upload_estimate(client, slug, _offer_bytes([("8. Кровля", 1_200_000.0)]))
    after = client.get(f"/projects/{slug}").get_data(as_text=True)

    assert "+100 000" in after
    assert "+300 000" not in after
