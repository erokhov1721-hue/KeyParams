from decimal import Decimal

import openpyxl
from openpyxl import Workbook

from app import estimate_sections, workbook_cache


def _offer(rows, *, header_row=9):
    """A workbook shaped like a real tender offer.

    ``rows`` are ``(item_no, section_no, article, works_name, total)`` tuples,
    written below a two-storey header: "Стоимость всего" spans four columns
    with "Всего" as the last of them, and the per-unit prices above sit under
    an identically-shaped heading of their own — which is what makes finding
    the right "Всего" the interesting part.
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="Предмет тендера:")

    ws.cell(row=header_row, column=1, value="№ п/п")
    ws.cell(row=header_row, column=2, value="№ раздела")
    ws.cell(row=header_row, column=3, value="Статья СМР")
    ws.cell(row=header_row, column=4, value="Наименование работ")

    ws.cell(row=header_row, column=5, value="Цена за ед. изм., RUB, с учетом НДС 20%")
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=8)
    ws.cell(row=header_row, column=9, value="Стоимость всего, RUB, с учетом НДС 20%")
    ws.merge_cells(start_row=header_row, start_column=9, end_row=header_row, end_column=12)
    ws.cell(row=header_row, column=13, value="Стоимость всего за объемы заказчика")

    for col, title in ((5, "Материалы"), (6, "СМР"), (7, "Косвенные расходы"), (8, "Всего")):
        ws.cell(row=header_row + 1, column=col, value=title)
    for col, title in ((9, "Материалы"), (10, "СМР"), (11, "Косвенные расходы"), (12, "Всего")):
        ws.cell(row=header_row + 1, column=col, value=title)

    for offset, (item, section, article, works, total) in enumerate(rows):
        row = header_row + 2 + offset
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=section)
        ws.cell(row=row, column=3, value=article)
        ws.cell(row=row, column=4, value=works)
        # A different number in the unit-price "Всего", so a reader that picks
        # the wrong column is caught rather than accidentally right.
        ws.cell(row=row, column=8, value=1)
        ws.cell(row=row, column=12, value=total)
    return wb


def _save(wb, tmp_path, name="smeta.xlsx"):
    path = tmp_path / name
    wb.save(path)
    return path


# --- classify ---

def test_classify_recognises_the_sections_of_a_real_offer():
    cases = {
        "1. Подготовительные работы, содержание площадки": "preparation",
        "2. Котлован": "excavation",
        "3. Устройство гидроизоляции подземной части": "waterproofing",
        "4. Конструктивные решения": "concrete",
        "5. Общестроительные работы (без отделки)": "partitions",
        "6. Фасадные работы": "facade",
        "7. Кровля": "roof",
        "8. Отделочные работы (паркинг, надземная часть МОП)": "finishing",
        "9. Лифты, подъемники": "lifts",
        "Инженерные системы": "utilities",
        "12. Благоустройство, дороги": "landscaping",
        "13. ТХ": "technology",
        "99. Прочее": "other",
        "14. MR-base": "mr_base",
        "16. Разработка рабочей документации": "rd",
        "Дополнительные работы": "other",
    }
    for name, expected in cases.items():
        assert estimate_sections.classify(name) == expected, name


def test_classify_ignores_a_leading_article_number():
    # The numbering is per-workbook and drifts; the name is what identifies
    # a section.
    assert estimate_sections.classify("12. Благоустройство") == "landscaping"
    assert estimate_sections.classify("Благоустройство") == "landscaping"


def test_classify_matches_a_short_name_as_a_whole_word():
    # "ТХ" is a section; the same two letters inside "отходов" are not.
    assert estimate_sections.classify("13. ТХ") == "technology"
    assert estimate_sections.classify("Вывоз отходов") is None


def test_classify_of_an_unknown_section_is_none():
    assert estimate_sections.classify("Содержание вертолётной площадки") is None
    assert estimate_sections.classify("") is None
    assert estimate_sections.classify(None) is None


# --- read_section_totals ---

def test_totals_are_read_per_section(tmp_path):
    path = _save(_offer([
        (1, 1, "1. Подготовительные работы", "Подготовка", 100.0),
        (2, None, None, "Ограждение", 40.0),
        (3, "1.1", "1.1. Мобилизация", "Мобилизация", 60.0),
        (4, 2, "2. Котлован", "Котлован", 200.0),
    ]), tmp_path)

    totals = estimate_sections.read_section_totals(path)

    # Only the top-level rows count: their sub-sections and line items are
    # already summed into them.
    assert totals == {"preparation": 100.0, "excavation": 200.0}


def test_totals_sum_exactly_not_with_a_floats_rounding_drift(tmp_path):
    # 100000.10 + 200000.20 + 300000.05 drifts to 600000.3500000001 summed
    # as float — read as Decimal, it doesn't.
    path = _save(_offer([
        (1, 1, "1. Подготовительные работы", "Подготовка", 100000.10),
        (2, 2, "1. Подготовительные работы", "Подготовка 2", 200000.20),
        (3, 3, "1. Подготовительные работы", "Подготовка 3", 300000.05),
    ]), tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals["preparation"] == Decimal("600000.35")
    assert isinstance(totals["preparation"], Decimal)


def test_the_lot_header_row_is_not_counted_as_a_section(tmp_path):
    # An offer opens with the lot itself: a section number, no article name,
    # and the whole offer as its total.
    path = _save(_offer([
        (1, 1, None, "Лот №1 - Генподряд", 999.0),
        (1, 1, "1. Котлован", "Котлован", 200.0),
    ]), tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"excavation": 200.0}


def test_a_section_named_only_in_the_works_column_is_still_read(tmp_path):
    # Seen in a real offer: one section of fifteen had its "Статья СМР" left
    # blank, and dropping it would have lost 1.9 billion roubles.
    path = _save(_offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
        (2, 2, None, "Инженерные системы", 1908214707.4),
    ]), tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals["utilities"] == Decimal("1908214707.4")


def test_a_section_that_costs_nothing_is_a_zero_not_a_gap(tmp_path):
    path = _save(_offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
        (2, 2, "12. Благоустройство", "Благоустройство", 0),
    ]), tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals["landscaping"] == 0.0


def test_an_unnumbered_addition_below_the_sections_is_counted(tmp_path):
    # "Дополнительные работы" carries neither a line number nor a section
    # number; skipping it left the report short of the offer's own total.
    path = _save(_offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
        (None, None, None, "Дополнительные работы", 12.5),
    ]), tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"excavation": 200.0, "other": 12.5}


def test_the_closing_total_rows_are_not_counted(tmp_path):
    # "ИТОГО" names itself in the margin, outside the columns this reads, so
    # it never acquires a name and falls out.
    wb = _offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
    ])
    ws = wb.active
    ws.cell(row=12, column=1, value="ИТОГО, руб. с учетом НДС")
    ws.cell(row=12, column=12, value=200.0)
    path = _save(wb, tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"excavation": 200.0}


def test_a_backup_breakdown_after_the_итого_row_is_not_double_counted(tmp_path):
    # One offer appended, right after its "ИТОГО", a second table restating
    # one of its sections (a numbered row, its own total) with the cost
    # broken down underneath into unnumbered parts that sum back to it.
    # Reading those parts as more standalone additions — the rule meant for
    # a section-less row like "Дополнительные работы" — added that
    # section's cost a second time on top of what it already contributed.
    wb = _offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
    ])
    ws = wb.active
    ws.cell(row=12, column=1, value="ИТОГО, руб. с учетом НДС")
    ws.cell(row=12, column=12, value=200.0)
    # The backup table: "2. Гидроизоляция" restated with its own total,
    # then the two unnumbered parts that make it up.
    ws.cell(row=14, column=2, value=2)
    ws.cell(row=14, column=4, value="Гидроизоляция")
    ws.cell(row=14, column=8, value=1)
    ws.cell(row=14, column=12, value=50.0)
    ws.cell(row=15, column=4, value="Праймер")
    ws.cell(row=15, column=8, value=1)
    ws.cell(row=15, column=12, value=30.0)
    ws.cell(row=16, column=4, value="Мембрана")
    ws.cell(row=16, column=8, value=1)
    ws.cell(row=16, column=12, value=20.0)
    ws.cell(row=17, column=1, value="ИТОГО, руб. с учетом НДС")
    ws.cell(row=17, column=12, value=50.0)
    path = _save(wb, tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"excavation": 200.0, "waterproofing": 50.0}


def test_a_third_block_after_the_backup_breakdown_is_not_counted(tmp_path):
    # Nothing in this codebase's real estimates goes three blocks deep — if
    # one ever does, the safer failure is to stop reading rather than to
    # guess at a second unfamiliar shape.
    wb = _offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
    ])
    ws = wb.active
    ws.cell(row=12, column=1, value="ИТОГО, руб. с учетом НДС")
    ws.cell(row=12, column=12, value=200.0)
    ws.cell(row=14, column=2, value=2)
    ws.cell(row=14, column=4, value="Гидроизоляция")
    ws.cell(row=14, column=8, value=1)
    ws.cell(row=14, column=12, value=50.0)
    ws.cell(row=15, column=1, value="ИТОГО, руб. с учетом НДС")
    ws.cell(row=15, column=12, value=50.0)
    ws.cell(row=17, column=2, value=3)
    ws.cell(row=17, column=4, value="Кровля")
    ws.cell(row=17, column=8, value=1)
    ws.cell(row=17, column=12, value=999.0)
    path = _save(wb, tmp_path)

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"excavation": 200.0, "waterproofing": 50.0}


def test_the_totals_column_is_the_one_under_the_total_cost_heading(tmp_path):
    # Both headings have a "Всего" beneath them; picking the unit-price one
    # would report 1 rouble per section.
    path = _save(_offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
    ]), tmp_path)

    assert estimate_sections.read_section_totals(path) == {"excavation": 200.0}


def test_sections_the_report_has_no_line_for_are_left_out(tmp_path):
    path = _save(_offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
        (2, 2, "2. Содержание вертолётной площадки", "Вертолёты", 50.0),
    ]), tmp_path)

    assert estimate_sections.read_section_totals(path) == {"excavation": 200.0}


def test_read_sections_with_warnings_reports_the_unmatched_section_name(tmp_path):
    path = _save(_offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
        (2, 2, "2. Содержание вертолётной площадки", "Вертолёты", 50.0),
    ]), tmp_path)

    sections, unmatched = estimate_sections.read_sections_with_warnings(path)

    assert [s.key for s in sections] == ["excavation"]
    assert unmatched == ["2. Содержание вертолётной площадки"]


def test_read_sections_with_warnings_is_empty_when_everything_matches(tmp_path):
    path = _save(_offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
    ]), tmp_path)

    sections, unmatched = estimate_sections.read_sections_with_warnings(path)

    assert unmatched == []


def test_a_workbook_that_is_not_an_offer_yields_nothing(tmp_path):
    wb = Workbook()
    wb.active.append(["Раздел", "Сумма"])
    wb.active.append(["Фундамент", 1000])
    path = _save(wb, tmp_path)

    assert estimate_sections.read_section_totals(path) == {}


def test_an_unreadable_file_is_reported_rather_than_crashing(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not a workbook at all")

    try:
        estimate_sections.read_section_totals(path)
    except estimate_sections.EstimateSectionsError:
        pass
    else:
        raise AssertionError("нечитаемый файл должен быть отклонён понятной ошибкой")


# --- смета, написанная уровнями («укрупнённая смета») ---

def _levels_estimate(rows, *, header_row=3):
    """A workbook shaped like the other kind of estimate: a column pair per
    level of nesting instead of one column of section numbers.

    ``rows`` are ``(level, number, name, total)`` with level 1, 2 or 3.
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=5, value="Приложение № 2 к Договору")
    ws.cell(row=header_row, column=2, value="номер 1")
    ws.cell(row=header_row, column=3, value="уровень 1")
    ws.cell(row=header_row, column=4, value="номер 2")
    ws.cell(row=header_row, column=5, value="уровень 2")
    ws.cell(row=header_row, column=6, value="номер 3")
    ws.cell(row=header_row, column=7, value="уровень 3")
    ws.cell(row=header_row, column=8, value="Ед. изм.")
    ws.cell(row=header_row, column=13, value="Всего, \nруб. \nс учетом НДС")

    for offset, (level, number, name, total) in enumerate(rows):
        row = header_row + 1 + offset
        number_col = {1: 2, 2: 4, 3: 6}[level]
        ws.cell(row=row, column=number_col, value=number)
        ws.cell(row=row, column=number_col + 1, value=name)
        if total is not None:
            ws.cell(row=row, column=13, value=total)
    return wb


def test_a_levels_estimate_sums_the_parts_of_each_section(tmp_path):
    # A section's own cell is empty and the money sits underneath it.
    path = _save(_levels_estimate([
        (1, 1, "Подготовительные работы, содержание площадки", None),
        (2, "1.1.", "Мобилизация площадки", 100.0),
        (2, "1.2.", "Содержание площадки", 200.0),
        (1, 2, "Котлован", None),
        (2, "2.1.", "Разработка грунта", 50.0),
    ]), tmp_path, "levels.xlsx")

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"preparation": 300.0, "excavation": 50.0}


def test_a_levels_estimate_goes_a_level_deeper_when_it_has_to(tmp_path):
    # The sub-section is empty in its turn and its parts carry the money;
    # counting both would double the section.
    path = _save(_levels_estimate([
        (1, 4, "Конструктивные решения", None),
        (2, "4.1.", "Подземная часть", None),
        (3, "4.1.1.", "Фундаментная плита", 70.0),
        (3, "4.1.2.", "Горизонтальные конструкции", 30.0),
        (2, "4.2.", "Надземная часть", 400.0),
        (3, "4.2.1.", "Эта часть уже посчитана выше", 999.0),
    ]), tmp_path, "levels.xlsx")

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"concrete": 500.0}


def test_a_levels_section_that_states_its_own_total_is_taken_at_its_word(tmp_path):
    path = _save(_levels_estimate([
        (1, 15, "Разработка рабочей документации", 278.0),
    ]), tmp_path, "levels.xlsx")

    assert estimate_sections.read_section_totals(path) == {"rd": 278.0}


def test_two_levels_sections_on_one_report_line_are_added_up(tmp_path):
    # "ВИС - механические системы" and "ВИС - электрические системы" are both
    # the report's utilities line.
    path = _save(_levels_estimate([
        (1, 10, "ВИС - механические системы", None),
        (2, "10.1.", "Отопление", 100.0),
        (1, 11, "ВИС - Электрические и слаботочные системы", None),
        (2, "11.1.", "Освещение", 200.0),
    ]), tmp_path, "levels.xlsx")

    assert estimate_sections.read_section_totals(path) == {"utilities": 300.0}


def test_the_closing_итого_row_of_a_levels_estimate_is_not_counted(tmp_path):
    # Jois's own estimate ends its levels with exactly this row: not a
    # section, but the estimate's own grand total. Classifying it would
    # either drop it as unmatched (harmless but noisy) or, worse, count the
    # whole estimate's total again as if it were a section of its own.
    path = _save(_levels_estimate([
        (1, 2, "Котлован", None),
        (2, "2.1.", "Разработка грунта", 50.0),
        (1, None, "ИТОГО, руб. с учетом НДС", 50.0),
    ]), tmp_path, "levels.xlsx")

    totals = estimate_sections.read_section_totals(path)

    assert totals == {"excavation": 50.0}


# --- формула без сохранённого значения ---

def test_an_uncached_formula_in_a_section_total_is_reported_not_silently_zero(tmp_path):
    # A total written as a formula (as a real offer's own subtotal rows are)
    # carries no cached result until Excel has actually opened and saved the
    # file — openpyxl reads that the same way it reads a genuinely empty
    # cell. The two must not be treated alike: one means "nothing here", the
    # other means "the file was never actually recalculated".
    wb = _offer([
        (1, 1, "1. Котлован", "Котлован", 200.0),
        (2, 2, "2. Кровля", "Кровля", "=100+100"),
    ])
    path = _save(wb, tmp_path, "uncached.xlsx")

    try:
        estimate_sections.read_section_totals(path)
    except estimate_sections.FormulaWithoutCacheError:
        pass
    else:
        raise AssertionError(
            "формула без кэша должна быть замечена, а не прочитана как пустая ячейка"
        )


# --- разбор шапки в разных написаниях ---

def test_the_totals_column_is_found_even_when_its_heading_is_not_merged(tmp_path):
    # One offer leaves "Стоимость всего" unmerged; insisting on the merge left
    # it without a totals column and so without any sections at all.
    wb = _offer([(1, 1, "1. Котлован", "Котлован", 200.0)], header_row=1)
    ws = wb.active
    ws.unmerge_cells(start_row=1, start_column=9, end_row=1, end_column=12)
    path = _save(wb, tmp_path, "unmerged.xlsx")

    assert estimate_sections.read_section_totals(path) == {"excavation": 200.0}


def test_the_article_column_is_recognised_however_it_is_headed(tmp_path):
    wb = _offer([(1, 1, "1. Котлован", "Котлован", 200.0)])
    wb.active.cell(row=9, column=3, value="Справочник статей СМР")
    path = _save(wb, tmp_path, "spravochnik.xlsx")

    assert estimate_sections.read_section_totals(path) == {"excavation": 200.0}


def test_a_section_numbered_rather_than_named_is_read_from_the_works_column(tmp_path):
    # One offer numbers its articles ("1", "1.1") and puts the wording in the
    # works column instead.
    path = _save(_offer([
        (1, 1, "1", "Устройство котлована", 200.0),
    ]), tmp_path, "numbered.xlsx")

    assert estimate_sections.read_section_totals(path) == {"excavation": 200.0}


def test_the_wordings_of_four_real_estimates_are_all_recognised():
    cases = {
        "Возведение несущих конструкций здания (22-341-П-КР)": "concrete",
        "Отделка паркинга, технических помещений, МОП, двери, ворота": "finishing",
        "ВИС - механические системы": "utilities",
        "ВИС - Электрические и слаботочные системы": "utilities",
        "отделка квартир": "mr_base",
        "ПРОЕКТИРОВАНИЕ": "rd",
        "Устройство фасадов": "facade",
        "Устройство кровли": "roof",
        "лифты, подъемники": "lifts",
        "Общестроительные работы - перегородки и стены": "partitions",
    }
    for name, expected in cases.items():
        assert estimate_sections.classify(name) == expected, name


# --- read_concrete_volume ---

def _offer_with_quantity(rows, *, header_row=9, qty_col=10, unit_col=7):
    """A workbook shaped like a real offer that also carries a "Предлагаемое
    количество" column — the header spans two rows, as it does for real, with
    the quantity heading on the second and the unit-of-measure on the first.

    ``rows`` are ``(section_no, article, works_name, qty, unit)`` tuples.
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=header_row, column=1, value="№ п/п")
    ws.cell(row=header_row, column=2, value="№ раздела")
    ws.cell(row=header_row, column=3, value="Статья СМР")
    ws.cell(row=header_row, column=4, value="Наименование работ")
    ws.cell(row=header_row, column=unit_col, value="Ед. изм")
    ws.cell(row=header_row + 1, column=qty_col, value="Предлагаемое количество")
    ws.cell(row=header_row, column=12, value="Стоимость всего")
    ws.cell(row=header_row + 1, column=12, value="Всего")

    for offset, (section, article, works, qty, unit) in enumerate(rows):
        row = header_row + 2 + offset
        ws.cell(row=row, column=1, value=offset + 1)
        ws.cell(row=row, column=2, value=section)
        ws.cell(row=row, column=3, value=article)
        ws.cell(row=row, column=4, value=works)
        ws.cell(row=row, column=unit_col, value=unit)
        ws.cell(row=row, column=qty_col, value=qty)
        ws.cell(row=row, column=12, value=1)
    return wb


def test_an_uncached_formula_in_a_quantity_cell_is_reported(tmp_path):
    path = _save(_offer_with_quantity([
        (4, "4. Конструктивные решения", "Возведение несущих конструкций здания", None, "м3"),
        (None, None, "Фундаментная плита", "=50+50", "м3"),
    ]), tmp_path)

    try:
        estimate_sections.read_concrete_volume(path)
    except estimate_sections.FormulaWithoutCacheError:
        pass
    else:
        raise AssertionError(
            "формула без кэша в колонке количества должна быть замечена"
        )


def test_concrete_volume_sums_the_leaf_quantities_under_the_section(tmp_path):
    # The section's own row and its sub-section headers never carry a
    # quantity in a real offer — only the priced line items underneath do.
    path = _save(_offer_with_quantity([
        (4, "4. Конструктивные решения", "Возведение несущих конструкций здания", None, "м3"),
        ("4.1", "4.1. Подземная часть", "Ж/Б конструкции подземной части", None, "м3"),
        (None, None, "Фундаментная плита", 100.0, "м3"),
        (None, None, "Плиты перекрытия", 50.5, "м3"),
        (5, "5. Общестроительные работы", "Перегородки и стены", None, "м3"),
        (None, None, "Перегородка типовая", 999.0, "м3"),
    ]), tmp_path)

    assert estimate_sections.read_concrete_volume(path) == 150.5


def test_concrete_volume_ignores_lines_measured_in_other_units(tmp_path):
    # Metalwork sometimes sits inside the same section, priced by the tonne
    # rather than the cubic metre — it must not inflate a concrete volume.
    path = _save(_offer_with_quantity([
        (4, "4. Конструктивные решения", "Возведение несущих конструкций здания", None, "м3"),
        (None, None, "Монолит", 200.0, "м3"),
        (None, None, "Металлоконструкции", 10.0, "т"),
    ]), tmp_path)

    assert estimate_sections.read_concrete_volume(path) == 200.0


def test_concrete_volume_is_none_without_a_matching_section(tmp_path):
    path = _save(_offer_with_quantity([
        (1, "1. Котлован", "Устройство котлована", 50.0, "м3"),
    ]), tmp_path)

    assert estimate_sections.read_concrete_volume(path) is None


def test_concrete_volume_is_none_without_a_quantity_column(tmp_path):
    # The plain offer fixture has no "Предлагаемое количество" heading at all.
    path = _save(_offer([
        (1, 4, "4. Конструктивные решения", "Возведение несущих конструкций здания", 200.0),
    ]), tmp_path)

    assert estimate_sections.read_concrete_volume(path) is None


def test_concrete_volume_of_an_unreadable_file_is_reported_rather_than_crashing(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not a workbook at all")

    try:
        estimate_sections.read_concrete_volume(path)
    except estimate_sections.EstimateSectionsError:
        pass
    else:
        raise AssertionError("нечитаемый файл должен быть отклонён понятной ошибкой")


# --- read_concrete_cost_breakdown ---

def test_concrete_cost_breakdown_excludes_the_steelwork_position(tmp_path):
    # Classifier position "4" is the concrete section's own row, stating
    # its Материалы/СМР split whole; "4.3" inside it prices steelwork, not
    # concrete, and must come back out rather than be counted with it.
    wb = _offer([
        (1, "4", "4", "Возведение несущих конструкций здания", 1000.0),
    ])
    ws = wb.active
    ws.cell(row=11, column=9, value=700.0)
    ws.cell(row=11, column=10, value=300.0)
    ws.cell(row=12, column=3, value="4.3")
    ws.cell(row=12, column=4, value="Металлические конструкции")
    ws.cell(row=12, column=9, value=50.0)
    ws.cell(row=12, column=10, value=20.0)
    ws.cell(row=12, column=12, value=70.0)
    path = _save(wb, tmp_path)

    materials, works = estimate_sections.read_concrete_cost_breakdown(path)

    assert materials == Decimal("650.0")
    assert works == Decimal("280.0")


def test_concrete_cost_breakdown_stays_whole_without_a_steelwork_position(tmp_path):
    wb = _offer([
        (1, "4", "4", "Возведение несущих конструкций здания", 1000.0),
    ])
    ws = wb.active
    ws.cell(row=11, column=9, value=700.0)
    ws.cell(row=11, column=10, value=300.0)
    path = _save(wb, tmp_path)

    materials, works = estimate_sections.read_concrete_cost_breakdown(path)

    assert materials == Decimal("700.0")
    assert works == Decimal("300.0")


def test_concrete_cost_breakdown_reads_the_code_with_its_name_attached(tmp_path):
    # The classifier column doesn't always carry a bare code — one estimate
    # writes "4. Конструктивные решения" there instead of a plain "4".
    wb = _offer([
        (1, "4", "4. Конструктивные решения", "Возведение несущих конструкций здания", 1000.0),
    ])
    ws = wb.active
    ws.cell(row=11, column=9, value=700.0)
    ws.cell(row=11, column=10, value=300.0)
    ws.cell(row=12, column=3, value="4.3 Металлические конструкции")
    ws.cell(row=12, column=9, value=50.0)
    ws.cell(row=12, column=10, value=20.0)
    path = _save(wb, tmp_path)

    materials, works = estimate_sections.read_concrete_cost_breakdown(path)

    assert materials == Decimal("650.0")
    assert works == Decimal("280.0")


def test_concrete_cost_breakdown_is_none_without_the_classifier_position(tmp_path):
    wb = _offer([
        (1, 1, "1", "Устройство котлована", 200.0),
    ])
    ws = wb.active
    ws.cell(row=11, column=9, value=120.0)
    ws.cell(row=11, column=10, value=80.0)
    path = _save(wb, tmp_path)

    assert estimate_sections.read_concrete_cost_breakdown(path) == (None, None)


def test_concrete_cost_breakdown_is_none_for_a_levels_estimate(tmp_path):
    # No "Статья СМР" classifier column exists in this shape of estimate.
    path = _save(_levels_estimate([
        (1, 4, "Возведение несущих конструкций здания", 1000.0),
    ]), tmp_path, "levels.xlsx")

    assert estimate_sections.read_concrete_cost_breakdown(path) == (None, None)


# --- read_facade_area ---

def test_facade_area_sums_leaf_rows_whose_article_names_the_cladding_layer(tmp_path):
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Устройство фасадов", None, "м2"),
        (None, "Устройство облицовки фасада", "Панель фасадная типовая", 5000.0, "м2"),
        (None, "Устройство облицовки фасада", "Панель угловая", 1200.5, "м2"),
        (7, "7. Кровля", "Устройство кровли", None, "м2"),
        (None, "Устройство облицовки фасада", "Панель кровельная", 999.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 6200.5


def test_facade_area_recognises_facade_named_with_any_of_its_wordings(tmp_path):
    # "фасад", "фасадные", "фасадов" — the classify() rule matches the
    # substring, so any of these must be found the same way.
    for name in ("Фасадные работы", "Устройство фасадов", "Фасад здания"):
        path = _save(_offer_with_quantity([
            (6, "6. " + name, name, None, "м2"),
            (None, "Устройство облицовки", "Панель фасадная", 100.0, "м2"),
        ]), tmp_path, name="facade.xlsx")
        assert estimate_sections.read_facade_area(path) == 100.0, name


def test_facade_area_ignores_rows_whose_article_names_a_different_layer(tmp_path):
    # A ventilated facade quotes the same square metres three times over —
    # substructure, insulation, and the cladding that finishes it — and the
    # classifier tells the three apart even though all three are м².
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Устройство навесного фасада", None, "м2"),
        (None, "Устройство подсистемы фасада", "Подсистема", 1000.0, "м2"),
        (None, "Устройство утеплителя, ветро и влагозащита", "Утеплитель", 1000.0, "м2"),
        (None, "Устройство облицовки фасада", "Панель из стеклофибробетона", 1000.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 1000.0


def test_facade_area_ignores_lines_measured_in_other_units(tmp_path):
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Устройство фасадов", None, "м2"),
        (None, "Устройство облицовки", "Панель навесная", 300.0, "м2"),
        (None, "Устройство облицовки", "Кронштейны", 50.0, "шт"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 300.0


def test_facade_area_prefers_a_rollup_row_over_its_own_leaves(tmp_path):
    # Where the classifier is filled on the свод, it already states the
    # figure whole — adding the leaf breakdown underneath it on top would
    # double it.
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Устройство фасадов", None, "м2"),
        ("6.3", "Устройство облицовки фасада", "Навесной фасад", 950.0, "м2"),
        (None, None, "Кассеты алюминиевые", 300.0, "м2"),
        (None, None, "Плиты керамогранита", 650.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 950.0


def test_facade_area_sums_leaves_when_no_rollup_states_its_own_figure(tmp_path):
    # The real case this was written against (159-ТУ): the classifier
    # column is filled on only four rows out of 109 in the section, none of
    # them a свод — and those four already give exactly the area the
    # estimate's own facade total agrees with.
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Устройство фасадов", None, "м2"),
        ("6.1", None, "Светопрозрачные конструкции", None, "м2"),
        (None, "Устройство заполнения", "Стеклопакет", 500.0, "м2"),
        (None, None, "Стеклянные консольные козырьки", 20.0, "м2"),
        ("6.3", None, "Навесной фасад", None, "м2"),
        (None, "Устройство подсистемы фасада", "Подсистема", 300.0, "м2"),
        (None, "Устройство облицовки фасада", "Кассеты алюминиевые", 300.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 800.0


def test_facade_area_recognises_light_transmitting_infill(tmp_path):
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Светопрозрачные конструкции", None, "м2"),
        (None, "Устройство заполнения", "Стеклопакет", 500.0, "м2"),
        (None, "Устройство профильной системы", "Импост", 500.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 500.0


def test_facade_area_recognises_modular_infill(tmp_path):
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Модульный фасад", None, "м2"),
        (None, "Профильная система / заполнение модуля", "Модуль", 400.0, "м2"),
        (None, "Несущие кронштейны", "Кронштейн", 400.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 400.0


def test_facade_area_recognises_wet_facade_plaster(tmp_path):
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Мокрый фасад", None, "м2"),
        (None, "Декоративная штукатурка по сетке", "Штукатурка", 250.0, "м2"),
        (None, "Утеплитель", "Плиты минваты", 250.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 250.0


def test_facade_area_recognises_restoration_plaster(tmp_path):
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Реставрация фасада", None, "м2"),
        (None, "Воссоздание штукатурного фасада", "Штукатурка", 180.0, "м2"),
        (None, "Демонтаж старой штукатурки", "Демонтаж", 180.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 180.0


def test_facade_area_matches_a_bare_numeric_classifier_code(tmp_path):
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Устройство фасадов", None, "м2"),
        (None, "6.3.3", "Панель фасадная", 220.0, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) == 220.0


def test_facade_area_is_none_when_the_article_column_names_no_layer(tmp_path):
    # Two estimates in five never fill "Статья СМР" at all on the rows that
    # carry an area — guessing a layer from the free-text works name is
    # exactly the behaviour this replaced, so it stays unset rather than
    # falling back to a name that means something different in every
    # estimate. The manual facade-area field on the passport is for this.
    path = _save(_offer_with_quantity([
        (6, "6. Фасады", "Устройство фасадов", None, "м2"),
        (None, None, "Панель фасадная типовая", 5000.0, "м2"),
        (None, None, "Панель угловая", 1200.5, "м2"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) is None


def test_facade_area_is_none_without_a_matching_section(tmp_path):
    path = _save(_offer_with_quantity([
        (1, "1. Котлован", "Устройство котлована", 50.0, "м3"),
    ]), tmp_path)

    assert estimate_sections.read_facade_area(path) is None


def test_facade_area_of_an_unreadable_file_is_reported_rather_than_crashing(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not a workbook at all")

    try:
        estimate_sections.read_facade_area(path)
    except estimate_sections.EstimateSectionsError:
        pass
    else:
        raise AssertionError("нечитаемый файл должен быть отклонён понятной ошибкой")


# --- количество в укрупнённой смете -----------------------------------------

def _levels_estimate_with_quantity(rows, *, header_row=3):
    """A levels estimate that also carries a "количество" column of its own
    — a real one calls it that rather than "предлагаемое количество", which
    is the offer's own wording, not this shape's.

    ``rows`` are ``(level, number, name, qty, unit)`` with level 1, 2 or 3.
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=5, value="Приложение № 2 к Договору")
    ws.cell(row=header_row, column=2, value="номер 1")
    ws.cell(row=header_row, column=3, value="уровень 1")
    ws.cell(row=header_row, column=4, value="номер 2")
    ws.cell(row=header_row, column=5, value="уровень 2")
    ws.cell(row=header_row, column=6, value="номер 3")
    ws.cell(row=header_row, column=7, value="уровень 3")
    ws.cell(row=header_row, column=8, value="Ед. изм.")
    ws.cell(row=header_row, column=9, value="количество")
    ws.cell(row=header_row, column=13, value="Всего, \nруб. \nс учетом НДС")

    for offset, (level, number, name, qty, unit) in enumerate(rows):
        row = header_row + 1 + offset
        number_col = {1: 2, 2: 4, 3: 6}[level]
        ws.cell(row=row, column=number_col, value=number)
        ws.cell(row=row, column=number_col + 1, value=name)
        if qty is not None:
            ws.cell(row=row, column=9, value=qty)
        if unit is not None:
            ws.cell(row=row, column=8, value=unit)
    return wb


def test_concrete_volume_sums_leaf_quantities_in_a_levels_estimate(tmp_path):
    # Jois' own estimate: "Конструктивные решения" written by levels, not as
    # an offer — the section header and sub-section rows carry no quantity
    # of their own, only the leaf lines under them do.
    path = _save(_levels_estimate_with_quantity([
        (1, 4, "Конструктивные решения", None, None),
        (2, "4.1.", "Подземная часть. Конструктивные решения", None, None),
        (3, "4.1.1.", "Устройство фундаментной плиты", 10014.17, "м3"),
        (3, "4.1.2.", "Горизонтальные конструкции подземной части", 3387.74, "м3"),
        (2, "4.2.", "Надземная часть. Конструктивные решения", None, None),
        (3, "4.2.1.", "Горизонтальные конструкции надземной части", 25666.78, "м3"),
        (1, 5, "Общестроительные работы", None, None),
        (3, "5.1.", "Перегородка типовая", 999.0, "м3"),
    ]), tmp_path, "levels_qty.xlsx")

    assert estimate_sections.read_concrete_volume(path) == 10014.17 + 3387.74 + 25666.78


def test_concrete_volume_in_a_levels_estimate_ignores_other_units(tmp_path):
    # Metalwork sits in the same section, priced by the tonne rather than
    # the cubic metre — it must not inflate a concrete volume here either.
    path = _save(_levels_estimate_with_quantity([
        (1, 4, "Конструктивные решения", None, None),
        (3, "4.1.", "Монолит", 200.0, "м3"),
        (3, "4.2.", "Металлоконструкции", 10.0, "т"),
    ]), tmp_path, "levels_qty.xlsx")

    assert estimate_sections.read_concrete_volume(path) == 200.0


def test_concrete_volume_in_a_levels_estimate_is_none_without_a_matching_section(tmp_path):
    path = _save(_levels_estimate_with_quantity([
        (1, 1, "Котлован", None, None),
        (3, "1.1.", "Разработка грунта", 50.0, "м3"),
    ]), tmp_path, "levels_qty.xlsx")

    assert estimate_sections.read_concrete_volume(path) is None


def test_facade_area_sums_leaf_quantities_in_a_levels_estimate(tmp_path):
    path = _save(_levels_estimate_with_quantity([
        (1, 6, "Фасады", None, None),
        (3, "6.1.", "Панель фасадная типовая", 5000.0, "м2"),
        (3, "6.2.", "Панель угловая", 1200.5, "м2"),
    ]), tmp_path, "levels_qty.xlsx")

    assert estimate_sections.read_facade_area(path) == 6200.5


# --- workbook_cache reuse ----------------------------------------------------

def test_reading_concrete_then_facade_parses_the_file_only_once(tmp_path, monkeypatch):
    # A project page reads this same file for its concrete volume, facade
    # area and section costs — each used to open and parse it again from
    # scratch, which was most of what made the page slow.
    workbook_cache.clear()
    path = _save(_offer_with_quantity([
        (4, "4. Конструктивные решения", "Возведение несущих конструкций здания", None, "м3"),
        (None, None, "Монолит", 200.0, "м3"),
        (6, "6. Фасады", "Устройство фасадов", None, "м2"),
        (None, "Устройство облицовки", "Панель фасадная", 100.0, "м2"),
    ]), tmp_path)

    calls = []
    real_load = openpyxl.load_workbook

    def counting_load(*args, **kwargs):
        calls.append(kwargs.get("data_only"))
        return real_load(*args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", counting_load)

    assert estimate_sections.read_concrete_volume(path) == 200.0
    assert estimate_sections.read_facade_area(path) == 100.0
    assert estimate_sections.read_sections(path)

    assert calls.count(True) == 1
    assert calls.count(False) == 1


def test_concrete_volume_in_a_levels_estimate_takes_a_sections_own_total_over_its_parts(tmp_path):
    # A "укрупнённая смета" section can state its own quantity directly, the
    # same way it can state its own cost (see _sections_from_levels). Summing
    # the sub-sections underneath it on top of that overcounts — on the real
    # sample this was written against, 159 342 m3 instead of the 39 835 m3
    # the section states on its own row.
    path = _save(_levels_estimate_with_quantity([
        (1, 4, "Конструктивные решения", 39835.0, "м3"),
        (2, "4.1.", "Подземная часть. Конструктивные решения", None, None),
        (3, "4.1.1.", "Устройство фундаментной плиты", 60000.0, "м3"),
    ]), tmp_path, "levels_qty.xlsx")

    assert estimate_sections.read_concrete_volume(path) == 39835.0


def test_concrete_volume_in_a_levels_estimate_takes_a_subsections_own_total_over_its_leaves(tmp_path):
    # Same rule one level down: a sub-section that states its own quantity is
    # taken at its word, not added to the leaves listed underneath it.
    path = _save(_levels_estimate_with_quantity([
        (1, 4, "Конструктивные решения", None, None),
        (2, "4.1.", "Подземная часть. Конструктивные решения", 100.0, "м3"),
        (3, "4.1.1.", "Устройство фундаментной плиты", 40.0, "м3"),
        (3, "4.1.2.", "Горизонтальные конструкции подземной части", 40.0, "м3"),
        (2, "4.2.", "Надземная часть. Конструктивные решения", None, None),
        (3, "4.2.1.", "Горизонтальные конструкции надземной части", 50.0, "м3"),
    ]), tmp_path, "levels_qty.xlsx")

    assert estimate_sections.read_concrete_volume(path) == 100.0 + 50.0
