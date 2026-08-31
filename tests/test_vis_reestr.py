import io
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import Workbook

from app import vis_reestr

# 1-based column numbers matching the 0-based COL_* indices in vis_reestr.py.
COL = {"type": 2, "object": 3, "date": 6, "initial": 9, "agreed": 12, "status": 16}


def _workbook(rows, *, sheet_name="Общая сводная", header_row=5):
    """A workbook shaped like the real claims registry: a blank first
    column, a header row somewhere in the first few rows with "Тип" in its
    own column, and one data row per entry after it."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(row=header_row, column=COL["type"], value="Тип запроса")
    ws.cell(row=header_row, column=COL["object"], value="Объект")
    ws.cell(row=header_row, column=COL["date"], value="Дата запроса от ГП")
    ws.cell(row=header_row, column=COL["initial"], value="Начальная сумма")
    ws.cell(row=header_row, column=COL["agreed"], value="Согласовано")
    ws.cell(row=header_row, column=COL["status"], value="Текущий статус")

    for offset, entry in enumerate(rows):
        row = header_row + 1 + offset
        ws.cell(row=row, column=COL["type"], value=entry.get("type"))
        ws.cell(row=row, column=COL["object"], value=entry.get("object"))
        ws.cell(row=row, column=COL["date"], value=entry.get("date"))
        ws.cell(row=row, column=COL["initial"], value=entry.get("initial"))
        ws.cell(row=row, column=COL["agreed"], value=entry.get("agreed"))
        ws.cell(row=row, column=COL["status"], value=entry.get("status"))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- parse_records ---

def test_parse_records_reads_rows_after_the_header():
    wb = _workbook([
        {"type": "удорожание", "object": "Nicole 1", "date": date(2024, 3, 1),
         "initial": 1000, "agreed": 900, "status": "в работе"},
        {"type": "претензия", "object": "Mira", "date": date(2024, 4, 1),
         "initial": 500, "agreed": 500, "status": "завершено"},
    ])
    records = vis_reestr.parse_records(wb)
    assert [r["object"] for r in records] == ["Nicole 1", "Mira"]
    assert records[0]["type"] == "УДОРОЖАНИЕ"
    assert records[0]["date"] == date(2024, 3, 1)
    assert records[0]["initial_sum"] == Decimal("1000")
    assert records[0]["agreed_sum"] == Decimal("900")
    assert records[0]["status"] == "В РАБОТЕ"


def test_parse_records_skips_rows_without_an_object():
    wb = _workbook([
        {"type": "удорожание", "object": None, "agreed": 100},
        {"type": "удорожание", "object": "Nicole 1", "agreed": 200},
    ])
    records = vis_reestr.parse_records(wb)
    assert len(records) == 1
    assert records[0]["object"] == "Nicole 1"


def test_parse_records_finds_header_row_wherever_it_is():
    wb = _workbook([{"object": "Nicole 1", "agreed": 100}], header_row=9)
    records = vis_reestr.parse_records(wb)
    assert len(records) == 1


def test_parse_records_missing_sheet_raises():
    wb = _workbook([{"object": "Nicole 1"}], sheet_name="Другой лист")
    with pytest.raises(vis_reestr.VisReestrError, match="Общая сводная"):
        vis_reestr.parse_records(wb)


def test_parse_records_excel_serial_date():
    wb = _workbook([{"object": "Nicole 1", "date": 45352, "agreed": 100}])  # 2024-03-01
    records = vis_reestr.parse_records(wb)
    assert records[0]["date"] == date(2024, 3, 1)


def test_parse_records_string_amount_with_thousands_separator():
    wb = _workbook([{"object": "Nicole 1", "agreed": "1 234,56"}])
    records = vis_reestr.parse_records(wb)
    assert records[0]["agreed_sum"] == Decimal("1234.56")


def test_parse_records_missing_amount_is_zero():
    wb = _workbook([{"object": "Nicole 1"}])
    records = vis_reestr.parse_records(wb)
    assert records[0]["initial_sum"] == Decimal("0")
    assert records[0]["agreed_sum"] == Decimal("0")


# --- build_analytics ---

def _records(*entries):
    """Plain records, as ``parse_records`` would return them — for testing
    ``build_analytics`` without needing a workbook."""
    defaults = {"type": "", "object": "", "date": None,
                "initial_sum": Decimal("0"), "agreed_sum": Decimal("0"), "status": ""}
    return [{**defaults, **e} for e in entries]


def test_build_analytics_kpi_totals():
    records = _records(
        {"object": "A", "initial_sum": Decimal("100"), "agreed_sum": Decimal("90"), "status": "ЗАВЕРШЕНО"},
        {"object": "B", "initial_sum": Decimal("50"), "agreed_sum": Decimal("50"), "status": "В РАБОТЕ"},
    )
    result = vis_reestr.build_analytics(records)
    assert result["kpi"]["total_requests"] == 2
    assert result["kpi"]["completed_count"] == 1
    assert result["kpi"]["total_initial_sum"] == Decimal("150")
    assert result["kpi"]["total_agreed_sum"] == Decimal("140")


def test_build_analytics_requests_by_object_sorted_descending():
    records = _records(
        {"object": "A"}, {"object": "A"}, {"object": "B"},
    )
    result = vis_reestr.build_analytics(records)
    assert result["requests_by_object"] == [
        {"name": "A", "count": 2}, {"name": "B", "count": 1},
    ]


def test_build_analytics_price_increase_only_counts_udorozhanie_type():
    records = _records(
        {"object": "A", "type": "УДОРОЖАНИЕ", "agreed_sum": Decimal("300")},
        {"object": "A", "type": "ПРЕТЕНЗИЯ", "agreed_sum": Decimal("999")},
        {"object": "B", "type": "УДОРОЖАНИЕ", "agreed_sum": Decimal("100")},
    )
    result = vis_reestr.build_analytics(records)
    assert result["price_increase"] == [
        {"name": "A", "sum": Decimal("300")}, {"name": "B", "sum": Decimal("100")},
    ]


def test_build_analytics_by_type_and_by_status():
    records = _records(
        {"object": "A", "type": "УДОРОЖАНИЕ", "status": "В РАБОТЕ"},
        {"object": "B", "type": "УДОРОЖАНИЕ", "status": "ЗАВЕРШЕНО"},
        {"object": "C", "type": "ПРЕТЕНЗИЯ", "status": "ЗАВЕРШЕНО"},
    )
    result = vis_reestr.build_analytics(records)
    assert {"type": "УДОРОЖАНИЕ", "count": 2} not in result["by_type"]  # shape uses "name", not "type"
    assert {"name": "УДОРОЖАНИЕ", "count": 2} in result["by_type"]
    assert {"name": "ЗАВЕРШЕНО", "count": 2} in result["by_status"]


def test_build_analytics_date_filter_excludes_undated_rows_when_range_set():
    records = _records(
        {"object": "A", "date": date(2024, 1, 15)},
        {"object": "B", "date": None},
    )
    result = vis_reestr.build_analytics(records, date_from=date(2024, 1, 1))
    assert result["meta"]["filtered_records"] == 1
    assert result["requests_by_object"] == [{"name": "A", "count": 1}]


def test_build_analytics_date_filter_includes_undated_rows_without_a_range():
    records = _records({"object": "A", "date": date(2024, 1, 15)}, {"object": "B", "date": None})
    result = vis_reestr.build_analytics(records)
    assert result["meta"]["filtered_records"] == 2


def test_build_analytics_date_filter_bounds_are_inclusive():
    records = _records(
        {"object": "A", "date": date(2024, 1, 1)},
        {"object": "B", "date": date(2024, 1, 31)},
        {"object": "C", "date": date(2024, 2, 1)},
    )
    result = vis_reestr.build_analytics(records, date_from=date(2024, 1, 1), date_to=date(2024, 1, 31))
    assert result["meta"]["filtered_records"] == 2


# --- finalize_* ---

def test_finalize_price_increase_adds_rank_and_share():
    rows = [{"name": "A", "sum": Decimal("300")}, {"name": "B", "sum": Decimal("100")}]
    result = vis_reestr.finalize_price_increase(rows)
    assert result[0]["rank"] == 1
    assert result[0]["share_pct"] == 75.0
    assert result[1]["rank"] == 2
    assert result[1]["share_pct"] == 25.0
    assert result[0]["width_pct"] == 100.0
    assert result[1]["width_pct"] == pytest.approx(33.3, abs=0.1)


def test_finalize_price_increase_empty_list():
    assert vis_reestr.finalize_price_increase([]) == []


def test_finalize_breakdown_share_pct_handles_zero_total():
    assert vis_reestr.finalize_breakdown([]) == []


# --- format_money_short ---

@pytest.mark.parametrize("value, expected", [
    (Decimal("24254622064.69"), "24.25 млрд ₽"),
    (Decimal("3500000"), "3.50 млн ₽"),
    (Decimal("12000"), "12.00 тыс ₽"),
    (Decimal("500"), "500 ₽"),
    (None, "—"),
])
def test_format_money_short(value, expected):
    assert vis_reestr.format_money_short(value) == expected
