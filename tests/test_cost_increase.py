import io

import pytest
from openpyxl import Workbook

from app import cost_increase


def _workbook(rows, *, header_row=1, total=None, name_header=None):
    """A workbook shaped like a real cost-increase file.

    ``rows`` are ``(name, was, now)``. The layout copies the real one closely,
    because most of what this reader has to get right is layout: the kind of
    work is named in a column with no heading of its own, the row number sits
    to its left, a ledger code sits to the right of the money, and the header
    carries a "разница" the reader is meant to ignore in favour of "стало".
    """
    wb = Workbook()
    ws = wb.active
    if name_header:
        ws.cell(row=header_row, column=2, value=name_header)
    ws.cell(row=header_row, column=3, value="было")
    ws.cell(row=header_row, column=4, value="стало")
    ws.cell(row=header_row, column=6, value="разница")

    for offset, (name, was, now) in enumerate(rows):
        row = header_row + 1 + offset
        ws.cell(row=row, column=1, value=offset + 1)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=was)
        ws.cell(row=row, column=4, value=now)
        # A wrong number in "разница", so a reader that trusts the column
        # instead of computing the change is caught rather than accidentally right.
        ws.cell(row=row, column=6, value=-1)
        ws.cell(row=row, column=8, value=5400 + offset)

    if total is not None:
        row = header_row + 2 + len(rows)
        ws.cell(row=row, column=2, value="СУММА")
        ws.cell(row=row, column=3, value=total[0])
        ws.cell(row=row, column=4, value=total[1])
    return wb


def _report(rows, *, estimate=None, **kwargs):
    buf = io.BytesIO()
    _workbook(rows, **kwargs).save(buf)
    buf.seek(0)
    return cost_increase.read_report(buf, estimate)


def _by_key(report):
    return {row.key: row for row in report.rows}


# --- reading the file ------------------------------------------------------

def test_reads_the_sections_of_a_real_file():
    report = _report([
        ("Подготовительные работы, содержание площадки", 1292416447.63, 1292416447.63),
        ("Котлован", 31873390.29, 31873390.29),
        ("Устройство гидроизоляции подземной части здания", 110099464.84, 123439548.48),
        ("Конструктивные решения", 2161902037.57, 2161902037.57),
        ("Общестроительные работы (без отделки)", 211684442.85, 211684442.85),
        ("Фасадные работы", 3084698944.75, 3149227927.59),
        ("Кровля", 145926597.99, 145926597.99),
        ("Отделочные работы (паркинг, надземная часть МОП)", 768338598.15, 768338598.15),
        ("Лифты, подъемники", 223666666.68, 223666666.68),
        ("Благоустройство, дороги", 0, 207987892.97),
        ("ТХ", 18259577.24, 18259577.24),
        ("Рабочая документация", 195565436.58, 195565436.58),
    ])
    rows = _by_key(report)

    assert report.unmatched == []
    assert set(rows) == {
        "preparation", "excavation", "waterproofing", "concrete", "partitions",
        "facade", "roof", "finishing", "lifts", "landscaping", "technology", "rd",
    }
    assert float(rows["waterproofing"].delta) == pytest.approx(13340083.64)
    assert rows["waterproofing"].percent == pytest.approx(12.1, abs=0.05)


def test_the_was_and_now_totals_sum_exactly_not_with_a_floats_rounding_drift():
    # 100000.10 + 200000.20 + 300000.05 drifts to 600000.3500000001 summed
    # as float — read as Decimal (see cost_increase._amount), it doesn't.
    from decimal import Decimal
    report = _report([
        ("Котлован 1", 100000.10, 100000.10),
        ("Котлован 2", 200000.20, 200000.20),
        ("Котлован 3", 300000.05, 300000.05),
    ])
    row = _by_key(report)["excavation"]

    assert row.was == Decimal("600000.35")
    assert isinstance(row.was, Decimal)


def test_the_two_vis_rows_become_one_line_of_engineering_systems():
    # Смета держит инженерные системы одним разделом, а файл удорожания делит
    # их на две строки «ВИС». Раз в отчёте строка одна, обе должны сложиться в
    # неё — иначе половина удорожания инженерии просто не видна.
    report = _report([
        ("ВИС - механические системы", 1120700487.64, 1130700487.64),
        ("ВИС - Электрические и слаботочные системы", 819317798.55, 819317798.55),
    ])

    assert [row.key for row in report.rows] == ["utilities"]
    row = report.rows[0]
    assert float(row.was) == pytest.approx(1940018286.19)
    assert float(row.now) == pytest.approx(1950018286.19)
    assert row.sources == [
        "ВИС - механические системы",
        "ВИС - Электрические и слаботочные системы",
    ]


def test_the_kind_of_work_is_found_in_a_column_with_no_heading():
    # В настоящем файле у колонки с названиями работ заголовка нет вовсе:
    # шапка подписывает только деньги.
    report = _report([("Кровля", 100.0, 110.0)])

    assert [row.key for row in report.rows] == ["roof"]


def test_the_kind_of_work_is_found_when_the_column_is_titled():
    report = _report([("Кровля", 100.0, 110.0)], name_header="Вид работ")

    assert [row.key for row in report.rows] == ["roof"]


def test_the_files_own_total_row_is_not_a_kind_of_work():
    report = _report([("Кровля", 100.0, 110.0)], total=(100.0, 110.0))

    assert [row.key for row in report.rows] == ["roof"]


def test_the_total_is_the_sum_of_the_lines_shown():
    # Итог считается по строкам таблицы, а не берётся из «СУММЫ» файла: иначе
    # строка, которую программа не разобрала, пряталась бы в итоге, и сойтись
    # он мог бы только случайно.
    report = _report(
        [("Кровля", 100.0, 110.0), ("Фасадные работы", 200.0, 260.0)],
        total=(999.0, 999.0),
    )

    assert report.total.was == pytest.approx(300.0)
    assert report.total.now == pytest.approx(370.0)
    assert report.total.delta == pytest.approx(70.0)
    assert report.total.percent == pytest.approx(70 / 300 * 100)


def test_lines_come_in_the_reports_order_not_the_files():
    report = _report([
        ("Благоустройство, дороги", 100.0, 100.0),
        ("Котлован", 100.0, 100.0),
        ("Кровля", 100.0, 100.0),
    ])

    assert [row.key for row in report.rows] == ["excavation", "roof", "landscaping"]


def test_a_section_neither_the_file_nor_the_estimate_prices_is_absent_not_zero():
    report = _report([("Кровля", 100.0, 110.0)])

    assert [row.key for row in report.rows] == ["roof"]


def test_rows_of_the_file_with_no_line_in_the_report_are_named_not_dropped_silently():
    report = _report([
        ("Кровля", 100.0, 110.0),
        ("Аренда вертолётной площадки", 50.0, 60.0),
    ])

    assert [row.key for row in report.rows] == ["roof"]
    assert report.unmatched == ["Аренда вертолётной площадки"]
    # И в итог такая строка тоже не попадает — иначе итог не сошёлся бы с
    # таблицей, которую человек видит.
    assert report.total.now == pytest.approx(110.0)


def test_a_row_with_no_numbers_at_all_is_skipped():
    report = _report([("Кровля", None, None), ("Фасадные работы", 100.0, 110.0)])

    assert [row.key for row in report.rows] == ["facade"]


def test_numbers_written_as_text_are_read():
    report = _report([("Кровля", "1 000,50", "1 100,50")])

    assert report.rows[0].delta == pytest.approx(100.0)


# --- the percentage --------------------------------------------------------

def test_work_that_did_not_move_is_zero_per_cent():
    report = _report([("Кровля", 100.0, 100.0)])

    assert report.rows[0].percent == 0.0
    assert cost_increase.format_percent(report.rows[0].percent) == "0 %"


def test_work_that_became_cheaper_gets_a_negative_percentage():
    report = _report([("Кровля", 100.0, 90.0)])

    assert report.rows[0].percent == pytest.approx(-10.0)
    assert report.rows[0].delta == pytest.approx(-10.0)


def test_work_that_was_not_in_the_estimate_has_no_percentage():
    # Благоустройство в настоящем файле начинается с нуля: доля, на которую
    # выросло ничто, — не число, и «+100%» здесь было бы утверждением о
    # работах, а не об арифметике.
    report = _report([("Благоустройство, дороги", 0, 207987892.97)])

    assert report.rows[0].percent is None
    assert float(report.rows[0].delta) == pytest.approx(207987892.97)
    assert cost_increase.format_percent(None) is None


# --- удорожание считается от сметы ------------------------------------------

def test_the_increase_is_measured_from_the_estimate_not_from_was():
    # Главное правило: база — смета, а «стало» — то, сколько работы стоят
    # теперь. «Было» здесь ни при чём: файл накопительный, и его «было» может
    # быть любым промежуточным значением.
    report = _report([("Кровля", 110.0, 130.0)], estimate={"roof": 100.0})
    row = report.rows[0]

    assert row.baseline == 100.0
    assert row.current == 130.0
    assert row.delta == pytest.approx(30.0)
    assert row.percent == pytest.approx(30.0)
    assert row.source == cost_increase.FROM_NOW


def test_stalo_below_the_estimate_is_a_saving_not_an_increase():
    report = _report([("Кровля", 100.0, 90.0)], estimate={"roof": 100.0})

    assert report.rows[0].delta == pytest.approx(-10.0)
    assert report.rows[0].percent == pytest.approx(-10.0)


def test_an_empty_stalo_falls_back_to_was():
    # Пустое «стало» — не нулевая стоимость, а «эту строку не пересчитывали»:
    # последняя цифра, которая у файла есть, стоит в «было».
    report = _report([("Кровля", 130.0, 0)], estimate={"roof": 100.0})
    row = report.rows[0]

    assert row.current == 130.0
    assert row.source == cost_increase.FROM_WAS
    assert row.delta == pytest.approx(30.0)


def test_a_whole_empty_stalo_column_falls_back_to_was_line_by_line():
    report = _report(
        [("Кровля", 130.0, 0), ("Фасадные работы", 250.0, None)],
        estimate={"roof": 100.0, "facade": 200.0},
    )

    assert [row.source for row in report.rows] == [
        cost_increase.FROM_WAS, cost_increase.FROM_WAS,
    ]
    assert report.total.delta == pytest.approx(80.0)


def test_a_line_both_columns_are_silent_about_is_left_at_the_estimate():
    # Ни «стало», ни «было»: файл про эти работы ничего не говорит. Это не то же
    # самое, что «работы убрали», — иначе вся стоимость раздела прочиталась бы
    # как экономия, которой никто не делал.
    report = _report([("Кровля", 0, 0)], estimate={"roof": 100.0})
    row = report.rows[0]

    assert row.current == 100.0
    assert row.delta == 0.0
    assert row.source == cost_increase.FROM_NOTHING


def test_a_section_of_the_estimate_the_file_leaves_out_still_gets_a_line():
    # Таблица показывает удорожание сметы, значит в ней должны быть все разделы
    # сметы. Раздел, которого в файле нет, стоит без изменений и помечен, а не
    # пропадает из таблицы и из итога.
    report = _report([("Кровля", 100.0, 110.0)], estimate={"roof": 100.0, "facade": 200.0})
    rows = _by_key(report)

    assert set(rows) == {"roof", "facade"}
    assert rows["facade"].current == 200.0
    assert rows["facade"].delta == 0.0
    assert rows["facade"].source == cost_increase.FROM_NOTHING
    assert rows["facade"].sources == []
    assert report.total.delta == pytest.approx(10.0)


def test_work_the_estimate_never_priced_is_increase_to_the_last_rouble():
    report = _report(
        [("Кровля", 100.0, 110.0), ("Благоустройство, дороги", 0, 207987892.97)],
        estimate={"roof": 100.0},
    )
    rows = _by_key(report)

    assert rows["landscaping"].estimate is None
    assert rows["landscaping"].baseline == 0.0
    assert float(rows["landscaping"].delta) == pytest.approx(207987892.97)
    # Процента нет: доля, на которую выросло ничто, — не число.
    assert rows["landscaping"].percent is None


def test_the_total_is_stalo_against_the_estimate():
    report = _report(
        [("Кровля", 110.0, 130.0), ("Фасадные работы", 190.0, 260.0)],
        estimate={"roof": 100.0, "facade": 200.0},
    )

    assert report.from_estimate is True
    assert report.total.baseline == pytest.approx(300.0)
    assert report.total.current == pytest.approx(390.0)
    assert report.total.delta == pytest.approx(90.0)
    assert report.total.percent == pytest.approx(30.0)


def test_without_an_estimate_the_baseline_is_the_files_own_was():
    report = _report([("Кровля", 100.0, 110.0)])

    assert report.from_estimate is False
    assert report.rows[0].estimate is None
    assert report.rows[0].baseline == 100.0
    assert report.rows[0].delta == pytest.approx(10.0)


def test_a_line_that_is_zero_everywhere_is_not_shown_at_all():
    # Ноль по смете, ноль в файле: строка не сообщает ничего, а в таблице
    # занимает место среди настоящих цифр.
    report = _report([("MR Base", 0, 0), ("Кровля", 100.0, 110.0)])

    assert [row.key for row in report.rows] == ["roof"]


def test_an_estimate_of_nothing_leaves_the_percentage_unsaid():
    # Делить на ноль нечем: процент не существует и должен остаться пустым, а
    # не превратиться в подпись «None» на странице.
    report = _report([("Кровля", 100.0, 110.0)], estimate={"roof": 0.0})

    assert report.rows[0].baseline == 0.0
    assert report.rows[0].delta == pytest.approx(110.0)
    assert report.rows[0].percent is None
    assert report.total.percent is None


# --- what a broken file does -----------------------------------------------

def test_a_file_without_the_two_money_columns_is_refused():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Раздел")
    ws.cell(row=1, column=2, value="Сумма")
    ws.cell(row=2, column=1, value="Кровля")
    ws.cell(row=2, column=2, value=100)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    with pytest.raises(cost_increase.CostIncreaseError):
        cost_increase.read_lines(buf)


def test_a_file_that_is_not_a_workbook_at_all_is_refused():
    with pytest.raises(cost_increase.CostIncreaseError):
        cost_increase.read_lines(io.BytesIO(b"this is not a spreadsheet"))


def test_a_header_further_down_the_sheet_is_still_found():
    report = _report([("Кровля", 100.0, 110.0)], header_row=4)

    assert [row.key for row in report.rows] == ["roof"]


def test_read_lines_works_on_an_open_file_so_an_upload_can_be_checked_before_saving():
    buf = io.BytesIO()
    _workbook([("Кровля", 100.0, 110.0)]).save(buf)
    buf.seek(0)

    lines = cost_increase.read_lines(buf)

    assert [line.name for line in lines] == ["Кровля"]


def test_an_uncached_formula_in_stalo_is_reported_not_silently_dropped():
    # "Кровля" would otherwise vanish from the report entirely (was is also
    # None, so the "was is None and now is None" check would skip the row)
    # — read as a real formula, not the same as an empty cell.
    buf = io.BytesIO()
    _workbook([("Кровля", None, "=100+30")]).save(buf)
    buf.seek(0)

    with pytest.raises(cost_increase.FormulaWithoutCacheError):
        cost_increase.read_lines(buf)


def test_reads_from_a_path_as_well(tmp_path):
    path = tmp_path / "udorozhanie.xlsx"
    _workbook([("Кровля", 100.0, 110.0)]).save(path)

    assert [line.name for line in cost_increase.read_lines(path)] == ["Кровля"]


# --- how the numbers are written -------------------------------------------

def test_a_percentage_is_written_the_way_the_comparison_writes_one():
    assert cost_increase.format_percent(12.05) == "+12,1 %"
    assert cost_increase.format_percent(-3.2) == "−3,2 %"
    assert cost_increase.format_percent(0.0) == "0 %"
    # Округляется до тех же десятых, что и показывается: «+0,04%» на экране
    # выглядело бы как «+0,0 %» — то есть как отсутствие изменения.
    assert cost_increase.format_percent(0.04) == "0 %"


def test_the_change_in_roubles_keeps_its_sign():
    assert cost_increase.format_delta(13340083.64) == "+13 340 083.64"
    assert cost_increase.format_delta(-500.0) == "−500"
    assert cost_increase.format_delta(0.0) == "0"
    assert cost_increase.format_delta(None) is None
